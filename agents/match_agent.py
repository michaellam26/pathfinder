"""
Match Agent — runs after every resume update (most frequent).

Responsibilities:
  1. Load resume text from PROFILE_DIR (default: ./profile/)
  2. Read all AI-TPM JDs from JD_Tracker sheet
  3. Two-stage evaluation:
       Stage 1 — Batch coarse scoring (10 JDs per Gemini call), write results
       Stage 2 — Fine evaluation of top 20% coarse-scored JDs (1 JD per call)
  4. Invalidate stale scores when resume changes (MD5 hash detection)
  5. Write results to Match_Results sheet

Run:
  python agents/match_agent.py

Profile dir can be overridden:
  PROFILE_DIR=/path/to/folder python agents/match_agent.py
"""
import os
import sys
import re
import json
import math
import hashlib
import asyncio
import time
import logging
from dotenv import load_dotenv
from google import genai
from google.genai import types

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from shared.excel_store import (
    EXCEL_PATH, PROJECT_ROOT, get_or_create_excel,
    get_jd_rows_for_match, get_match_pairs, upsert_match_record,
    batch_upsert_match_records, MATCH_HEADERS,
)
from shared.prompts import get_prompt_pair, HM_PROMPTS, TRACKS
from shared.schemas import CoarseItem, BatchCoarseResult, MatchResult

# ── JD Markdown cache ──────────────────────────────────────────────────────────
# JD_CACHE_DIR sourced from shared.config (P1-5).

def _load_jd_markdown(url: str) -> str | None:
    """Load JD Markdown from cache. Prefer _structured.md, fallback to raw .md."""
    md5_hex = hashlib.md5(url.encode()).hexdigest()
    structured = os.path.join(JD_CACHE_DIR, f"{md5_hex}_structured.md")
    if os.path.exists(structured):
        with open(structured, encoding="utf-8") as f:
            return f.read()
    raw = os.path.join(JD_CACHE_DIR, f"{md5_hex}.md")
    if os.path.exists(raw):
        with open(raw, encoding="utf-8") as f:
            return f.read()
    return None

from shared.gemini_pool import _GeminiKeyPoolBase
from shared.rate_limiter import _RateLimiter
from shared.config import MODEL, JD_CACHE_DIR
from shared.exceptions import GeminiTransientError, GeminiStructuralError
from shared.run_summary import RunSummary
from shared.ats_matcher import compute_coverage
from shared.resume_io import load_resume

# PRJ-002: Coverage % below this gets a ⚠️ marker in the printed summary.
# Soft signal only — JDs are NOT dropped or excluded from fine eval.
ATS_COVERAGE_LOW_THRESHOLD = 30.0

# BUG-41: single shared limiter across Stage 1 and Stage 2
_GEMINI_LIMITER = _RateLimiter(rpm=13)


# BUG-31: use _GeminiKeyPoolBase directly with genai_mod parameter
_GeminiKeyPool = _GeminiKeyPoolBase  # alias for backward compat (tests)

load_dotenv()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(levelname)s] - %(message)s')

PROFILE_DIR = os.getenv("PROFILE_DIR", os.path.join(PROJECT_ROOT, "profile"))

# Stage 2 fine-eval gating. JDs enter fine eval if EITHER:
#   • coarse score >= MATCH_FINE_SCORE_THRESHOLD, or
#   • coarse score is in the top MATCH_FINE_TOP_PERCENT% of this run.
# Union semantics protect against both flat-high distributions (where
# top-N% would discard genuine fits) and flat-low ones (where the
# absolute threshold would select nothing).
FINE_SCORE_THRESHOLD = int(os.getenv("MATCH_FINE_SCORE_THRESHOLD", "60"))
FINE_TOP_PERCENT     = float(os.getenv("MATCH_FINE_TOP_PERCENT", "60"))


_KEY_POOL: "_GeminiKeyPool | None" = None  # initialised in main()

# P0-1 / PRJ-004 REQ-004-22: optional Gemini Context Caches, one per track.
# A cache binds its system_instruction, and each track has its own HM prompt,
# so one cache cannot serve five tracks. main() creates caches lazily for
# tracks with ≥2 fine-eval candidates; evaluate_match looks up its track's
# cache and falls back transparently to the uncached path on a miss.
_FINE_CACHE_NAMES: dict = {}


# ── Pydantic schemas: see shared/schemas.py ───────────────────────────────────


# ── Resume loader ─────────────────────────────────────────────────────────────
# load_resume now lives in shared/resume_io.py and supports .md/.txt/.pdf
# (PDFs are converted via pdfplumber and cached under profile/.cache/).


# ── Stage 1: batch coarse scoring ────────────────────────────────────────────
def _format_jd_for_coarse(jd: dict) -> str:
    """Format JD for coarse scoring using Requirements, Additional Qualifications,
    and Responsibilities fields from structured Excel data."""
    try:
        d = json.loads(jd["jd_json"])
        parts = [
            f"Job Title: {d.get('job_title', 'N/A')}",
            f"Company: {d.get('company', 'N/A')}",
            f"Location: {d.get('location', 'N/A')}",
        ]
        if d.get("requirements"):
            parts.append("Requirements:\n" + "\n".join(f"- {r}" for r in d["requirements"]))
        if d.get("additional_qualifications"):
            parts.append("Additional Qualifications:\n" + "\n".join(f"- {q}" for q in d["additional_qualifications"]))
        if d.get("key_responsibilities"):
            parts.append("Responsibilities:\n" + "\n".join(f"- {r}" for r in d["key_responsibilities"]))
        return "\n\n".join(parts)
    except Exception:
        return jd["jd_json"]


def _extract_ats_keywords(jd: dict) -> list[str]:
    """Extract ats_keywords from a JD row's cached JSON. Empty list on legacy
    JDs (cached before PR 1 added the field) or malformed JSON."""
    try:
        d = json.loads(jd["jd_json"])
        kws = d.get("ats_keywords") or []
        return [str(k) for k in kws if k]
    except Exception as e:
        # Phase 4 review: surface the failure so a malformed jd_json doesn't
        # silently degrade ATS coverage to 0 across many JDs unnoticed.
        logging.debug(f"_extract_ats_keywords failed for {jd.get('url', '?')}: {e}")
        return []


def compute_ats_for_jds(resume_text: str, jds: list[dict]) -> dict[str, dict]:
    """Compute ATS coverage for each JD. Returns {url: coverage_dict}.

    Deterministic, no LLM, no rate limiter. Skips JDs whose ats_keywords
    field is empty (legacy JDs); coverage_dict has percent=None for those.

    Phase 4 perf note: compute_coverage internally normalizes the resume
    once per call. For N JDs we'd re-normalize the same resume N times.
    The matcher API doesn't expose a pre-normalized form, but normalize()
    is O(text length) and ~100 JDs × small resume is still milliseconds —
    not worth complicating the matcher API for.
    """
    out: dict[str, dict] = {}
    for jd in jds:
        kws = _extract_ats_keywords(jd)
        out[jd["url"]] = compute_coverage(kws, resume_text)
    return out


def _track_batches(jds: list, size: int = 10) -> list:
    """Partition JDs by job_domain, then chunk each group by `size`.
    Returns [(track, [jd, ...]), ...] — a batch never mixes tracks because a
    Gemini batch call shares one system prompt (PRJ-004 REQ-004-22)."""
    by_track: dict = {}
    for jd in jds:
        by_track.setdefault(jd.get("job_domain", "AI"), []).append(jd)
    return [(track, group[i * size : (i + 1) * size])
            for track, group in by_track.items()
            for i in range(math.ceil(len(group) / size))]


def batch_coarse_score(resume_text: str, jds_batch: list[dict],
                       job_domain: str = "AI") -> list[int]:
    """Send resume + up to 10 JDs in one Gemini call; return list of int scores.

    PRJ-004 REQ-004-22: a batch shares one system prompt, so every JD in
    jds_batch must belong to the same track — callers partition by
    job_domain before chunking.

    P0-4: returns [] (empty list) on a structural error so the caller can
    drop the affected JDs without writing fake score=1 records. Transient
    errors propagate as GeminiTransientError — the caller should let them
    bubble up to main() so the run fails loudly.
    """
    if _KEY_POOL is None:
        raise RuntimeError("_KEY_POOL not initialized — call main() first or set _KEY_POOL before invoking batch_coarse_score()")
    numbered = "\n\n".join(
        f"[JD {i}]\n{_format_jd_for_coarse(jd)}" for i, jd in enumerate(jds_batch)
    )
    cfg = types.GenerateContentConfig(
        system_instruction=get_prompt_pair(job_domain)[0],
        temperature=0.0,
        response_mime_type="application/json",
        response_schema=BatchCoarseResult,
    )
    try:
        resp = _KEY_POOL.generate_content(
            model=MODEL,
            contents=(
                "--- CANDIDATE PROFILE ---\n"
                f"<scraped_content>\n{resume_text}\n</scraped_content>\n\n"
                "--- JDs TO SCORE ---\n"
                f"<scraped_content>\n{numbered}\n</scraped_content>\n\n"
                "Return BatchCoarseResult JSON with one CoarseItem per JD."
            ),
            config=cfg,
        )
        result = BatchCoarseResult.model_validate_json(resp.text)
        scores = [1] * len(jds_batch)
        for item in result.items:
            if 0 <= item.index < len(jds_batch):
                scores[item.index] = max(1, item.score)
        return scores
    except GeminiTransientError:
        raise
    except GeminiStructuralError as e:
        logging.error(f"Batch coarse score structural failure (skipping {len(jds_batch)} JDs): {e}")
        return []
    except Exception as e:
        # JSON / Pydantic validation on the response text — also structural.
        logging.error(f"Batch coarse score response parse failed (skipping {len(jds_batch)} JDs): {e}")
        return []


# ── Stage 2: fine match evaluation ───────────────────────────────────────────
def evaluate_match(resume_text: str, jd_json: str,
                   job_domain: str = "AI") -> str | None:
    """Run Stage 2 fine evaluation. Returns Gemini JSON string on success.

    PRJ-004 REQ-004-22/23: the HM prompt is selected by the JD's track via
    get_prompt_pair — the identical constant resume_optimizer.re_score uses,
    so before/after deltas stay comparable per track (REQ-052).

    P0-1: when the track has an entry in _FINE_CACHE_NAMES, uses Gemini
    Context Caching — the resume + that track's HM prompt are referenced via
    cached_content so only the JD is sent fresh per call (~30-40% input-token
    savings). P0-4: returns None on structural error so the caller can drop
    the record. Transient errors propagate as GeminiTransientError.
    """
    if _KEY_POOL is None:
        raise RuntimeError("_KEY_POOL not initialized — call main() first or set _KEY_POOL before invoking evaluate_match()")
    cache_name = _FINE_CACHE_NAMES.get(job_domain)
    if cache_name:
        cfg = types.GenerateContentConfig(
            cached_content=cache_name,
            temperature=0.0,
            response_mime_type="application/json",
            response_schema=MatchResult,
        )
        contents = (
            "--- TARGET JD ---\n"
            f"<scraped_content>\n{jd_json}\n</scraped_content>\n\n"
            "Provide MatchResult JSON."
        )
    else:
        cfg = types.GenerateContentConfig(
            system_instruction=get_prompt_pair(job_domain)[1],
            temperature=0.0,
            response_mime_type="application/json",
            response_schema=MatchResult,
        )
        contents = (
            "--- CANDIDATE PROFILE ---\n"
            f"<scraped_content>\n{resume_text}\n</scraped_content>\n\n"
            "--- TARGET JD ---\n"
            f"<scraped_content>\n{jd_json}\n</scraped_content>\n\n"
            "Provide MatchResult JSON."
        )
    try:
        resp = _KEY_POOL.generate_content(
            model=MODEL,
            contents=contents,
            config=cfg,
        )
        return resp.text
    except GeminiTransientError:
        raise
    except GeminiStructuralError as e:
        logging.error(f"Fine match eval structural failure (record dropped): {e}")
        return None


def _select_fine_candidates(
    scored: dict,
    score_threshold: int,
    top_percent: float,
) -> tuple[list, dict]:
    """Pick which coarse-scored JDs proceed to fine eval.

    Selection is the UNION of two criteria:
      • absolute: score >= score_threshold
      • relative: score in the top `top_percent`% of this run

    Only pairs still at stage="coarse" are returned (already-fine pairs are
    skipped). Returns (to_fine_keys, stats) where stats is a dict with
    keys n, top_count, top_cutoff, threshold_count for printing.
    """
    n = len(scored)
    if n == 0:
        return [], {"n": 0, "top_count": 0, "top_cutoff": 0, "threshold_count": 0}
    sorted_keys = sorted(
        scored.keys(),
        key=lambda k: scored[k]["score"],
        reverse=True,
    )
    top_n      = max(1, math.ceil(n * top_percent / 100))
    top_keys   = set(sorted_keys[:top_n])
    top_cutoff = scored[sorted_keys[top_n - 1]]["score"]
    threshold_keys = {
        k for k, v in scored.items() if v["score"] >= score_threshold
    }
    candidates = top_keys | threshold_keys
    to_fine = [k for k in candidates if scored[k]["stage"] == "coarse"]
    stats = {
        "n": n,
        "top_count": len(top_keys),
        "top_cutoff": top_cutoff,
        "threshold_count": len(threshold_keys),
    }
    return to_fine, stats


# ── Main ──────────────────────────────────────────────────────────────────────
async def main():
    global _KEY_POOL
    summary = RunSummary(agent="match")
    try:
        await _main_inner(summary)
    except GeminiTransientError as e:
        summary.transient_errors += 1
        summary.note(f"Run aborted (transient): {e}")
        raise
    except Exception as e:
        summary.note(f"Run aborted: {type(e).__name__}: {e}")
        raise
    finally:
        # PRJ-004 REQ-004-25/26: token-usage snapshot in every run log —
        # the measurement carrier for the trial-run cost gate.
        try:
            from shared.gemini_pool import get_usage_summary
            summary.note(f"gemini usage: {get_usage_summary()}")
        except Exception:
            pass
        summary.mark_finished()
        log_path = summary.write()
        print(f"📊 Run summary: {log_path}")
        print(summary.to_json())


async def _main_inner(summary: RunSummary):
    global _KEY_POOL, _FINE_CACHE_NAMES
    gemini_keys = [k for k in [
        os.getenv("GEMINI_API_KEY"),
        os.getenv("GEMINI_API_KEY_2"),
    ] if k]
    if not gemini_keys:
        print("❌ Missing GEMINI_API_KEY in .env")
        summary.note("Missing GEMINI_API_KEY")
        return
    _KEY_POOL = _GeminiKeyPoolBase(gemini_keys, genai_mod=genai)
    logging.info(f"[KeyPool] Loaded {len(gemini_keys)} Gemini API key(s).")
    _FINE_CACHE_NAMES = {}  # reset between runs

    print("\n" + "="*60)
    print("MATCH AGENT")
    print("="*60 + "\n")

    resume_text, resume_id = load_resume(PROFILE_DIR)
    if not resume_text:
        print(f"⚠️  No resume found in {PROFILE_DIR}\n"
              f"   Create {PROFILE_DIR}/ and add a .md or .txt resume file.")
        summary.note(f"No resume in {PROFILE_DIR}")
        return

    resume_hash = hashlib.md5(resume_text.encode("utf-8")).hexdigest()
    print(f"📄 Resume: {resume_id}  (MD5: {resume_hash[:8]}...)")

    xlsx_path = get_or_create_excel()
    jds       = get_jd_rows_for_match(xlsx_path)
    if not jds:
        print("⚠️  No domain-qualified JDs in tracker. Run job_agent.py first.")
        summary.note("No JDs in tracker")
        return

    # ── Stage 1: batch coarse scoring ─────────────────────────────────────────
    # Trust job_agent's Job Domain classification — no second-pass keyword filter.
    existing_pairs = get_match_pairs(xlsx_path)

    # Detect stale pairs (resume changed since last score)
    stale_keys = {
        k for k, v in existing_pairs.items()
        if k[0] == resume_id and v["hash"] and v["hash"] != resume_hash
    }
    if stale_keys:
        logging.info(f"[Stage1] {len(stale_keys)} stale pair(s) detected — resume changed, will re-score.")

    already_fine   = {k for k, v in existing_pairs.items()
                      if k[0] == resume_id and v["stage"] == "fine" and k not in stale_keys}
    already_coarse = {k for k, v in existing_pairs.items()
                      if k[0] == resume_id and v["stage"] == "coarse" and k not in stale_keys}

    pending_jds = [
        jd for jd in jds
        if (resume_id, jd["url"]) not in already_fine
        and (resume_id, jd["url"]) not in already_coarse
    ]

    print(f"📋 JDs in tracker (domain-qualified): {len(jds)}")
    print(f"✅ Already fine-scored:      {len(already_fine)}")
    print(f"~  Already coarse-scored:   {len(already_coarse)}")
    print(f"🔍 Needs coarse scoring:     {len(pending_jds)}\n")
    summary.skipped += len(already_fine) + len(already_coarse)
    summary.attempted += len(pending_jds)

    # ── ATS dim (PRJ-002): deterministic keyword coverage, no LLM ─────────────
    # Run on the same set of JDs that go through Recruiter scoring (pending).
    # Already-scored JDs from previous runs keep whatever ATS % was written
    # last time; clearing Match_Results triggers a full re-score with fresh ATS.
    ats_results: dict[str, dict] = compute_ats_for_jds(resume_text, pending_jds)
    if ats_results:
        with_data = [u for u, c in ats_results.items() if c.get("percent") is not None]
        if with_data:
            avg_pct = sum(ats_results[u]["percent"] for u in with_data) / len(with_data)
            low_count = sum(
                1 for u in with_data
                if ats_results[u]["percent"] < ATS_COVERAGE_LOW_THRESHOLD
            )
            print(f"🔍 ATS coverage: {len(with_data)} JD(s) scored, "
                  f"avg {avg_pct:.1f}%, {low_count} below {ATS_COVERAGE_LOW_THRESHOLD:.0f}% ⚠️")
        legacy = len(ats_results) - len(with_data)
        if legacy:
            print(f"   ({legacy} JD(s) lack ats_keywords — re-run job_agent to populate.)")

    # Batch coarse scoring in groups of 10. PRJ-004 REQ-004-22: a batch shares
    # one system prompt, so partition by track first, then chunk within each
    # group — same total call count (±4 partial batches worst case).
    coarse_scores: dict[str, int] = {}
    if pending_jds:
        limiter    = _GEMINI_LIMITER
        pass_jds   = pending_jds
        batches    = _track_batches(pass_jds)

        for b, (track, batch) in enumerate(batches):
            print(f"[Coarse batch {b+1}/{len(batches)}] Scoring {len(batch)} "
                  f"{track} JDs...")
            await limiter.acquire()
            scores = await asyncio.to_thread(batch_coarse_score, resume_text,
                                             batch, track)
            # P0-4: empty list = structural error; do not write fake score=1.
            if not scores:
                logging.warning(
                    f"[Stage1] Batch {b+1} returned no scores (structural error); "
                    f"skipping {len(batch)} JD(s)."
                )
                summary.structural_errors += 1
                summary.failed += len(batch)
                continue
            for jd, score in zip(batch, scores):
                coarse_scores[jd["url"]] = score
            logging.info(f"[Stage1] Batch {b+1} ({track}) scores: {scores}")
        summary.succeeded += len(coarse_scores)

        # Stage 1 records: dict format. Score column mirrors Recruiter Score.
        # ATS / Recruiter / HM populated; HM=None until Stage 2 runs.
        coarse_records = []
        for jd in pass_jds:
            url = jd["url"]
            if url not in coarse_scores:
                continue
            ats = ats_results.get(url, {})
            coarse_records.append({
                "resume_id":  resume_id,
                "jd_url":     url,
                "match_json": json.dumps({
                    "compatibility_score": coarse_scores[url],
                    "key_strengths":       [],
                    "critical_gaps":       [],
                    "recommendation_reason": "Recruiter (Stage 1) score — pending HM evaluation.",
                }),
                "resume_hash":           resume_hash,
                "stage":                 "coarse",
                "ats_coverage_percent":  ats.get("percent"),
                "ats_missing":           ats.get("missing", []),
                "recruiter_score":       coarse_scores[url],
                "hm_score":              None,
            })
        if coarse_records:
            batch_upsert_match_records(xlsx_path, coarse_records)
        logging.info(f"[Stage1] Wrote {len(coarse_records)} coarse records "
                     f"(skipped {len(pass_jds) - len(coarse_records)} due to structural errors).")

    # ── Stage 2: fine evaluation of top 20% coarse-scored JDs ─────────────────
    all_pairs = get_match_pairs(xlsx_path)
    scored_for_resume = {
        k: v for k, v in all_pairs.items()
        if k[0] == resume_id and v["score"] > 0
    }

    if not scored_for_resume:
        print("No scored JDs found; skipping fine evaluation.")
        _print_top_results(xlsx_path, resume_id)
        return

    to_fine, stats = _select_fine_candidates(
        scored_for_resume,
        score_threshold=FINE_SCORE_THRESHOLD,
        top_percent=FINE_TOP_PERCENT,
    )

    print(f"\n[Stage2] Total scored JDs:           {stats['n']}")
    print(f"[Stage2] Threshold (>= {FINE_SCORE_THRESHOLD}):{'':>13}{stats['threshold_count']} JD(s)")
    print(f"[Stage2] Top {FINE_TOP_PERCENT:g}% (cutoff {stats['top_cutoff']}):{'':>5}{stats['top_count']} JD(s)")
    print(f"[Stage2] Union → fine eval queued:   {len(to_fine)}\n")

    if to_fine:
        jd_lookup  = {jd["url"]: jd for jd in jds}
        limiter    = _GEMINI_LIMITER
        sem        = asyncio.Semaphore(3)
        total_fine = len(to_fine)
        fine_records: list[tuple] = []

        # P0-1 / PRJ-004: try Gemini Context Caching per track. A cache binds
        # its system_instruction (one HM prompt), so each track with ≥2 fine
        # candidates gets its own cache of resume + HM_PROMPTS[track]; caching
        # cuts ~30-40% of input tokens. Falls back transparently if the model
        # doesn't support caching (preview models often don't).
        cache_contents = [types.Content(
            role="user",
            parts=[types.Part.from_text(text=(
                "--- CANDIDATE PROFILE ---\n"
                f"<scraped_content>\n{resume_text}\n</scraped_content>"
            ))],
        )] if hasattr(types, "Content") else None
        if cache_contents is not None:
            fine_track_counts: dict[str, int] = {}
            for key in to_fine:
                jd = jd_lookup.get(key[1])
                if jd:
                    t = jd.get("job_domain", "AI")
                    fine_track_counts[t] = fine_track_counts.get(t, 0) + 1
            for track, count in fine_track_counts.items():
                if count < 2 or track not in HM_PROMPTS:
                    continue
                cache_name = _KEY_POOL.create_cache(
                    model=MODEL,
                    system_instruction=HM_PROMPTS[track],
                    contents=cache_contents,
                    ttl="3600s",
                    display_name=f"match-{summary.run_id}-{track}",
                )
                if cache_name:
                    _FINE_CACHE_NAMES[track] = cache_name
            if _FINE_CACHE_NAMES:
                summary.note(f"Using context caches: {_FINE_CACHE_NAMES}")

        async def fine_one(key: tuple, i: int) -> None:
            async with sem:
                url = key[1]
                jd  = jd_lookup.get(url)
                if not jd:
                    return
                print(f"[Fine {i}/{total_fine}] Evaluating: {url}")
                await limiter.acquire()
                jd_content  = _load_jd_markdown(url) or jd["jd_json"]
                result_json = await asyncio.to_thread(
                    evaluate_match, resume_text, jd_content,
                    jd.get("job_domain", "AI"),
                )
                # P0-4: structural error — drop the record, keep coarse score in Excel.
                if result_json is None:
                    print(f"    Structural error; record dropped (coarse score kept).")
                    summary.structural_errors += 1
                    summary.failed += 1
                    return
                try:
                    parsed = json.loads(result_json)
                    score = max(1, parsed.get("compatibility_score", 1))
                    if parsed.get("compatibility_score", 1) < 1:
                        parsed["compatibility_score"] = score
                        result_json = json.dumps(parsed)
                except Exception as e:
                    # JSON parse on Gemini response with response_schema set is
                    # essentially never structural — but if it does happen, drop.
                    logging.error(f"[Fine] Response parse failed for {url}: {e}")
                    summary.structural_errors += 1
                    summary.failed += 1
                    return
                # Dict format with hm_score only — preserves ATS Coverage %
                # and Recruiter Score that Stage 1 wrote (the upsert uses
                # "key absent → leave column unchanged" semantics).
                fine_records.append({
                    "resume_id":   resume_id,
                    "jd_url":      url,
                    "match_json":  result_json,
                    "resume_hash": resume_hash,
                    "stage":       "fine",
                    "hm_score":    score,
                })
                summary.succeeded += 1
                print(f"    Score: {score}/100")

        summary.attempted += len(to_fine)
        try:
            await asyncio.gather(*[fine_one(k, i) for i, k in enumerate(to_fine, 1)])
        finally:
            # P0-1: best-effort cache cleanup; never let teardown crash main.
            for cache_name in list(_FINE_CACHE_NAMES.values()):
                _KEY_POOL.delete_cache(cache_name)
            _FINE_CACHE_NAMES.clear()
        batch_upsert_match_records(xlsx_path, fine_records)
        logging.info(f"[Stage2] Wrote {len(fine_records)} fine records.")

    print(f"\n🎉 Match Agent complete. Results in {xlsx_path}")
    _print_top_results(xlsx_path, resume_id)


def _print_top_results(xlsx_path: str, resume_id: str):
    """Print top-5 matches to console. ★ = fine evaluated, ~ = coarse only."""
    from openpyxl import load_workbook
    # BUG-51: dynamic column lookup instead of hardcoded indices
    col_rid   = MATCH_HEADERS.index("Resume ID") + 1
    col_url   = MATCH_HEADERS.index("JD URL") + 1
    col_score = MATCH_HEADERS.index("Score") + 1
    col_stage = MATCH_HEADERS.index("Stage") + 1
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Match_Results"]
        rows = []
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, col_rid).value
            if rid != resume_id:
                continue
            url   = ws.cell(r, col_url).value or ""
            score = ws.cell(r, col_score).value or 0
            stage = ws.cell(r, col_stage).value or "fine"
            rows.append((score, url, stage))
        rows.sort(reverse=True)
        if not rows:
            return
        print("\n🏆 Top matches:")
        for rank, (score, url, stage) in enumerate(rows[:5], 1):
            indicator = "★" if stage == "fine" else "~"
            print(f"  {rank}. {indicator} [{score}/100] {url}")
        print("  (★ = fine-evaluated  ~ = coarse-only)")
    finally:
        wb.close()


if __name__ == "__main__":
    asyncio.run(main())

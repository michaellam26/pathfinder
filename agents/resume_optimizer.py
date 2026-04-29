"""
Resume Optimizer Agent — the 4th agent in the pipeline.

Responsibilities:
  1. Load all scored matches (score >= 0) from Match_Results
  2. For each match, tailor the resume to the specific JD using Gemini
  3. Re-score the tailored resume using the same fine-evaluation prompt
  4. Save tailored resumes to tailored_resumes/{resume_id}/{url_md5}.md
  5. Write results to Tailored_Match_Results sheet

Run:
  python agents/resume_optimizer.py
"""
import os
import sys
import json
import hashlib
import asyncio
import logging
import math
from dotenv import load_dotenv
from google import genai
from google.genai import types

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from shared.excel_store import (
    EXCEL_PATH, PROJECT_ROOT, TAILORED_HEADERS, get_or_create_excel,
    get_scored_matches, get_tailored_match_pairs, batch_upsert_tailored_records,
    get_jd_rows_for_match,
)
from shared.gemini_pool import _GeminiKeyPoolBase
from shared.rate_limiter import _RateLimiter
from shared.config import MODEL
from shared.exceptions import GeminiTransientError, GeminiStructuralError
from shared.run_summary import RunSummary
from shared.prompts import (
    FINE_SYSTEM_PROMPT,
    TAILOR_SYSTEM_PROMPT, BATCH_TAILOR_SYSTEM_PROMPT,
)
from shared.schemas import (
    MatchResult, TailoredResume,
    BatchTailoredItem, BatchTailoredResult,
)

# BUG-41: single shared limiter for all Gemini calls
_GEMINI_LIMITER = _RateLimiter(rpm=13)

# ── JD cache ──────────────────────────────────────────────────────────────────
JD_CACHE_DIR = os.path.join(PROJECT_ROOT, "jd_cache")
TAILORED_DIR = os.path.join(PROJECT_ROOT, "tailored_resumes")
PROFILE_DIR  = os.getenv("PROFILE_DIR", os.path.join(PROJECT_ROOT, "profile"))


# BUG-31: use _GeminiKeyPoolBase directly with genai_mod parameter
_GeminiKeyPool = _GeminiKeyPoolBase  # alias for backward compat (tests)

load_dotenv()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - [%(levelname)s] - %(message)s')

_KEY_POOL: "_GeminiKeyPool | None" = None


# ── Pydantic schemas: see shared/schemas.py ───────────────────────────────────
# ── Prompts: see shared/prompts.py ────────────────────────────────────────────

BATCH_TAILOR_SIZE = 2   # JDs per batch for tailor (output is large — full Markdown resume each)
RESCORE_CONCURRENCY = 3  # parallel single-JD re-score calls (matches match_agent's fine eval)


# ── Helpers ──────────────────────────────────────────────────────────────────
def load_resume(folder: str) -> tuple:
    """Returns (resume_text, resume_id). resume_id = filename without extension."""
    if not os.path.exists(folder):
        logging.error(f"Profile folder not found: {folder}")
        return "", ""
    files = [f for f in os.listdir(folder)
             if not f.startswith('.') and f.lower().endswith(('.md', '.txt'))]
    if not files:
        logging.error(f"No .md/.txt files in {folder}")
        return "", ""
    fname = files[0]
    with open(os.path.join(folder, fname), encoding="utf-8") as fh:
        text = fh.read()
    resume_id = os.path.splitext(fname)[0]
    logging.info(f"Loaded resume: {fname}  ({len(text)} chars)")
    return text, resume_id


def _load_jd_markdown(url: str) -> str | None:
    """Load JD Markdown from cache. Prefer _structured.md, fallback to raw .md."""
    md5 = hashlib.md5(url.encode()).hexdigest()
    structured = os.path.join(JD_CACHE_DIR, f"{md5}_structured.md")
    if os.path.exists(structured):
        with open(structured, encoding="utf-8") as f:
            return f.read()
    raw = os.path.join(JD_CACHE_DIR, f"{md5}.md")
    if os.path.exists(raw):
        with open(raw, encoding="utf-8") as f:
            return f.read()
    return None


def _save_tailored_resume(resume_id: str, url: str, content: str) -> str:
    """Save tailored resume to tailored_resumes/{resume_id}/{url_md5}.md. Returns path."""
    subdir = os.path.join(TAILORED_DIR, resume_id)
    os.makedirs(subdir, exist_ok=True)
    md5 = hashlib.md5(url.encode()).hexdigest()
    path = os.path.join(subdir, f"{md5}.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    return path


# ── Gemini calls ─────────────────────────────────────────────────────────────
def tailor_resume(resume_text: str, jd_content: str) -> str | None:
    """Call Gemini to tailor resume for a specific JD. Returns JSON string.

    P0-4: returns None on structural error so the caller can drop the
    record. Transient errors propagate as GeminiTransientError.
    """
    if _KEY_POOL is None:
        raise RuntimeError("_KEY_POOL not initialized — call main() first or set _KEY_POOL before invoking tailor_resume()")
    cfg = types.GenerateContentConfig(
        system_instruction=TAILOR_SYSTEM_PROMPT,
        temperature=0.3,
        response_mime_type="application/json",
        response_schema=TailoredResume,
    )
    try:
        resp = _KEY_POOL.generate_content(
            model=MODEL,
            contents=(
                "--- ORIGINAL RESUME ---\n"
                f"<scraped_content>\n{resume_text}\n</scraped_content>\n\n"
                "--- TARGET JD ---\n"
                f"<scraped_content>\n{jd_content}\n</scraped_content>\n\n"
                "Tailor the resume for this specific job. Return TailoredResume JSON."
            ),
            config=cfg,
        )
        return resp.text
    except GeminiTransientError:
        raise
    except GeminiStructuralError as e:
        logging.error(f"Tailor resume structural failure (record dropped): {e}")
        return None


def batch_tailor_resume(resume_text: str, jd_contents: list[str]) -> list[dict]:
    """Batch-tailor resume for multiple JDs in one Gemini call. Returns list[dict] aligned to input."""
    if _KEY_POOL is None:
        raise RuntimeError("_KEY_POOL not initialized — call main() first or set _KEY_POOL before invoking batch_tailor_resume()")
    numbered = "\n\n".join(
        f"[JD {i}]\n{jd}" for i, jd in enumerate(jd_contents)
    )
    cfg = types.GenerateContentConfig(
        system_instruction=BATCH_TAILOR_SYSTEM_PROMPT,
        temperature=0.3,
        response_mime_type="application/json",
        response_schema=BatchTailoredResult,
    )
    try:
        resp = _KEY_POOL.generate_content(
            model=MODEL,
            contents=(
                "--- ORIGINAL RESUME ---\n"
                f"<scraped_content>\n{resume_text}\n</scraped_content>\n\n"
                "--- TARGET JDs ---\n"
                f"<scraped_content>\n{numbered}\n</scraped_content>\n\n"
                "Tailor the resume for each JD. Return BatchTailoredResult JSON."
            ),
            config=cfg,
        )
        result = BatchTailoredResult.model_validate_json(resp.text)
        out = [{} for _ in range(len(jd_contents))]
        for item in result.items:
            if 0 <= item.index < len(jd_contents):
                out[item.index] = {
                    "tailored_resume_markdown": item.tailored_resume_markdown,
                    "optimization_summary": item.optimization_summary,
                }
        return out
    except GeminiTransientError:
        raise
    except GeminiStructuralError as e:
        logging.error(f"Batch tailor structural failure (skipping {len(jd_contents)} jobs): {e}")
        return [{} for _ in range(len(jd_contents))]
    except Exception as e:
        logging.error(f"Batch tailor response parse failed (skipping {len(jd_contents)} jobs): {e}")
        return [{} for _ in range(len(jd_contents))]


def re_score(tailored_resume: str, jd_content: str) -> str | None:
    """Re-score the tailored resume using the same fine-evaluation prompt.

    P0-4: returns None on structural error so the caller can drop the
    record. Transient errors propagate as GeminiTransientError.
    """
    if _KEY_POOL is None:
        raise RuntimeError("_KEY_POOL not initialized — call main() first or set _KEY_POOL before invoking re_score()")
    cfg = types.GenerateContentConfig(
        system_instruction=FINE_SYSTEM_PROMPT,
        temperature=0.0,
        response_mime_type="application/json",
        response_schema=MatchResult,
    )
    try:
        resp = _KEY_POOL.generate_content(
            model=MODEL,
            contents=(
                "--- CANDIDATE PROFILE ---\n"
                f"<scraped_content>\n{tailored_resume}\n</scraped_content>\n\n"
                "--- TARGET JD ---\n"
                f"<scraped_content>\n{jd_content}\n</scraped_content>\n\n"
                "Provide MatchResult JSON."
            ),
            config=cfg,
        )
        return resp.text
    except GeminiTransientError:
        raise
    except GeminiStructuralError as e:
        logging.error(f"Re-score structural failure (record dropped): {e}")
        return None


# ── Main ─────────────────────────────────────────────────────────────────────
async def main():
    summary = RunSummary(agent="optimizer")
    try:
        await _main_inner(summary)
    except GeminiTransientError as e:
        summary.transient_errors += 1
        summary.note(f"Run aborted (transient): {e}")
        raise
    except Exception as e:
        summary.note(f"Run aborted: {type(e).__name__}: {e}")
        raise
    finally:
        summary.mark_finished()
        log_path = summary.write()
        print(f"📊 Run summary: {log_path}")
        print(summary.to_json())


async def _main_inner(summary: RunSummary):
    global _KEY_POOL
    gemini_keys = [k for k in [
        os.getenv("GEMINI_API_KEY"),
        os.getenv("GEMINI_API_KEY_2"),
    ] if k]
    if not gemini_keys:
        print("Missing GEMINI_API_KEY in .env")
        summary.note("Missing GEMINI_API_KEY")
        return
    _KEY_POOL = _GeminiKeyPoolBase(gemini_keys, genai_mod=genai)
    logging.info(f"[KeyPool] Loaded {len(gemini_keys)} Gemini API key(s).")

    print("\n" + "=" * 60)
    print("RESUME OPTIMIZER AGENT")
    print("=" * 60 + "\n")

    # 1. Load resume
    resume_text, resume_id = load_resume(PROFILE_DIR)
    if not resume_text:
        print(f"No resume found in {PROFILE_DIR}")
        return

    resume_hash = hashlib.md5(resume_text.encode("utf-8")).hexdigest()
    print(f"Resume: {resume_id}  (MD5: {resume_hash[:8]}...)")

    # 2. Load scored matches + existing tailored pairs
    # P0-6: stage="fine" is the default but pass explicitly so the intent is
    # visible: only Stage 2 Gemini fine-eval scores feed score-delta math.
    xlsx_path = get_or_create_excel()
    scored = get_scored_matches(xlsx_path, stage="fine")
    if not scored:
        print("No scored matches found. Run match_agent.py first.")
        return

    existing_tailored = get_tailored_match_pairs(xlsx_path)

    # 3. Build JD lookup for title/company
    jd_rows = get_jd_rows_for_match(xlsx_path)
    jd_meta = {}
    for jd in jd_rows:
        try:
            d = json.loads(jd["jd_json"])
            jd_meta[jd["url"]] = {
                "job_title": d.get("job_title", ""),
                "company": d.get("company", ""),
                "jd_json": jd["jd_json"],
            }
        except Exception:
            pass

    # 4. Filter: skip already-optimized pairs with same resume hash
    to_process = []
    for match in scored:
        key = (match["resume_id"], match["jd_url"])
        if key in existing_tailored:
            if existing_tailored[key]["resume_hash"] == resume_hash:
                continue  # already optimized, resume unchanged
        to_process.append(match)

    print(f"Scored matches (score >= 0): {len(scored)}")
    print(f"Already optimized:          {len(scored) - len(to_process)}")
    print(f"To process:                 {len(to_process)}\n")
    summary.skipped += len(scored) - len(to_process)
    summary.attempted += len(to_process)

    if not to_process:
        print("All matches already optimized. Nothing to do.")
        _print_summary(xlsx_path, resume_id)
        return

    # 5. Prepare job items (pre-load JD content)
    limiter = _GEMINI_LIMITER
    total   = len(to_process)

    job_items = []
    for match in to_process:
        url = match["jd_url"]
        meta = jd_meta.get(url, {})
        jd_content = _load_jd_markdown(url) or meta.get("jd_json", "")
        if not jd_content:
            logging.warning(f"No JD content for {url}, skipping.")
            continue
        job_items.append({
            "match": match,
            "url": url,
            "meta": meta,
            "jd_content": jd_content,
        })

    # ── Phase 1: Batch Tailor ──────────────────────────────────────
    n_tailor_batches = math.ceil(len(job_items) / BATCH_TAILOR_SIZE)
    print(f"[Phase 1] Batch tailoring {len(job_items)} jobs in {n_tailor_batches} batches "
          f"(batch_size={BATCH_TAILOR_SIZE})...\n")

    tailor_results: dict[str, dict] = {}   # url -> {"tailored_md", "opt_summary"}
    tailor_fallback: list[dict] = []

    for b in range(n_tailor_batches):
        batch = job_items[b * BATCH_TAILOR_SIZE : (b + 1) * BATCH_TAILOR_SIZE]
        jd_contents = [item["jd_content"] for item in batch]

        labels = ", ".join(
            f"{item['meta'].get('company', '?')}" for item in batch
        )
        print(f"  [Tailor {b+1}/{n_tailor_batches}] {labels}")

        await limiter.acquire()
        batch_results = await asyncio.to_thread(
            batch_tailor_resume, resume_text, jd_contents
        )

        for item, result in zip(batch, batch_results):
            if result and result.get("tailored_resume_markdown"):
                url = item["url"]
                tailor_results[url] = {
                    "tailored_md": result["tailored_resume_markdown"],
                    "opt_summary": result.get("optimization_summary", ""),
                }
                _save_tailored_resume(resume_id, url, result["tailored_resume_markdown"])
            else:
                tailor_fallback.append(item)

    # ── Phase 1.5: Fallback — retry failed items individually ──────
    if tailor_fallback:
        print(f"\n  [Fallback] Retrying {len(tailor_fallback)} failed tailor items individually...")
        for item in tailor_fallback:
            await limiter.acquire()
            tailor_json = await asyncio.to_thread(
                tailor_resume, resume_text, item["jd_content"]
            )
            # P0-4: None = structural error; drop the record (do not write empty MD).
            if tailor_json is None:
                logging.error(f"  Fallback tailor also failed for {item['url']} (structural)")
                summary.structural_errors += 1
                continue
            try:
                data = json.loads(tailor_json)
                md = data.get("tailored_resume_markdown", "")
                if md:
                    url = item["url"]
                    tailor_results[url] = {
                        "tailored_md": md,
                        "opt_summary": data.get("optimization_summary", ""),
                    }
                    _save_tailored_resume(resume_id, url, md)
            except Exception as e:
                logging.error(f"  Fallback tailor parse failed for {item['url']}: {e}")
                summary.structural_errors += 1

    print(f"\n[Phase 1 done] {len(tailor_results)}/{len(job_items)} resumes tailored.")

    # ── Phase 2: Re-score (single JD/call, FINE_SYSTEM_PROMPT) ─────
    # Mirrors agents/match_agent.py:411-448 so original fine score and
    # tailored re-score share the same call shape — Score Delta then
    # reflects the resume change, not batch-context anchoring.
    rescore_items = [item for item in job_items if item["url"] in tailor_results]
    print(f"\n[Phase 2] Re-scoring {len(rescore_items)} tailored resumes "
          f"(1 JD/call, FINE_SYSTEM_PROMPT)...\n")

    score_results: dict[str, dict] = {}   # url -> score data dict
    sem = asyncio.Semaphore(RESCORE_CONCURRENCY)

    async def rescore_one(item: dict, idx: int, total: int) -> None:
        async with sem:
            url = item["url"]
            company = item["meta"].get("company", "?")
            await limiter.acquire()
            score_json = await asyncio.to_thread(
                re_score, tailor_results[url]["tailored_md"], item["jd_content"]
            )
            # P0-4: None = structural error; drop the record (do not write fake score=0).
            if score_json is None:
                logging.error(f"  [Re-score {idx}/{total}] {company} — structural error, dropped")
                summary.structural_errors += 1
                return
            try:
                parsed = json.loads(score_json)
            except Exception as e:
                logging.error(f"  [Re-score {idx}/{total}] {company} — parse failed: {e}; dropped")
                summary.structural_errors += 1
                return
            score_results[url] = parsed
            score = parsed.get("compatibility_score", 0)
            print(f"  [Re-score {idx}/{total}] {company} — {score}/100")

    if rescore_items:
        total = len(rescore_items)
        await asyncio.gather(*[
            rescore_one(it, i, total) for i, it in enumerate(rescore_items, 1)
        ])

    # ── Phase 3: Assemble results + write Excel ────────────────────
    results: list[dict] = []
    for item in job_items:
        url = item["url"]
        if url not in tailor_results or url not in score_results:
            continue
        meta = item["meta"]
        match = item["match"]
        score_data = score_results[url]
        tailored_score = score_data.get("compatibility_score", 0)
        original_score = match["score"]
        delta = tailored_score - original_score

        job_title  = meta.get("job_title", "Unknown")
        company    = meta.get("company", "Unknown")
        regression = delta < 0
        marker     = "⚠️ regressed → keep base" if regression else ""
        print(f"  {company} — {job_title}: {original_score} → {tailored_score} ({delta:+d}) {marker}")

        md5 = hashlib.md5(url.encode()).hexdigest()
        path = os.path.join(TAILORED_DIR, resume_id, f"{md5}.md")

        results.append({
            "resume_id": resume_id,
            "jd_url": url,
            "job_title": job_title,
            "company": company,
            "original_score": original_score,
            "tailored_score": tailored_score,
            "score_delta": delta,
            "tailored_resume_path": path,
            "optimization_summary": tailor_results[url]["opt_summary"],
            "resume_hash": resume_hash,
            "regression": regression,
        })

    if results:
        batch_upsert_tailored_records(xlsx_path, results)
        logging.info(f"Wrote {len(results)} tailored records.")
    summary.succeeded += len(results)
    summary.failed += max(0, len(to_process) - len(results))

    api_calls = n_tailor_batches + len(tailor_fallback) + len(rescore_items)
    print(f"\nResume Optimizer complete. {len(results)}/{total} jobs processed.")
    print(f"API calls: {api_calls} (tailor: {n_tailor_batches} batch + {len(tailor_fallback)} fallback; "
          f"re-score: {len(rescore_items)} single)")
    _print_summary(xlsx_path, resume_id)


def _print_summary(xlsx_path: str, resume_id: str):
    """Print tailored match summary sorted by score delta.

    Also surfaces regressions (tailored < base) prominently — for those
    JDs the user should keep using the base resume in profile/.
    """
    pairs = get_tailored_match_pairs(xlsx_path)
    if not pairs:
        return

    from openpyxl import load_workbook
    # BUG-50: dynamic column lookup instead of hardcoded indices
    col_rid      = TAILORED_HEADERS.index("Resume ID") + 1
    col_company  = TAILORED_HEADERS.index("Company") + 1
    col_title    = TAILORED_HEADERS.index("Job Title") + 1
    col_original = TAILORED_HEADERS.index("Original Score") + 1
    col_tailored = TAILORED_HEADERS.index("Tailored Score") + 1
    col_delta    = TAILORED_HEADERS.index("Score Delta") + 1
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Tailored_Match_Results"]
        rows = []
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, col_rid).value
            if rid != resume_id:
                continue
            company   = ws.cell(r, col_company).value or ""
            job_title = ws.cell(r, col_title).value or ""
            original  = ws.cell(r, col_original).value or 0
            tailored  = ws.cell(r, col_tailored).value or 0
            delta     = ws.cell(r, col_delta).value or 0
            rows.append((delta, company, job_title, original, tailored))
        rows.sort(reverse=True)
        if not rows:
            return
        print(f"\n{'='*70}")
        print(f"{'Company':<20} {'Job Title':<20} {'Orig':>5} {'Tail':>5} {'Delta':>6}")
        print(f"{'-'*70}")
        total_delta = 0
        for delta, company, title, orig, tail in rows[:20]:
            c = (company[:18] + "..") if len(company) > 20 else company
            t = (title[:18] + "..") if len(title) > 20 else title
            print(f"{c:<20} {t:<20} {orig:>5} {tail:>5} {delta:>+6}")
            total_delta += delta
        if rows:
            avg = total_delta / len(rows)
            print(f"{'-'*70}")
            print(f"Average delta: {avg:+.1f}  ({len(rows)} jobs)")

        regressed = [r for r in rows if r[0] < 0]
        if regressed:
            print(f"\n⚠️  Regressed (tailored < base) — keep using base resume from profile/")
            print(f"   {len(regressed)} of {len(rows)} job(s):")
            for delta, company, title, orig, tail in regressed:
                c = (company[:18] + "..") if len(company) > 20 else company
                t = (title[:18] + "..") if len(title) > 20 else title
                print(f"   [{delta:+d}] {c:<20} {t:<20} ({orig} → {tail})")
    finally:
        wb.close()


if __name__ == "__main__":
    asyncio.run(main())

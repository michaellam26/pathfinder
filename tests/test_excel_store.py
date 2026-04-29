"""
Tests for shared/excel_store.py

Coverage:
  - get_or_create_excel: creates file with correct sheets/headers
  - count_company_rows / get_company_rows: reads rows
  - upsert_companies: insert + update
  - get_company_names_without_tpm: reads second sheet
  - update_company_career_url: updates URL column
  - get_jd_urls / get_jd_url_meta: JD read helpers
  - batch_update_jd_timestamps: timestamp update
  - get_jd_rows_for_match: filter by is_ai_tpm
  - upsert_jd_record / batch_upsert_jd_records: JD write
  - get_incomplete_jd_rows: detects missing fields
  - count_tpm_jobs_by_company: job count aggregation
  - update_company_job_counts: write job counts
  - get_match_pairs / upsert_match_record / batch_upsert_match_records
"""
import os
import sys
import json
import tempfile
import unittest
import unittest.mock
from datetime import datetime

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from shared.excel_store import (
    get_or_create_excel, count_company_rows, get_company_rows,
    get_company_rows_with_row_num,
    upsert_companies, get_company_names_without_tpm, update_company_career_url,
    get_jd_urls, get_jd_url_meta, batch_update_jd_timestamps,
    get_jd_rows_for_match, upsert_jd_record, batch_upsert_jd_records,
    get_incomplete_jd_rows, count_tpm_jobs_by_company, update_company_job_counts,
    get_match_pairs, upsert_match_record, batch_upsert_match_records,
    get_scored_matches, get_tailored_match_pairs, batch_upsert_tailored_records,
    get_archived_companies, update_archive_status, unarchive_company,
    get_company_archive_info, count_valid_tpm_jobs_by_company,
    COMPANY_HEADERS, JD_HEADERS, MATCH_HEADERS, TAILORED_HEADERS, _JD_COL,
)
import openpyxl


def _tmp_xlsx() -> str:
    """Return path to a temp xlsx that doesn't exist yet."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    os.remove(path)  # get_or_create_excel will create it
    return path


class TestGetOrCreateExcel(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_creates_file(self):
        get_or_create_excel(self.path)
        self.assertTrue(os.path.exists(self.path))

    def test_creates_required_sheets(self):
        get_or_create_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        for sheet in ["Company_List", "Company_Without_TPM", "JD_Tracker", "Match_Results"]:
            self.assertIn(sheet, wb.sheetnames, f"Missing sheet: {sheet}")

    def test_returns_path(self):
        returned = get_or_create_excel(self.path)
        self.assertEqual(returned, self.path)

    def test_idempotent_second_call(self):
        """Calling twice should not error and return same path."""
        get_or_create_excel(self.path)
        returned = get_or_create_excel(self.path)
        self.assertEqual(returned, self.path)

    def test_company_list_headers(self):
        get_or_create_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for h in ["Company Name", "AI Domain", "Career URL"]:
            self.assertIn(h, headers)

    def test_match_results_headers(self):
        get_or_create_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Match_Results"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for h in ["Resume ID", "JD URL", "Score", "Resume Hash", "Stage"]:
            self.assertIn(h, headers)

    def test_bug06_move_fails_raises_runtimeerror_not_recursion(self):
        """BUG-06 regression: if shutil.move fails during corrupt-file recovery,
        get_or_create_excel must raise RuntimeError instead of infinite recursion."""
        import tempfile
        from unittest.mock import patch

        fd, corrupt_path = tempfile.mkstemp(suffix=".xlsx")
        os.write(fd, b"CORRUPT NOT A ZIP")
        os.close(fd)
        try:
            with patch("shutil.move", side_effect=OSError("disk full")):
                with self.assertRaises(RuntimeError):
                    get_or_create_excel(corrupt_path)
        finally:
            if os.path.exists(corrupt_path):
                os.remove(corrupt_path)

    def test_bug06_move_succeeds_recreates_fresh_file(self):
        """BUG-06: when shutil.move succeeds, recursive call must create a fresh
        valid xlsx (not loop again)."""
        import tempfile

        fd, corrupt_path = tempfile.mkstemp(suffix=".xlsx")
        os.write(fd, b"CORRUPT NOT A ZIP")
        os.close(fd)
        bak = corrupt_path + ".bak"
        try:
            result = get_or_create_excel(corrupt_path)
            self.assertEqual(result, corrupt_path)
            self.assertTrue(os.path.exists(corrupt_path))
            wb = openpyxl.load_workbook(corrupt_path)
            self.assertIn("Company_List", wb.sheetnames)
        finally:
            if os.path.exists(corrupt_path):
                os.remove(corrupt_path)
            if os.path.exists(bak):
                os.remove(bak)


class TestCountAndGetCompanyRows(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_empty_returns_zero(self):
        self.assertEqual(count_company_rows(self.path), 0)

    def test_count_after_insert(self):
        upsert_companies(self.path, [
            {"company_name": "TestCo", "ai_domain": "AI Startups",
             "business_focus": "Focus", "career_url": "https://testco.com/jobs"},
        ])
        self.assertEqual(count_company_rows(self.path), 1)

    def test_get_rows_returns_list(self):
        upsert_companies(self.path, [
            {"company_name": "Alpha", "ai_domain": "AI Startups",
             "business_focus": "B", "career_url": "https://alpha.com/careers"},
        ])
        rows = get_company_rows(self.path)
        self.assertIsInstance(rows, list)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], "Alpha")

    def test_empty_get_rows(self):
        rows = get_company_rows(self.path)
        self.assertEqual(rows, [])


class TestUpsertCompanies(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_insert_two_companies(self):
        upsert_companies(self.path, [
            {"company_name": "A", "ai_domain": "AI Startups",
             "business_focus": "Focus A", "career_url": "https://a.com/jobs"},
            {"company_name": "B", "ai_domain": "Large Model Labs",
             "business_focus": "Focus B", "career_url": "https://b.ai/careers"},
        ])
        self.assertEqual(count_company_rows(self.path), 2)

    def test_update_existing_company(self):
        upsert_companies(self.path, [
            {"company_name": "Corp", "ai_domain": "AI Startups",
             "business_focus": "Old", "career_url": "https://old.com/jobs"},
        ])
        upsert_companies(self.path, [
            {"company_name": "Corp", "ai_domain": "Big Tech (AI Investment)",
             "business_focus": "New focus", "career_url": "https://new.com/careers"},
        ])
        # Should still be 1 row (updated, not duplicated)
        self.assertEqual(count_company_rows(self.path), 1)
        rows = get_company_rows(self.path)
        self.assertEqual(rows[0][1], "Big Tech (AI Investment)")

    def test_missing_career_url_defaults_na(self):
        upsert_companies(self.path, [
            {"company_name": "NoCareers", "ai_domain": "AI Startups",
             "business_focus": "Focus"},
        ])
        rows = get_company_rows(self.path)
        self.assertEqual(rows[0][3], "N/A")


class TestP0_5UpsertCompaniesPreservesCounts(unittest.TestCase):
    """P0-5: upsert_companies must NOT reset cols 6-9 (TPM Jobs / AI TPM Jobs /
    No TPM Count / Auto Archived) on existing rows. Those columns are managed
    by update_company_job_counts and the auto-archival pipeline; resetting
    them on upsert silently destroys data when the pipeline crashes between
    company discovery and JD scraping."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _seed_company_with_counts(self, name: str, tpm: int, ai_tpm: int,
                                  no_tpm: int, auto_archived: str):
        upsert_companies(self.path, [{
            "company_name": name, "ai_domain": "NLP",
            "business_focus": "X", "career_url": "https://example.com/jobs"
        }])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        # Find the row by name
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == name:
                ws.cell(r, 6).value = tpm
                ws.cell(r, 7).value = ai_tpm
                ws.cell(r, 8).value = no_tpm
                ws.cell(r, 9).value = auto_archived
                break
        wb.save(self.path)
        wb.close()

    def _read_row(self, name: str):
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        try:
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, 1).value == name:
                    return [ws.cell(r, c).value for c in range(1, 10)]
            return None
        finally:
            wb.close()

    def test_existing_row_preserves_tpm_counts(self):
        self._seed_company_with_counts("Acme", tpm=5, ai_tpm=2,
                                       no_tpm=1, auto_archived="Yes")
        # Re-upsert with NEW career_url
        upsert_companies(self.path, [{
            "company_name": "Acme", "ai_domain": "NLP",
            "business_focus": "X", "career_url": "https://acme.io/careers"
        }])
        row = self._read_row("Acme")
        self.assertEqual(row[3], "https://acme.io/careers", "career_url should update")
        self.assertEqual(row[5], 5, "TPM Jobs must be preserved")
        self.assertEqual(row[6], 2, "AI TPM Jobs must be preserved")
        self.assertEqual(row[7], 1, "No TPM Count must be preserved")
        self.assertEqual(row[8], "Yes", "Auto Archived must be preserved")

    def test_new_row_initializes_counts_to_zero(self):
        upsert_companies(self.path, [{
            "company_name": "NewCo", "ai_domain": "NLP",
            "business_focus": "X", "career_url": "https://newco.com"
        }])
        row = self._read_row("NewCo")
        self.assertEqual(row[5], 0)
        self.assertEqual(row[6], 0)
        self.assertEqual(row[7], 0)
        self.assertEqual(row[8], "No")

    def test_existing_row_updates_cols_1_to_5(self):
        self._seed_company_with_counts("MutCo", tpm=9, ai_tpm=3,
                                       no_tpm=0, auto_archived="No")
        upsert_companies(self.path, [{
            "company_name": "MutCo", "ai_domain": "Robotics",
            "business_focus": "Updated focus",
            "career_url": "https://mutco.io/jobs"
        }])
        row = self._read_row("MutCo")
        self.assertEqual(row[1], "Robotics")
        self.assertEqual(row[2], "Updated focus")
        self.assertEqual(row[3], "https://mutco.io/jobs")
        # Updated At (col 5, index 4) should be a non-empty string
        self.assertIsNotNone(row[4])
        self.assertNotEqual(row[4], "")
        # Counts preserved
        self.assertEqual(row[5], 9)
        self.assertEqual(row[6], 3)


class TestGetCompanyNamesWithoutTPM(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_empty_when_no_data(self):
        names = get_company_names_without_tpm(self.path)
        self.assertIsInstance(names, set)
        self.assertEqual(len(names), 0)

    def test_reads_without_tpm_sheet(self):
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_Without_TPM"]
        ws.append(["CompanyX", "AI Startups", "Focus", "https://x.com", "2026-01-01"])
        wb.save(self.path)
        names = get_company_names_without_tpm(self.path)
        self.assertIn("CompanyX", names)


class TestUpdateCompanyCareerUrl(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)
        upsert_companies(self.path, [
            {"company_name": "UrlCo", "ai_domain": "AI Startups",
             "business_focus": "F", "career_url": "https://old.com/jobs"},
        ])

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_updates_url(self):
        # First data row is Excel row 2 (row 1 = header)
        update_company_career_url(self.path, 2, "https://new.com/careers")
        rows = get_company_rows(self.path)
        self.assertEqual(rows[0][3], "https://new.com/careers")


class TestBug05CareerUrlRowAlignment(unittest.TestCase):
    """Regression test for BUG-05: update_company_career_url must write to the
    correct row even when empty rows exist between data rows in the sheet."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _insert_with_gap(self):
        """Insert CompanyA at row 2, leave row 3 empty, insert CompanyB at row 4."""
        import openpyxl
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        # row 2: CompanyA
        ws.cell(2, 1, "CompanyA"); ws.cell(2, 4, "https://a.com/jobs")
        # row 3: intentionally empty (simulates real-world sparse sheets)
        # row 4: CompanyB
        ws.cell(4, 1, "CompanyB"); ws.cell(4, 4, "https://b.com/jobs")
        wb.save(self.path)

    def test_get_company_rows_with_row_num_returns_actual_excel_rows(self):
        """get_company_rows_with_row_num() must return actual Excel row numbers,
        not list indices — critical for correct write-back when gaps exist."""
        self._insert_with_gap()
        indexed = get_company_rows_with_row_num(self.path)
        # Should return two entries with actual row numbers 2 and 4
        self.assertEqual(len(indexed), 2)
        excel_rows = [r for r, _ in indexed]
        self.assertIn(2, excel_rows, "CompanyA should be at Excel row 2")
        self.assertIn(4, excel_rows, "CompanyB should be at Excel row 4")

    def test_update_second_company_with_gap_does_not_corrupt_first(self):
        """BUG-05: when an empty row separates two companies, updating the
        second company must write to Excel row 4, not row 3."""
        self._insert_with_gap()
        indexed = get_company_rows_with_row_num(self.path)
        # Find CompanyB's actual Excel row
        company_b_entry = next((r, row) for r, row in indexed if row[0] == "CompanyB")
        excel_row_b, _ = company_b_entry
        self.assertEqual(excel_row_b, 4, "CompanyB must be at Excel row 4")

        update_company_career_url(self.path, excel_row_b, "https://b.com/careers-new")

        rows = get_company_rows(self.path)
        names = [r[0] for r in rows]
        urls  = {r[0]: r[3] for r in rows}

        # CompanyA must be untouched
        self.assertEqual(urls.get("CompanyA"), "https://a.com/jobs",
                         "CompanyA URL must NOT be modified (BUG-05 regression)")
        # CompanyB must have the new URL
        self.assertEqual(urls.get("CompanyB"), "https://b.com/careers-new",
                         "CompanyB URL must be updated correctly")


class TestJDHelpers(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _insert_jd(self, url, company="TestCorp", is_ai_tpm=True,
                   location="Remote", tech="pytorch", resp="Lead AI"):
        jd_json = json.dumps({
            "job_title": "TPM",
            "company": company,
            "location": location,
            "salary_range": "$200k",
            "requirements": [tech],
            "additional_qualifications": ["Nice to have"],
            "key_responsibilities": [resp],
            "is_ai_tpm": is_ai_tpm,
        })
        upsert_jd_record(self.path, url, jd_json, "abc123")

    def test_get_jd_urls_empty(self):
        urls = get_jd_urls(self.path)
        self.assertEqual(urls, [])

    def test_get_jd_urls_after_insert(self):
        self._insert_jd("https://jobs.lever.co/testcorp/123")
        urls = get_jd_urls(self.path)
        self.assertIn("https://jobs.lever.co/testcorp/123", urls)

    def test_get_jd_urls_excludes_na_company(self):
        # Manually insert a row with N/A company
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        ws.append(["https://example.com/job1", "Bad JD", "N/A",
                   "", "", "", "", "False", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
        wb.save(self.path)
        urls = get_jd_urls(self.path)
        self.assertNotIn("https://example.com/job1", urls)

    def test_get_jd_url_meta_has_hash(self):
        self._insert_jd("https://greenhouse.io/testco/job/1")
        meta = get_jd_url_meta(self.path)
        self.assertIn("https://greenhouse.io/testco/job/1", meta)
        self.assertIn("hash", meta["https://greenhouse.io/testco/job/1"])

    def test_get_jd_url_meta_age_days(self):
        self._insert_jd("https://greenhouse.io/testco/job/2")
        meta = get_jd_url_meta(self.path)
        age = meta["https://greenhouse.io/testco/job/2"]["age_days"]
        self.assertLess(age, 1.0)  # just inserted, should be < 1 day

    def test_get_jd_rows_for_match_only_ai_tpm(self):
        self._insert_jd("https://lever.co/x/ai", is_ai_tpm=True)
        self._insert_jd("https://lever.co/x/notai", is_ai_tpm=False)
        rows = get_jd_rows_for_match(self.path)
        urls = [r["url"] for r in rows]
        self.assertIn("https://lever.co/x/ai", urls)
        self.assertNotIn("https://lever.co/x/notai", urls)

    def test_upsert_jd_record_update_existing(self):
        url = "https://jobs.lever.co/testco/xyz"
        self._insert_jd(url, location="NYC")
        jd_json = json.dumps({
            "job_title": "Sr TPM", "company": "TestCorp",
            "location": "San Francisco", "salary_range": "$250k",
            "core_ai_tech_stack": ["tensorflow"],
            "key_responsibilities": ["Lead ML"],
            "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, url, jd_json, "newhash")
        # Should not duplicate
        urls = get_jd_urls(self.path)
        self.assertEqual(urls.count(url), 1)

    def test_upsert_jd_record_bad_json(self):
        url = "https://example.com/bad"
        upsert_jd_record(self.path, url, "NOT JSON", "hash")
        # Should not raise; URL should be in tracker with JSON ERROR
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        found = False
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == url:
                found = True
                self.assertEqual(ws.cell(r, 2).value, "JSON ERROR")
        self.assertTrue(found)

    def test_batch_upsert_jd_records(self):
        records = [
            ("https://gh.io/a/1", json.dumps({
                "job_title": "TPM", "company": "CompA", "location": "Remote",
                "salary_range": "N/A", "core_ai_tech_stack": ["llm"],
                "key_responsibilities": ["ship"], "is_ai_tpm": True,
            }), "hash1"),
            ("https://gh.io/a/2", json.dumps({
                "job_title": "PM", "company": "CompB", "location": "NYC",
                "salary_range": "N/A", "core_ai_tech_stack": [],
                "key_responsibilities": [], "is_ai_tpm": False,
            }), "hash2"),
        ]
        n = batch_upsert_jd_records(self.path, records)
        self.assertEqual(n, 2)
        urls = get_jd_urls(self.path)
        self.assertIn("https://gh.io/a/1", urls)


class TestGetIncompleteJdRows(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_complete_jd_not_returned(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "FullCo", "location": "Remote",
            "salary_range": "$200k", "requirements": ["pytorch"],
            "additional_qualifications": ["Nice to have"],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://jd.com/1", jd_json, "h")
        incomplete = get_incomplete_jd_rows(self.path)
        urls = [r["url"] for r in incomplete]
        self.assertNotIn("https://jd.com/1", urls)

    def test_missing_location_returned(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "NoCo", "location": "N/A",
            "salary_range": "N/A", "core_ai_tech_stack": ["pytorch"],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://jd.com/2", jd_json, "h")
        incomplete = get_incomplete_jd_rows(self.path)
        urls = [r["url"] for r in incomplete]
        self.assertIn("https://jd.com/2", urls)


class TestCountTpmJobsByCompany(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_counts_jobs_correctly(self):
        for i, ai in enumerate([True, True, False]):
            jd_json = json.dumps({
                "job_title": "TPM", "company": "BigCo", "location": "Remote",
                "salary_range": "N/A", "core_ai_tech_stack": [],
                "key_responsibilities": [], "is_ai_tpm": ai,
            })
            upsert_jd_record(self.path, f"https://bco.com/job/{i}", jd_json, f"h{i}")
        counts = count_tpm_jobs_by_company(self.path)
        self.assertIn("BigCo", counts)
        self.assertEqual(counts["BigCo"]["tpm"], 3)
        self.assertEqual(counts["BigCo"]["ai_tpm"], 2)

    def test_empty_tracker_returns_empty_dict(self):
        counts = count_tpm_jobs_by_company(self.path)
        self.assertEqual(counts, {})


class TestUpdateCompanyJobCounts(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)
        upsert_companies(self.path, [
            {"company_name": "BigCo", "ai_domain": "Big Tech (AI Investment)",
             "business_focus": "F", "career_url": "https://bigco.com/jobs"},
        ])

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_writes_tpm_counts(self):
        update_company_job_counts(self.path, {
            "BigCo": {"tpm": 5, "ai_tpm": 3}
        })
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        tpm_col = headers.get("TPM Jobs")
        ai_tpm_col = headers.get("AI TPM Jobs")
        self.assertIsNotNone(tpm_col)
        self.assertEqual(ws.cell(2, tpm_col).value, 5)
        self.assertEqual(ws.cell(2, ai_tpm_col).value, 3)


class TestBatchUpdateJdTimestamps(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_updates_known_urls(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "TimeCo", "location": "Remote",
            "salary_range": "N/A", "core_ai_tech_stack": [],
            "key_responsibilities": [], "is_ai_tpm": True,
        })
        url = "https://timeco.com/job/1"
        upsert_jd_record(self.path, url, jd_json, "h")
        n = batch_update_jd_timestamps(self.path, [url])
        self.assertEqual(n, 1)

    def test_empty_list_returns_zero(self):
        n = batch_update_jd_timestamps(self.path, [])
        self.assertEqual(n, 0)


class TestMatchHelpers(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _match_json(self, score=75):
        return json.dumps({
            "compatibility_score": score,
            "key_strengths": ["Strong LLM background"],
            "critical_gaps": ["No GenAI prod deployment"],
            "recommendation_reason": "Good fit overall.",
        })

    def test_get_match_pairs_empty(self):
        pairs = get_match_pairs(self.path)
        self.assertEqual(pairs, {})

    def test_upsert_match_record_insert(self):
        upsert_match_record(
            self.path, "resume_v1", "https://jobs.lever.co/x/1",
            self._match_json(80), "abc", "fine"
        )
        pairs = get_match_pairs(self.path)
        key = ("resume_v1", "https://jobs.lever.co/x/1")
        self.assertIn(key, pairs)
        self.assertEqual(pairs[key]["score"], 80)

    def test_upsert_match_record_update(self):
        url = "https://jobs.lever.co/x/2"
        upsert_match_record(self.path, "r1", url, self._match_json(50), "h", "coarse")
        upsert_match_record(self.path, "r1", url, self._match_json(85), "h2", "fine")
        pairs = get_match_pairs(self.path)
        self.assertEqual(pairs[("r1", url)]["score"], 85)
        self.assertEqual(pairs[("r1", url)]["stage"], "fine")

    def test_get_match_pairs_stage_and_hash(self):
        upsert_match_record(
            self.path, "rv2", "https://gh.io/c/1",
            self._match_json(60), "resumehash", "coarse"
        )
        pairs = get_match_pairs(self.path)
        info = pairs[("rv2", "https://gh.io/c/1")]
        self.assertEqual(info["stage"], "coarse")
        self.assertEqual(info["hash"], "resumehash")

    def test_batch_upsert_match_records(self):
        records = [
            ("r1", "https://url1.com", self._match_json(70), "h1", "coarse"),
            ("r1", "https://url2.com", self._match_json(90), "h1", "fine"),
        ]
        n = batch_upsert_match_records(self.path, records)
        self.assertEqual(n, 2)
        pairs = get_match_pairs(self.path)
        self.assertIn(("r1", "https://url1.com"), pairs)
        self.assertIn(("r1", "https://url2.com"), pairs)

    def test_batch_upsert_empty_returns_zero(self):
        n = batch_upsert_match_records(self.path, [])
        self.assertEqual(n, 0)

    def test_upsert_match_bad_json(self):
        upsert_match_record(self.path, "rx", "https://bad.com/jd", "NOT JSON", "", "fine")
        pairs = get_match_pairs(self.path)
        self.assertIn(("rx", "https://bad.com/jd"), pairs)
        self.assertEqual(pairs[("rx", "https://bad.com/jd")]["score"], 0)


class TestBug12WorkbookClose(unittest.TestCase):
    """BUG-12 regression: every load_workbook() call must be followed by wb.close()."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _assert_close_called(self, fn, *args):
        """Wrap load_workbook to spy on close(); fail if close() is never called."""
        import openpyxl as _openpyxl
        closed = []
        original_load = _openpyxl.load_workbook

        def spy_load(*a, **kw):
            wb = original_load(*a, **kw)
            orig_close = wb.close
            def tracked_close():
                closed.append(True)
                orig_close()
            wb.close = tracked_close
            return wb

        with unittest.mock.patch("shared.excel_store.load_workbook", side_effect=spy_load):
            fn(*args)

        self.assertTrue(
            len(closed) > 0,
            f"{fn.__name__} did not call wb.close() — file handle leak (BUG-12)",
        )

    def test_count_company_rows_closes_workbook(self):
        self._assert_close_called(count_company_rows, self.path)

    def test_get_company_rows_closes_workbook(self):
        self._assert_close_called(get_company_rows, self.path)

    def test_get_company_rows_with_row_num_closes_workbook(self):
        self._assert_close_called(get_company_rows_with_row_num, self.path)

    def test_get_company_names_without_tpm_closes_workbook(self):
        self._assert_close_called(get_company_names_without_tpm, self.path)

    def test_get_jd_urls_closes_workbook(self):
        self._assert_close_called(get_jd_urls, self.path)

    def test_get_jd_url_meta_closes_workbook(self):
        self._assert_close_called(get_jd_url_meta, self.path)

    def test_get_jd_rows_for_match_closes_workbook(self):
        self._assert_close_called(get_jd_rows_for_match, self.path)

    def test_get_incomplete_jd_rows_closes_workbook(self):
        self._assert_close_called(get_incomplete_jd_rows, self.path)

    def test_count_tpm_jobs_by_company_closes_workbook(self):
        self._assert_close_called(count_tpm_jobs_by_company, self.path)

    def test_get_match_pairs_closes_workbook(self):
        self._assert_close_called(get_match_pairs, self.path)

    def test_upsert_companies_closes_workbook(self):
        self._assert_close_called(
            upsert_companies, self.path,
            [{"company_name": "CloseTestCo", "ai_domain": "AI", "business_focus": "F"}],
        )

    def test_update_company_career_url_closes_workbook(self):
        upsert_companies(self.path, [
            {"company_name": "CloseCo", "ai_domain": "AI",
             "business_focus": "F", "career_url": "https://old.com"},
        ])
        self._assert_close_called(update_company_career_url, self.path, 2, "https://new.com")

    def test_upsert_jd_record_closes_workbook(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "CloseCo", "location": "Remote",
            "salary_range": "N/A", "core_ai_tech_stack": [], "key_responsibilities": [],
            "is_ai_tpm": True,
        })
        self._assert_close_called(
            upsert_jd_record, self.path, "https://close.test/1", jd_json, "h"
        )

    def test_batch_upsert_jd_records_closes_workbook(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "CloseCo", "location": "Remote",
            "salary_range": "N/A", "core_ai_tech_stack": [], "key_responsibilities": [],
            "is_ai_tpm": True,
        })
        self._assert_close_called(
            batch_upsert_jd_records, self.path,
            [("https://close.test/2", jd_json, "h")],
        )

    def test_upsert_match_record_closes_workbook_on_early_return(self):
        """Ensure close() is called even when upsert_match_record takes the early-return path."""
        match_json = json.dumps({
            "compatibility_score": 80, "key_strengths": [], "critical_gaps": [],
            "recommendation_reason": "ok",
        })
        # Insert once so the update (early-return) path is exercised on second call
        upsert_match_record(self.path, "r_close", "https://close.test/mr", match_json, "h", "fine")
        self._assert_close_called(
            upsert_match_record, self.path, "r_close", "https://close.test/mr",
            match_json, "h2", "fine",
        )

    def test_batch_upsert_match_records_closes_workbook(self):
        match_json = json.dumps({
            "compatibility_score": 75, "key_strengths": [], "critical_gaps": [],
            "recommendation_reason": "ok",
        })
        self._assert_close_called(
            batch_upsert_match_records, self.path,
            [("r_close2", "https://close.test/3", match_json, "h", "fine")],
        )

    def test_update_company_job_counts_closes_workbook(self):
        self._assert_close_called(
            update_company_job_counts, self.path, {"CloseCo": {"tpm": 1, "ai_tpm": 0}}
        )

    def test_batch_update_jd_timestamps_closes_workbook(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "CloseCo", "location": "Remote",
            "salary_range": "N/A", "core_ai_tech_stack": [], "key_responsibilities": [],
            "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://close.test/ts", jd_json, "h")
        self._assert_close_called(
            batch_update_jd_timestamps, self.path, ["https://close.test/ts"]
        )


class TestTailoredMatchResultsSheet(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_sheet_created(self):
        wb = openpyxl.load_workbook(self.path)
        self.assertIn("Tailored_Match_Results", wb.sheetnames)

    def test_headers_correct(self):
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Tailored_Match_Results"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for h in TAILORED_HEADERS:
            self.assertIn(h, headers)

    def test_migration_adds_sheet_to_existing_file(self):
        """If an older Excel has no Tailored_Match_Results sheet, migration adds it."""
        wb = openpyxl.load_workbook(self.path)
        if "Tailored_Match_Results" in wb.sheetnames:
            del wb["Tailored_Match_Results"]
        wb.save(self.path)
        wb.close()
        get_or_create_excel(self.path)
        wb2 = openpyxl.load_workbook(self.path)
        self.assertIn("Tailored_Match_Results", wb2.sheetnames)


class TestGetScoredMatches(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _match_json(self, score=75):
        return json.dumps({
            "compatibility_score": score,
            "key_strengths": ["Strong LLM background"],
            "critical_gaps": ["No GenAI prod deployment"],
            "recommendation_reason": "Good fit overall.",
        })

    def test_empty_returns_empty(self):
        results = get_scored_matches(self.path)
        self.assertEqual(results, [])

    def test_default_returns_only_fine_stage(self):
        """P0-6: by default get_scored_matches returns only Stage 2 'fine' rows.
        Stage 1 'coarse' heuristic scores are excluded so they never enter
        the tailored-resume score-delta computation."""
        upsert_match_record(self.path, "r1", "https://a.com/1", self._match_json(80), "h1", "fine")
        upsert_match_record(self.path, "r1", "https://a.com/2", self._match_json(0), "h1", "coarse")
        upsert_match_record(self.path, "r1", "https://a.com/3", self._match_json(45), "h1", "coarse")
        results = get_scored_matches(self.path)
        urls = [r["jd_url"] for r in results]
        self.assertEqual(urls, ["https://a.com/1"])
        self.assertEqual(results[0]["stage"], "fine")

    def test_stage_none_returns_all_stages(self):
        """P0-6: stage=None disables the filter (used by tests / migrations)."""
        upsert_match_record(self.path, "r1", "https://a.com/1", self._match_json(80), "h1", "fine")
        upsert_match_record(self.path, "r1", "https://a.com/2", self._match_json(45), "h1", "coarse")
        results = get_scored_matches(self.path, stage=None)
        urls = sorted(r["jd_url"] for r in results)
        self.assertEqual(urls, ["https://a.com/1", "https://a.com/2"])

    def test_explicit_stage_coarse_returns_only_coarse(self):
        upsert_match_record(self.path, "r1", "https://a.com/1", self._match_json(80), "h1", "fine")
        upsert_match_record(self.path, "r1", "https://a.com/2", self._match_json(45), "h1", "coarse")
        results = get_scored_matches(self.path, stage="coarse")
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["jd_url"], "https://a.com/2")

    def test_returns_correct_fields(self):
        upsert_match_record(self.path, "r1", "https://a.com/1", self._match_json(70), "hash1", "fine")
        results = get_scored_matches(self.path)
        self.assertEqual(len(results), 1)
        rec = results[0]
        self.assertEqual(rec["resume_id"], "r1")
        self.assertEqual(rec["jd_url"], "https://a.com/1")
        self.assertEqual(rec["score"], 70)
        self.assertEqual(rec["stage"], "fine")
        self.assertEqual(rec["resume_hash"], "hash1")


class TestGetTailoredMatchPairs(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_empty_returns_empty(self):
        pairs = get_tailored_match_pairs(self.path)
        self.assertEqual(pairs, {})

    def test_returns_inserted_records(self):
        batch_upsert_tailored_records(self.path, [{
            "resume_id": "r1", "jd_url": "https://a.com/1",
            "job_title": "TPM", "company": "TestCo",
            "original_score": 60, "tailored_score": 80, "score_delta": 20,
            "tailored_resume_path": "tailored_resumes/r1/abc.md",
            "optimization_summary": "Improved keywords",
            "resume_hash": "hash1",
        }])
        pairs = get_tailored_match_pairs(self.path)
        key = ("r1", "https://a.com/1")
        self.assertIn(key, pairs)
        self.assertEqual(pairs[key]["tailored_score"], 80)
        self.assertEqual(pairs[key]["resume_hash"], "hash1")

    def test_missing_sheet_returns_empty(self):
        """If sheet doesn't exist (e.g. old file), return empty dict."""
        wb = openpyxl.load_workbook(self.path)
        if "Tailored_Match_Results" in wb.sheetnames:
            del wb["Tailored_Match_Results"]
        wb.save(self.path)
        wb.close()
        pairs = get_tailored_match_pairs(self.path)
        self.assertEqual(pairs, {})


class TestBatchUpsertTailoredRecords(unittest.TestCase):

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _rec(self, url="https://a.com/1", original=60, tailored=80):
        return {
            "resume_id": "r1", "jd_url": url,
            "job_title": "TPM", "company": "TestCo",
            "original_score": original, "tailored_score": tailored,
            "score_delta": tailored - original,
            "tailored_resume_path": f"tailored_resumes/r1/abc.md",
            "optimization_summary": "Improved keywords",
            "resume_hash": "hash1",
        }

    def test_insert_records(self):
        n = batch_upsert_tailored_records(self.path, [
            self._rec("https://a.com/1"),
            self._rec("https://a.com/2", 50, 70),
        ])
        self.assertEqual(n, 2)
        pairs = get_tailored_match_pairs(self.path)
        self.assertEqual(len(pairs), 2)

    def test_update_existing_record(self):
        batch_upsert_tailored_records(self.path, [self._rec("https://a.com/1", 60, 80)])
        batch_upsert_tailored_records(self.path, [self._rec("https://a.com/1", 60, 90)])
        pairs = get_tailored_match_pairs(self.path)
        self.assertEqual(len(pairs), 1)
        self.assertEqual(pairs[("r1", "https://a.com/1")]["tailored_score"], 90)

    def test_empty_returns_zero(self):
        n = batch_upsert_tailored_records(self.path, [])
        self.assertEqual(n, 0)

    def test_writes_all_columns(self):
        batch_upsert_tailored_records(self.path, [self._rec()])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Tailored_Match_Results"]
        self.assertEqual(ws.cell(2, 1).value, "r1")  # Resume ID
        self.assertEqual(ws.cell(2, 3).value, "TPM")  # Job Title
        self.assertEqual(ws.cell(2, 4).value, "TestCo")  # Company
        self.assertEqual(ws.cell(2, 5).value, 60)  # Original Score
        self.assertEqual(ws.cell(2, 6).value, 80)  # Tailored Score
        self.assertEqual(ws.cell(2, 7).value, 20)  # Score Delta
        self.assertIn("tailored_resumes", str(ws.cell(2, 8).value))  # Path
        self.assertEqual(ws.cell(2, 11).value, "hash1")  # Resume Hash
        self.assertEqual(ws.cell(2, 12).value, False)  # Regression (delta=20 → False)

    def test_regression_explicit_true(self):
        """When caller passes regression=True, it round-trips."""
        rec = self._rec("https://a.com/1", 80, 75)  # delta=-5
        rec["regression"] = True
        batch_upsert_tailored_records(self.path, [rec])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Tailored_Match_Results"]
        self.assertEqual(ws.cell(2, 12).value, True)

    def test_regression_inferred_from_negative_delta(self):
        """If caller omits regression, it's inferred from score_delta < 0."""
        rec = self._rec("https://a.com/1", 80, 75)  # delta=-5
        rec.pop("regression", None)
        batch_upsert_tailored_records(self.path, [rec])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Tailored_Match_Results"]
        self.assertEqual(ws.cell(2, 12).value, True)

    def test_regression_false_for_positive_delta(self):
        """No regression when tailored >= original."""
        rec = self._rec("https://a.com/1", 60, 75)  # delta=+15
        rec.pop("regression", None)
        batch_upsert_tailored_records(self.path, [rec])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Tailored_Match_Results"]
        self.assertEqual(ws.cell(2, 12).value, False)

    def test_regression_false_for_zero_delta(self):
        """Tailored == base is not a regression."""
        rec = self._rec("https://a.com/1", 70, 70)  # delta=0
        rec.pop("regression", None)
        batch_upsert_tailored_records(self.path, [rec])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Tailored_Match_Results"]
        self.assertEqual(ws.cell(2, 12).value, False)


class TestTailoredHeadersHasRegression(unittest.TestCase):
    """TAILORED_HEADERS must include the Regression column at column 12."""

    def test_regression_in_headers(self):
        self.assertIn("Regression", TAILORED_HEADERS)

    def test_regression_position_stable(self):
        """Regression must stay at column 12 (index 11) so existing readers
        with hardcoded column references keep working. PR 2 appends NEW
        per-dim columns AFTER Regression — Regression itself doesn't move.
        """
        self.assertEqual(TAILORED_HEADERS.index("Regression"), 11)


class TestTailoredMigration(unittest.TestCase):
    """Existing files without Regression column should auto-migrate on open."""

    # Pre-P0-12 schema: 11 columns, no Regression, no per-dim columns.
    # Hardcoded so the test is stable across future schema additions.
    _PRE_P0_12_HEADERS = [
        "Resume ID", "JD URL", "Job Title", "Company", "Original Score",
        "Tailored Score", "Score Delta", "Tailored Resume Path",
        "Optimization Summary", "Updated At", "Resume Hash",
    ]

    def test_migration_adds_regression_column(self):
        path = _tmp_xlsx()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tailored_Match_Results"
            ws.append(self._PRE_P0_12_HEADERS)
            ws.append(["r1", "https://a.com/1", "TPM", "Co", 80, 70, -10,
                       "p", "s", "2026-01-01", "h"])
            ws.append(["r1", "https://a.com/2", "TPM", "Co", 60, 80, 20,
                       "p", "s", "2026-01-01", "h"])
            for required in ("Company_List", "Company_Without_TPM", "JD_Tracker",
                             "Match_Results"):
                wb.create_sheet(required)
            wb.save(path)
            wb.close()

            get_or_create_excel(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Tailored_Match_Results"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            self.assertIn("Regression", headers)
            reg_col = headers.index("Regression") + 1
            self.assertEqual(ws.cell(2, reg_col).value, True)   # delta=-10
            self.assertEqual(ws.cell(3, reg_col).value, False)  # delta=+20
            wb.close()
        finally:
            if os.path.exists(path):
                os.remove(path)


class TestPRJ002Headers(unittest.TestCase):
    """PRJ-002 PR 2 — 3-dimension scoring columns must be in headers."""

    def test_match_headers_has_ats_coverage(self):
        self.assertIn("ATS Coverage %", MATCH_HEADERS)

    def test_match_headers_has_recruiter_score(self):
        self.assertIn("Recruiter Score", MATCH_HEADERS)

    def test_match_headers_has_hm_score(self):
        self.assertIn("HM Score", MATCH_HEADERS)

    def test_match_headers_has_ats_missing(self):
        self.assertIn("ATS Missing", MATCH_HEADERS)

    def test_match_legacy_columns_unchanged(self):
        # Old column indices must stay valid for back-compat readers.
        self.assertEqual(MATCH_HEADERS.index("Score"), 2)
        self.assertEqual(MATCH_HEADERS.index("Resume Hash"), 7)
        self.assertEqual(MATCH_HEADERS.index("Stage"), 8)

    def test_tailored_headers_has_per_dim_columns(self):
        for col in ("Original ATS", "Tailored ATS", "ATS Delta",
                    "Original Recruiter", "Tailored Recruiter", "Recruiter Delta",
                    "Original HM", "Tailored HM", "HM Delta"):
            self.assertIn(col, TAILORED_HEADERS, f"Missing column: {col}")

    def test_tailored_legacy_columns_unchanged(self):
        # Old column indices must stay valid.
        self.assertEqual(TAILORED_HEADERS.index("Original Score"), 4)
        self.assertEqual(TAILORED_HEADERS.index("Tailored Score"), 5)
        self.assertEqual(TAILORED_HEADERS.index("Score Delta"), 6)
        self.assertEqual(TAILORED_HEADERS.index("Resume Hash"), 10)
        self.assertEqual(TAILORED_HEADERS.index("Regression"), 11)


class TestPRJ002MatchResultsMigration(unittest.TestCase):
    """PRJ-002 PR 2 — Match_Results gains 4 columns; old files auto-migrate."""

    # Pre-PRJ-002 Match_Results schema: 9 columns (post-P0 stage column).
    _PRE_PRJ002_HEADERS = [
        "Resume ID", "JD URL", "Score", "Strengths", "Gaps", "Reason",
        "Updated At", "Resume Hash", "Stage",
    ]

    def test_migration_adds_three_dim_columns(self):
        path = _tmp_xlsx()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Match_Results"
            ws.append(self._PRE_PRJ002_HEADERS)
            ws.append(["r1", "https://a.com/1", 75, "s1", "g1", "ok",
                       "2026-01-01", "h", "fine"])
            for required in ("Company_List", "Company_Without_TPM", "JD_Tracker",
                             "Tailored_Match_Results"):
                wb.create_sheet(required)
            wb.save(path)
            wb.close()

            get_or_create_excel(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Match_Results"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            for new_col in ("ATS Coverage %", "Recruiter Score", "HM Score", "ATS Missing"):
                self.assertIn(new_col, headers)
            # Existing row's legacy columns untouched.
            self.assertEqual(ws.cell(2, 3).value, 75)
            self.assertEqual(ws.cell(2, 9).value, "fine")
            # New columns blank for the existing row.
            for new_col in ("ATS Coverage %", "Recruiter Score", "HM Score", "ATS Missing"):
                col = headers.index(new_col) + 1
                self.assertIsNone(ws.cell(2, col).value)
            wb.close()
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_migration_idempotent(self):
        """Running get_or_create_excel twice should not duplicate new columns."""
        path = _tmp_xlsx()
        try:
            get_or_create_excel(path)
            get_or_create_excel(path)
            wb = openpyxl.load_workbook(path)
            ws = wb["Match_Results"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            self.assertEqual(headers.count("ATS Coverage %"), 1)
            self.assertEqual(headers.count("HM Score"), 1)
            wb.close()
        finally:
            if os.path.exists(path):
                os.remove(path)


class TestPRJ002TailoredMigration(unittest.TestCase):
    """PRJ-002 PR 2 — Tailored_Match_Results gains 9 per-dim columns."""

    # Post-P0-12, pre-PRJ-002: 12 columns.
    _PRE_PRJ002_HEADERS = [
        "Resume ID", "JD URL", "Job Title", "Company", "Original Score",
        "Tailored Score", "Score Delta", "Tailored Resume Path",
        "Optimization Summary", "Updated At", "Resume Hash", "Regression",
    ]

    def test_migration_adds_nine_per_dim_columns(self):
        path = _tmp_xlsx()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tailored_Match_Results"
            ws.append(self._PRE_PRJ002_HEADERS)
            ws.append(["r1", "https://a.com/1", "TPM", "Co", 60, 80, 20,
                       "p", "s", "2026-01-01", "h", False])
            for required in ("Company_List", "Company_Without_TPM", "JD_Tracker",
                             "Match_Results"):
                wb.create_sheet(required)
            wb.save(path)
            wb.close()

            get_or_create_excel(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Tailored_Match_Results"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            for new_col in ("Original ATS", "Tailored ATS", "ATS Delta",
                            "Original Recruiter", "Tailored Recruiter", "Recruiter Delta",
                            "Original HM", "Tailored HM", "HM Delta"):
                self.assertIn(new_col, headers)
            # Legacy columns and values untouched.
            self.assertEqual(ws.cell(2, 5).value, 60)   # Original Score
            self.assertEqual(ws.cell(2, 6).value, 80)   # Tailored Score
            self.assertEqual(ws.cell(2, 7).value, 20)   # Score Delta
            self.assertEqual(ws.cell(2, 12).value, False)  # Regression
            wb.close()
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_full_migration_pre_p0_12_to_latest(self):
        """End-to-end: pre-P0-12 schema (no Regression, no dim cols) → latest.

        This verifies that BOTH the P0-12 Regression migration and the
        PRJ-002 dim-column migration run successfully on a single
        outdated file.
        """
        path = _tmp_xlsx()
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tailored_Match_Results"
            ws.append([
                "Resume ID", "JD URL", "Job Title", "Company", "Original Score",
                "Tailored Score", "Score Delta", "Tailored Resume Path",
                "Optimization Summary", "Updated At", "Resume Hash",
            ])
            ws.append(["r1", "https://a.com/1", "TPM", "Co", 80, 70, -10,
                       "p", "s", "2026-01-01", "h"])
            for required in ("Company_List", "Company_Without_TPM", "JD_Tracker",
                             "Match_Results"):
                wb.create_sheet(required)
            wb.save(path)
            wb.close()

            get_or_create_excel(path)

            wb = openpyxl.load_workbook(path)
            ws = wb["Tailored_Match_Results"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            # All target columns now present
            for col in ("Regression", "Original ATS", "HM Delta"):
                self.assertIn(col, headers)
            # Regression backfilled from delta < 0
            reg_col = headers.index("Regression") + 1
            self.assertTrue(ws.cell(2, reg_col).value)
            wb.close()
        finally:
            if os.path.exists(path):
                os.remove(path)


class TestTailoredWorkbookClose(unittest.TestCase):
    """Ensure all tailored functions properly close workbook handles."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def _assert_close_called(self, fn, *args):
        import openpyxl as _openpyxl
        closed = []
        original_load = _openpyxl.load_workbook

        def spy_load(*a, **kw):
            wb = original_load(*a, **kw)
            orig_close = wb.close
            def tracked_close():
                closed.append(True)
                orig_close()
            wb.close = tracked_close
            return wb

        with unittest.mock.patch("shared.excel_store.load_workbook", side_effect=spy_load):
            fn(*args)

        self.assertTrue(
            len(closed) > 0,
            f"{fn.__name__} did not call wb.close() — file handle leak",
        )

    def test_get_scored_matches_closes(self):
        self._assert_close_called(get_scored_matches, self.path)

    def test_get_tailored_match_pairs_closes(self):
        self._assert_close_called(get_tailored_match_pairs, self.path)

    def test_batch_upsert_tailored_records_closes(self):
        rec = {
            "resume_id": "r1", "jd_url": "https://a.com/1",
            "job_title": "TPM", "company": "Co",
            "original_score": 60, "tailored_score": 80, "score_delta": 20,
            "tailored_resume_path": "p", "optimization_summary": "s",
            "resume_hash": "h",
        }
        self._assert_close_called(batch_upsert_tailored_records, self.path, [rec])


class TestDataQualityColumn(unittest.TestCase):
    """REQ-060: Data Quality column in JD_Tracker sheet."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_jd_headers_include_data_quality(self):
        self.assertIn("Data Quality", JD_HEADERS)

    def test_new_excel_has_data_quality_header(self):
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("Data Quality", headers)
        wb.close()

    def test_upsert_jd_record_writes_data_quality(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "QCo", "location": "Remote",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": [],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
            "data_quality": "complete",
        })
        upsert_jd_record(self.path, "https://qco.com/1", jd_json, "h")
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dq_col = headers.index("Data Quality") + 1
        self.assertEqual(ws.cell(2, dq_col).value, "complete")
        wb.close()

    def test_upsert_jd_record_writes_partial_quality(self):
        jd_json = json.dumps({
            "job_title": "TPM", "company": "QCo", "location": "",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": [],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
            "data_quality": "partial",
        })
        upsert_jd_record(self.path, "https://qco.com/2", jd_json, "h")
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dq_col = headers.index("Data Quality") + 1
        self.assertEqual(ws.cell(2, dq_col).value, "partial")
        wb.close()

    def test_batch_upsert_jd_records_writes_data_quality(self):
        records = [
            ("https://qco.com/3", json.dumps({
                "job_title": "TPM", "company": "QCo", "location": "NYC",
                "salary_range": "N/A", "requirements": ["ML"],
                "additional_qualifications": [],
                "key_responsibilities": ["Ship"], "is_ai_tpm": True,
                "data_quality": "complete",
            }), "h1"),
            ("https://qco.com/4", json.dumps({
                "job_title": "PM", "company": "QCo", "location": "",
                "salary_range": "N/A", "requirements": [],
                "additional_qualifications": [],
                "key_responsibilities": [], "is_ai_tpm": False,
                "data_quality": "failed",
            }), "h2"),
        ]
        batch_upsert_jd_records(self.path, records)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dq_col = headers.index("Data Quality") + 1
        self.assertEqual(ws.cell(2, dq_col).value, "complete")
        self.assertEqual(ws.cell(3, dq_col).value, "failed")
        wb.close()

    def test_upsert_bad_json_writes_failed_quality(self):
        upsert_jd_record(self.path, "https://bad.com/1", "NOT JSON", "h")
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dq_col = headers.index("Data Quality") + 1
        self.assertEqual(ws.cell(2, dq_col).value, "failed")
        wb.close()

    def test_batch_upsert_bad_json_writes_failed_quality(self):
        records = [("https://bad.com/2", "NOT JSON", "h")]
        batch_upsert_jd_records(self.path, records)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dq_col = headers.index("Data Quality") + 1
        self.assertEqual(ws.cell(2, dq_col).value, "failed")
        wb.close()

    def test_migration_adds_data_quality_column(self):
        """Existing Excel without Data Quality column gets it via migration."""
        # Remove the Data Quality column to simulate old file
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "Data Quality" in headers:
            dq_col = headers.index("Data Quality") + 1
            ws.delete_cols(dq_col)
        wb.save(self.path)
        wb.close()
        # Run migration
        get_or_create_excel(self.path)
        wb2 = openpyxl.load_workbook(self.path)
        ws2 = wb2["JD_Tracker"]
        headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
        self.assertIn("Data Quality", headers2)
        wb2.close()

    def test_data_quality_empty_when_not_provided(self):
        """If jd_json has no data_quality key, column should be empty/None."""
        jd_json = json.dumps({
            "job_title": "TPM", "company": "OldCo", "location": "Remote",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": [],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://oldco.com/1", jd_json, "h")
        wb = openpyxl.load_workbook(self.path)
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        dq_col = headers.index("Data Quality") + 1
        # openpyxl returns None for empty cells; "" is written but read back as None
        self.assertIn(ws.cell(2, dq_col).value, (None, ""))
        wb.close()


class TestAutoArchive(unittest.TestCase):
    """REQ-063: Auto-archive companies with no TPM jobs."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)
        # Seed two companies
        upsert_companies(self.path, [
            {"company_name": "ArchiveCo", "ai_domain": "LLM",
             "business_focus": "AI", "career_url": "https://archive.co/careers"},
            {"company_name": "ActiveCo", "ai_domain": "Vision",
             "business_focus": "AI", "career_url": "https://active.co/careers"},
        ])

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_company_headers_include_archive_columns(self):
        self.assertIn("No TPM Count", COMPANY_HEADERS)
        self.assertIn("Auto Archived", COMPANY_HEADERS)

    def test_new_excel_has_archive_headers(self):
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("No TPM Count", headers)
        self.assertIn("Auto Archived", headers)
        wb.close()

    def test_get_archived_companies_empty_initially(self):
        result = get_archived_companies(self.path)
        self.assertEqual(result, set())

    def test_get_archived_companies_returns_archived(self):
        update_archive_status(self.path, "ArchiveCo", 3, "yes")
        result = get_archived_companies(self.path)
        self.assertEqual(result, {"ArchiveCo"})

    def test_get_archived_companies_excludes_non_archived(self):
        update_archive_status(self.path, "ArchiveCo", 3, "yes")
        update_archive_status(self.path, "ActiveCo", 1, "no")
        result = get_archived_companies(self.path)
        self.assertIn("ArchiveCo", result)
        self.assertNotIn("ActiveCo", result)

    def test_update_archive_status_writes_count_and_flag(self):
        update_archive_status(self.path, "ArchiveCo", 2, "no")
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        cnt_col = headers["No TPM Count"]
        arch_col = headers["Auto Archived"]
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value).strip() == "ArchiveCo":
                self.assertEqual(ws.cell(r, cnt_col).value, 2)
                self.assertEqual(ws.cell(r, arch_col).value, "no")
                break
        else:
            self.fail("ArchiveCo not found in sheet")
        wb.close()

    def test_update_archive_status_sets_yes(self):
        update_archive_status(self.path, "ArchiveCo", 3, "yes")
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        arch_col = headers["Auto Archived"]
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value).strip() == "ArchiveCo":
                self.assertEqual(ws.cell(r, arch_col).value, "yes")
                break
        wb.close()

    def test_unarchive_company_resets(self):
        update_archive_status(self.path, "ArchiveCo", 3, "yes")
        self.assertIn("ArchiveCo", get_archived_companies(self.path))
        unarchive_company(self.path, "ArchiveCo")
        self.assertNotIn("ArchiveCo", get_archived_companies(self.path))
        info = get_company_archive_info(self.path)
        self.assertEqual(info["ArchiveCo"]["no_tpm_count"], 0)
        self.assertEqual(info["ArchiveCo"]["archived"], "no")

    def test_get_company_archive_info_returns_all(self):
        update_archive_status(self.path, "ArchiveCo", 3, "yes")
        update_archive_status(self.path, "ActiveCo", 1, "no")
        info = get_company_archive_info(self.path)
        self.assertEqual(info["ArchiveCo"]["no_tpm_count"], 3)
        self.assertEqual(info["ArchiveCo"]["archived"], "yes")
        self.assertEqual(info["ActiveCo"]["no_tpm_count"], 1)
        self.assertEqual(info["ActiveCo"]["archived"], "no")

    def test_migration_adds_archive_columns(self):
        """Old Excel without archive columns gets them via migration."""
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # Remove archive columns to simulate old file
        for col_name in ("No TPM Count", "Auto Archived"):
            if col_name in headers:
                idx = headers.index(col_name) + 1
                ws.delete_cols(idx)
                headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        wb.save(self.path)
        wb.close()
        # Run migration
        get_or_create_excel(self.path)
        wb2 = openpyxl.load_workbook(self.path)
        ws2 = wb2["Company_List"]
        headers2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
        self.assertIn("No TPM Count", headers2)
        self.assertIn("Auto Archived", headers2)
        wb2.close()

    def test_update_nonexistent_company_no_error(self):
        """Updating a company that doesn't exist should not raise."""
        update_archive_status(self.path, "GhostCo", 5, "yes")
        # GhostCo not in sheet, so nothing should change
        result = get_archived_companies(self.path)
        self.assertNotIn("GhostCo", result)


class TestCountValidTpmJobsByCompany(unittest.TestCase):
    """REQ-063: count_valid_tpm_jobs_by_company excludes data_quality='failed'."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_empty_returns_empty(self):
        result = count_valid_tpm_jobs_by_company(self.path)
        self.assertEqual(result, {})

    def test_counts_non_failed_records(self):
        jd_ok = json.dumps({
            "job_title": "TPM", "company": "TestCo", "location": "Remote",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": [],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
            "data_quality": "complete",
        })
        upsert_jd_record(self.path, "https://test.co/1", jd_ok, "h1")
        result = count_valid_tpm_jobs_by_company(self.path)
        self.assertEqual(result.get("TestCo"), 1)

    def test_excludes_failed_records(self):
        jd_ok = json.dumps({
            "job_title": "TPM", "company": "TestCo", "location": "Remote",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": [],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
            "data_quality": "complete",
        })
        jd_fail = json.dumps({
            "job_title": "TPM", "company": "TestCo", "location": "",
            "salary_range": "N/A", "requirements": [],
            "additional_qualifications": [],
            "key_responsibilities": [], "is_ai_tpm": False,
            "data_quality": "failed",
        })
        upsert_jd_record(self.path, "https://test.co/1", jd_ok, "h1")
        upsert_jd_record(self.path, "https://test.co/2", jd_fail, "h2")
        result = count_valid_tpm_jobs_by_company(self.path)
        self.assertEqual(result.get("TestCo"), 1)

    def test_company_only_failed_not_counted(self):
        jd_fail = json.dumps({
            "job_title": "TPM", "company": "FailCo", "location": "",
            "salary_range": "N/A", "requirements": [],
            "additional_qualifications": [],
            "key_responsibilities": [], "is_ai_tpm": False,
            "data_quality": "failed",
        })
        upsert_jd_record(self.path, "https://fail.co/1", jd_fail, "h1")
        result = count_valid_tpm_jobs_by_company(self.path)
        self.assertNotIn("FailCo", result)

    def test_partial_quality_counted(self):
        jd_partial = json.dumps({
            "job_title": "TPM", "company": "PartCo", "location": "",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": [],
            "key_responsibilities": ["Lead"], "is_ai_tpm": True,
            "data_quality": "partial",
        })
        upsert_jd_record(self.path, "https://part.co/1", jd_partial, "h1")
        result = count_valid_tpm_jobs_by_company(self.path)
        self.assertEqual(result.get("PartCo"), 1)


# ─────────────────────────────────────────────────────────────────────────────
class TestBug33NullListFields(unittest.TestCase):
    """BUG-33: d.get('requirements', []) returns None when Gemini returns null,
    causing TypeError in '\\n'.join(). Must handle null gracefully."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_requirements_null_no_crash(self):
        """When requirements is explicitly null, upsert must not raise TypeError."""
        jd_json = json.dumps({
            "job_title": "TPM", "company": "NullCo", "location": "SF, CA",
            "salary_range": "N/A", "requirements": None,
            "additional_qualifications": ["Nice to have"],
            "key_responsibilities": ["Lead projects"], "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://null.co/1", jd_json, "h1")
        ws = openpyxl.load_workbook(self.path)["JD_Tracker"]
        self.assertEqual(ws.cell(2, 6).value, "None")  # requirements → "None"
        ws.parent.close()

    def test_additional_qualifications_null_no_crash(self):
        """When additional_qualifications is null, upsert must not raise TypeError."""
        jd_json = json.dumps({
            "job_title": "TPM", "company": "NullCo2", "location": "NYC",
            "salary_range": "N/A", "requirements": ["Python"],
            "additional_qualifications": None,
            "key_responsibilities": ["Ship"], "is_ai_tpm": False,
        })
        upsert_jd_record(self.path, "https://null.co/2", jd_json, "h2")
        ws = openpyxl.load_workbook(self.path)["JD_Tracker"]
        self.assertEqual(ws.cell(2, 7).value, "None")  # addq → "None"
        ws.parent.close()

    def test_key_responsibilities_null_no_crash(self):
        """When key_responsibilities is null, upsert must not raise TypeError."""
        jd_json = json.dumps({
            "job_title": "TPM", "company": "NullCo3", "location": "LA",
            "salary_range": "N/A", "requirements": ["Go"],
            "additional_qualifications": ["K8s"],
            "key_responsibilities": None, "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://null.co/3", jd_json, "h3")
        ws = openpyxl.load_workbook(self.path)["JD_Tracker"]
        self.assertEqual(ws.cell(2, 8).value, "None")  # resp → "None"
        ws.parent.close()

    def test_all_three_null_no_crash(self):
        """When all three list fields are null, upsert must not raise TypeError."""
        jd_json = json.dumps({
            "job_title": "TPM", "company": "AllNull", "location": "Remote",
            "salary_range": "N/A", "requirements": None,
            "additional_qualifications": None,
            "key_responsibilities": None, "is_ai_tpm": False,
        })
        upsert_jd_record(self.path, "https://null.co/all", jd_json, "hall")
        ws = openpyxl.load_workbook(self.path)["JD_Tracker"]
        self.assertEqual(ws.cell(2, 6).value, "None")
        self.assertEqual(ws.cell(2, 7).value, "None")
        self.assertEqual(ws.cell(2, 8).value, "None")
        ws.parent.close()

    def test_missing_keys_still_work(self):
        """When keys are completely absent (not null), should still work."""
        jd_json = json.dumps({
            "job_title": "TPM", "company": "MissCo", "location": "Austin",
            "salary_range": "N/A", "is_ai_tpm": True,
        })
        upsert_jd_record(self.path, "https://miss.co/1", jd_json, "hm")
        ws = openpyxl.load_workbook(self.path)["JD_Tracker"]
        self.assertEqual(ws.cell(2, 6).value, "None")
        self.assertEqual(ws.cell(2, 7).value, "None")
        self.assertEqual(ws.cell(2, 8).value, "None")
        ws.parent.close()


# ─────────────────────────────────────────────────────────────────────────────
class TestBug40OperatorPrecedence(unittest.TestCase):
    """BUG-40: Operator precedence in get_match_pairs/get_scored_matches must be correct."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_parenthesized_correctly_in_source(self):
        """Verify the expressions use explicit parentheses for correct precedence."""
        import inspect
        from shared.excel_store import get_match_pairs, get_scored_matches
        for fn in (get_match_pairs, get_scored_matches):
            source = inspect.getsource(fn)
            # Should NOT have the ambiguous pattern without parens
            self.assertNotIn('.value or "" if ws.max_column', source,
                             f"{fn.__name__} has ambiguous operator precedence")
            self.assertNotIn(".value or \"fine\" if ws.max_column", source,
                             f"{fn.__name__} has ambiguous operator precedence")

    def test_match_pairs_hash_and_stage_with_values(self):
        """Verify hash and stage are correctly read when values exist."""
        upsert_match_record(self.path, "r1", "https://j.co/1",
                            '{"score":85}', "hashABC", stage="coarse")
        pairs = get_match_pairs(self.path)
        key = ("r1", "https://j.co/1")
        self.assertIn(key, pairs)
        self.assertEqual(pairs[key]["hash"], "hashABC")
        self.assertEqual(pairs[key]["stage"], "coarse")


# ─────────────────────────────────────────────────────────────────────────────
class TestBug45UpsertCompaniesInitAllColumns(unittest.TestCase):
    """BUG-45: upsert_companies must initialize all 9 COMPANY_HEADERS columns."""

    def setUp(self):
        self.path = _tmp_xlsx()
        get_or_create_excel(self.path)

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_new_company_has_all_columns(self):
        upsert_companies(self.path, [{"company_name": "TestCo", "ai_domain": "NLP"}])
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        # Row 2 should have 9 columns
        vals = [ws.cell(2, c).value for c in range(1, 10)]
        wb.close()
        self.assertEqual(vals[0], "TestCo")  # Company Name
        self.assertEqual(vals[5], 0)  # TPM Jobs
        self.assertEqual(vals[6], 0)  # AI TPM Jobs
        self.assertEqual(vals[7], 0)  # No TPM Count
        self.assertEqual(vals[8], "No")  # Auto Archived


class TestBug46DocstringColumnNumber(unittest.TestCase):
    """BUG-46: batch_update_jd_timestamps docstring must not reference wrong column number."""

    def test_docstring_no_wrong_column(self):
        # After BUG-52 fix, docstring no longer references hardcoded column numbers at all
        self.assertNotIn("col 9", batch_update_jd_timestamps.__doc__)
        self.assertIn("Updated At", batch_update_jd_timestamps.__doc__)


class TestBug55WithoutTpmHeadersMigration(unittest.TestCase):
    """BUG-55: WITHOUT_TPM_HEADERS must include TPM Jobs and AI TPM Jobs."""

    def setUp(self):
        self.path = _tmp_xlsx()

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    def test_headers_include_tpm_columns(self):
        from shared.excel_store import WITHOUT_TPM_HEADERS
        self.assertIn("TPM Jobs", WITHOUT_TPM_HEADERS)
        self.assertIn("AI TPM Jobs", WITHOUT_TPM_HEADERS)

    def test_new_excel_has_correct_without_tpm_headers(self):
        get_or_create_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_Without_TPM"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        wb.close()
        self.assertIn("TPM Jobs", headers)
        self.assertIn("AI TPM Jobs", headers)

    def test_migration_adds_columns_to_existing(self):
        # Create old-style file with only 5 columns
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Company_Without_TPM"
        ws.append(["Company Name", "AI Domain", "Business Focus", "Career URL", "Updated At"])
        wb.save(self.path)
        wb.close()
        # Run migration
        get_or_create_excel(self.path)
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_Without_TPM"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        wb.close()
        self.assertIn("TPM Jobs", headers)
        self.assertIn("AI TPM Jobs", headers)


class TestBug52JdTrackerDynamicColumns(unittest.TestCase):
    """BUG-52: JD_Tracker read functions must use _JD_COL for column lookup, not hardcoded numbers."""

    def test_jd_col_mapping_matches_headers(self):
        """_JD_COL must map every JD_HEADERS entry to correct 1-based index."""
        for i, h in enumerate(JD_HEADERS):
            self.assertEqual(_JD_COL[h], i + 1, f"_JD_COL['{h}'] should be {i + 1}")

    def test_no_hardcoded_column_numbers_in_jd_functions(self):
        """JD_Tracker read functions must not contain hardcoded ws.cell(r, <number>) patterns."""
        import inspect
        funcs = [get_jd_urls, get_jd_url_meta, batch_update_jd_timestamps,
                 get_jd_rows_for_match, get_incomplete_jd_rows,
                 count_tpm_jobs_by_company, count_valid_tpm_jobs_by_company]
        import re
        pattern = re.compile(r'ws\.cell\(r,\s*\d+\)')
        for fn in funcs:
            source = inspect.getsource(fn)
            matches = pattern.findall(source)
            self.assertEqual(matches, [],
                             f"{fn.__name__} still has hardcoded column numbers: {matches}")

    def test_get_jd_urls_reads_correct_columns(self):
        """End-to-end: get_jd_urls reads URL and Company from correct columns."""
        path = _tmp_xlsx()
        try:
            get_or_create_excel(path)
            wb = openpyxl.load_workbook(path)
            ws = wb["JD_Tracker"]
            # Write a valid row
            row = [None] * len(JD_HEADERS)
            row[JD_HEADERS.index("JD URL")] = "https://example.com/job/1"
            row[JD_HEADERS.index("Company")] = "TestCorp"
            ws.append(row)
            wb.save(path)
            wb.close()
            urls = get_jd_urls(path)
            self.assertIn("https://example.com/job/1", urls)
        finally:
            if os.path.exists(path):
                os.remove(path)

    def test_count_tpm_jobs_reads_correct_columns(self):
        """End-to-end: count_tpm_jobs_by_company reads Company and Is AI TPM correctly."""
        path = _tmp_xlsx()
        try:
            get_or_create_excel(path)
            wb = openpyxl.load_workbook(path)
            ws = wb["JD_Tracker"]
            row = [None] * len(JD_HEADERS)
            row[JD_HEADERS.index("JD URL")] = "https://example.com/job/1"
            row[JD_HEADERS.index("Company")] = "AcmeCorp"
            row[JD_HEADERS.index("Is AI TPM")] = "True"
            ws.append(row)
            row2 = [None] * len(JD_HEADERS)
            row2[JD_HEADERS.index("JD URL")] = "https://example.com/job/2"
            row2[JD_HEADERS.index("Company")] = "AcmeCorp"
            row2[JD_HEADERS.index("Is AI TPM")] = "False"
            ws.append(row2)
            wb.save(path)
            wb.close()
            counts = count_tpm_jobs_by_company(path)
            self.assertEqual(counts["AcmeCorp"]["tpm"], 2)
            self.assertEqual(counts["AcmeCorp"]["ai_tpm"], 1)
        finally:
            if os.path.exists(path):
                os.remove(path)


if __name__ == "__main__":
    unittest.main(verbosity=2)

"""
Tests for agents/resume_optimizer.py

Coverage:
  - Pydantic schemas: TailoredResume, MatchResult
  - load_resume: reads .md/.txt from folder
  - _load_jd_markdown: loads from jd_cache
  - _save_tailored_resume: writes to tailored_resumes/
  - tailor_resume: mocked Gemini call
  - re_score: mocked Gemini call
  - _GeminiKeyPool: key rotation
"""
import sys
import os
import json
import hashlib
import tempfile
import unittest
from unittest.mock import MagicMock, patch

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

# ── Stub heavy deps before import ────────────────────────────────────────────
for mod in ["google", "google.genai", "google.genai.types", "dotenv"]:
    if mod not in sys.modules:
        sys.modules[mod] = MagicMock()

sys.modules["dotenv"].load_dotenv = lambda: None

import google
google.genai = MagicMock()
google.genai.types = MagicMock()

from shared.prompts import (
    TAILOR_SYSTEM_PROMPT as _TAILOR_SYSTEM_PROMPT,
    FINE_SYSTEM_PROMPT as _FINE_SYSTEM_PROMPT,
)
from shared.schemas import (
    TailoredResume,
    MatchResult,
    BatchTailoredItem,
    BatchTailoredResult,
)
from agents.resume_optimizer import (
    BATCH_TAILOR_SIZE,
    RESCORE_CONCURRENCY,
    load_resume,
    _load_jd_markdown,
    _save_tailored_resume,
    _print_summary,
    tailor_resume,
    re_score,
    batch_tailor_resume,
    _GeminiKeyPool,
)
from shared.excel_store import TAILORED_HEADERS
import agents.resume_optimizer as optimizer_mod


# ─────────────────────────────────────────────────────────────────────────────
class TestPydanticSchemas(unittest.TestCase):

    def test_tailored_resume_valid(self):
        tr = TailoredResume(
            tailored_resume_markdown="# Resume\nContent here",
            optimization_summary="1. Reordered skills 2. Added keywords",
        )
        self.assertIn("Resume", tr.tailored_resume_markdown)
        self.assertIn("Reordered", tr.optimization_summary)

    def test_match_result_valid(self):
        m = MatchResult(
            compatibility_score=85,
            key_strengths=["LLM experience"],
            critical_gaps=["No MLOps"],
            recommendation_reason="Strong candidate.",
        )
        self.assertEqual(m.compatibility_score, 85)

    def test_tailored_resume_model_validate_json(self):
        raw = json.dumps({
            "tailored_resume_markdown": "# Resume",
            "optimization_summary": "Added keywords",
        })
        result = TailoredResume.model_validate_json(raw)
        self.assertEqual(result.tailored_resume_markdown, "# Resume")


# ─────────────────────────────────────────────────────────────────────────────
class TestLoadResume(unittest.TestCase):

    def _write_resume(self, folder, name, content):
        path = os.path.join(folder, name)
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)

    def test_loads_md_file(self):
        with tempfile.TemporaryDirectory() as d:
            self._write_resume(d, "resume.md", "# Resume\nAI TPM expert")
            text, rid = load_resume(d)
            self.assertIn("AI TPM", text)
            self.assertEqual(rid, "resume")

    def test_returns_empty_if_no_files(self):
        with tempfile.TemporaryDirectory() as d:
            text, rid = load_resume(d)
            self.assertEqual(text, "")

    def test_returns_empty_if_folder_missing(self):
        text, rid = load_resume("/nonexistent/path/xyz")
        self.assertEqual(text, "")


# ─────────────────────────────────────────────────────────────────────────────
class TestLoadJdMarkdown(unittest.TestCase):

    def test_prefers_structured(self):
        with tempfile.TemporaryDirectory() as d:
            url = "https://example.com/job/1"
            md5 = hashlib.md5(url.encode()).hexdigest()
            with open(os.path.join(d, f"{md5}.md"), "w") as f:
                f.write("raw content")
            with open(os.path.join(d, f"{md5}_structured.md"), "w") as f:
                f.write("structured content")
            with patch.object(optimizer_mod, "JD_CACHE_DIR", d):
                result = _load_jd_markdown(url)
            self.assertEqual(result, "structured content")

    def test_falls_back_to_raw(self):
        with tempfile.TemporaryDirectory() as d:
            url = "https://example.com/job/2"
            md5 = hashlib.md5(url.encode()).hexdigest()
            with open(os.path.join(d, f"{md5}.md"), "w") as f:
                f.write("raw content")
            with patch.object(optimizer_mod, "JD_CACHE_DIR", d):
                result = _load_jd_markdown(url)
            self.assertEqual(result, "raw content")

    def test_returns_none_if_missing(self):
        with tempfile.TemporaryDirectory() as d:
            with patch.object(optimizer_mod, "JD_CACHE_DIR", d):
                result = _load_jd_markdown("https://example.com/nonexistent")
            self.assertIsNone(result)


# ─────────────────────────────────────────────────────────────────────────────
class TestSaveTailoredResume(unittest.TestCase):

    def test_creates_file(self):
        with tempfile.TemporaryDirectory() as d:
            with patch.object(optimizer_mod, "TAILORED_DIR", d):
                path = _save_tailored_resume("r1", "https://a.com/1", "# Tailored")
            self.assertTrue(os.path.exists(path))
            with open(path) as f:
                self.assertEqual(f.read(), "# Tailored")

    def test_path_structure(self):
        with tempfile.TemporaryDirectory() as d:
            with patch.object(optimizer_mod, "TAILORED_DIR", d):
                path = _save_tailored_resume("r1", "https://a.com/1", "content")
            self.assertIn("r1", path)
            md5 = hashlib.md5("https://a.com/1".encode()).hexdigest()
            self.assertIn(md5, path)
            self.assertTrue(path.endswith(".md"))


# ─────────────────────────────────────────────────────────────────────────────
class TestTailorResume(unittest.TestCase):

    def setUp(self):
        mock_pool = MagicMock()
        mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({
                "tailored_resume_markdown": "# Tailored Resume",
                "optimization_summary": "1. Added keywords",
            })
        )
        optimizer_mod._KEY_POOL = mock_pool

    def tearDown(self):
        optimizer_mod._KEY_POOL = None

    def test_returns_json_string(self):
        result = tailor_resume("resume text", "jd content")
        data = json.loads(result)
        self.assertIn("tailored_resume_markdown", data)
        self.assertEqual(data["tailored_resume_markdown"], "# Tailored Resume")

    def test_returns_none_on_structural_error(self):
        """P0-4: structural error returns None, not '{}' fake JSON."""
        from shared.exceptions import GeminiStructuralError
        optimizer_mod._KEY_POOL.generate_content.side_effect = GeminiStructuralError("bad")
        self.assertIsNone(tailor_resume("resume", "jd"))

    def test_reraises_transient_error(self):
        from shared.exceptions import GeminiTransientError
        optimizer_mod._KEY_POOL.generate_content.side_effect = GeminiTransientError("429")
        with self.assertRaises(GeminiTransientError):
            tailor_resume("resume", "jd")


# ─────────────────────────────────────────────────────────────────────────────
class TestReScore(unittest.TestCase):

    def setUp(self):
        mock_pool = MagicMock()
        mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({
                "compatibility_score": 85,
                "key_strengths": ["Better alignment"],
                "critical_gaps": ["Still no MLOps"],
                "recommendation_reason": "Improved fit.",
            })
        )
        optimizer_mod._KEY_POOL = mock_pool

    def tearDown(self):
        optimizer_mod._KEY_POOL = None

    def test_returns_json_string(self):
        result = re_score("tailored resume", "jd content")
        data = json.loads(result)
        self.assertEqual(data["compatibility_score"], 85)

    def test_returns_none_on_structural_error(self):
        """P0-4: structural error returns None, not '{}' fake JSON."""
        from shared.exceptions import GeminiStructuralError
        optimizer_mod._KEY_POOL.generate_content.side_effect = GeminiStructuralError("bad")
        self.assertIsNone(re_score("resume", "jd"))

    def test_reraises_transient_error(self):
        from shared.exceptions import GeminiTransientError
        optimizer_mod._KEY_POOL.generate_content.side_effect = GeminiTransientError("429")
        with self.assertRaises(GeminiTransientError):
            re_score("resume", "jd")


# ─────────────────────────────────────────────────────────────────────────────
class TestBatchSchemas(unittest.TestCase):

    def test_batch_tailored_result_valid(self):
        raw = json.dumps({
            "items": [
                {"index": 0, "tailored_resume_markdown": "# R0", "optimization_summary": "s0"},
                {"index": 1, "tailored_resume_markdown": "# R1", "optimization_summary": "s1"},
            ]
        })
        result = BatchTailoredResult.model_validate_json(raw)
        self.assertEqual(len(result.items), 2)
        self.assertEqual(result.items[0].index, 0)
        self.assertIn("R0", result.items[0].tailored_resume_markdown)

    def test_batch_constants(self):
        self.assertGreater(BATCH_TAILOR_SIZE, 0)
        self.assertGreater(RESCORE_CONCURRENCY, 0)


# ─────────────────────────────────────────────────────────────────────────────
class TestBatchTailorResume(unittest.TestCase):

    def setUp(self):
        self.mock_pool = MagicMock()
        optimizer_mod._KEY_POOL = self.mock_pool

    def tearDown(self):
        optimizer_mod._KEY_POOL = None

    def test_returns_aligned_results(self):
        self.mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({
                "items": [
                    {"index": 0, "tailored_resume_markdown": "# Tailored0", "optimization_summary": "s0"},
                    {"index": 1, "tailored_resume_markdown": "# Tailored1", "optimization_summary": "s1"},
                ]
            })
        )
        results = batch_tailor_resume("resume", ["jd0", "jd1"])
        self.assertEqual(len(results), 2)
        self.assertEqual(results[0]["tailored_resume_markdown"], "# Tailored0")
        self.assertEqual(results[1]["tailored_resume_markdown"], "# Tailored1")

    def test_returns_empty_dicts_on_structural_error(self):
        """P0-4: structural error yields empty dict per JD; not fake content."""
        from shared.exceptions import GeminiStructuralError
        self.mock_pool.generate_content.side_effect = GeminiStructuralError("bad")
        results = batch_tailor_resume("resume", ["jd0", "jd1"])
        self.assertEqual(results, [{}, {}])

    def test_reraises_transient_error(self):
        from shared.exceptions import GeminiTransientError
        self.mock_pool.generate_content.side_effect = GeminiTransientError("429")
        with self.assertRaises(GeminiTransientError):
            batch_tailor_resume("resume", ["jd0", "jd1"])

    def test_missing_index_returns_empty_dict(self):
        """Gemini returns only index 1, index 0 should be empty."""
        self.mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({
                "items": [
                    {"index": 1, "tailored_resume_markdown": "# Only1", "optimization_summary": "s1"},
                ]
            })
        )
        results = batch_tailor_resume("resume", ["jd0", "jd1"])
        self.assertEqual(results[0], {})
        self.assertEqual(results[1]["tailored_resume_markdown"], "# Only1")

    def test_out_of_bounds_index_ignored(self):
        self.mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({
                "items": [
                    {"index": 0, "tailored_resume_markdown": "# T0", "optimization_summary": "s0"},
                    {"index": 99, "tailored_resume_markdown": "# Bad", "optimization_summary": "bad"},
                ]
            })
        )
        results = batch_tailor_resume("resume", ["jd0"])
        self.assertEqual(len(results), 1)
        self.assertEqual(results[0]["tailored_resume_markdown"], "# T0")


# ─────────────────────────────────────────────────────────────────────────────
class TestGeminiKeyPool(unittest.TestCase):

    def test_initial_key(self):
        pool = _GeminiKeyPool(["key1", "key2"])
        self.assertEqual(pool.current, "key1")

    def test_rotate_advances_key(self):
        pool = _GeminiKeyPool(["k1", "k2"])
        self.assertTrue(pool.rotate())
        self.assertEqual(pool.current, "k2")

    def test_rotate_returns_false_when_exhausted(self):
        pool = _GeminiKeyPool(["only_key"])
        self.assertFalse(pool.rotate())

    def test_rotates_on_429(self):
        pool = _GeminiKeyPool(["k1", "k2"], genai_mod=optimizer_mod.genai)
        mock_client = MagicMock()
        mock_client.models.generate_content.side_effect = [
            Exception("429 RESOURCE_EXHAUSTED"),
            MagicMock(text='{}'),
        ]
        with patch("agents.resume_optimizer.genai.Client", return_value=mock_client):
            pool.generate_content("model", "content", MagicMock())
        self.assertEqual(pool.current, "k2")


# ─────────────────────────────────────────────────────────────────────────────
class TestPrompts(unittest.TestCase):
    """Verify prompts contain expected content."""

    def test_tailor_prompt_has_strict_rules(self):
        self.assertIn("NEVER fabricate", _TAILOR_SYSTEM_PROMPT)
        self.assertIn("ONLY use information already in the original resume", _TAILOR_SYSTEM_PROMPT)

    def test_fine_prompt_matches_match_agent(self):
        self.assertIn("AI/ML Tech Depth (30%)", _FINE_SYSTEM_PROMPT)
        self.assertIn("TPM Function Match (30%)", _FINE_SYSTEM_PROMPT)
        self.assertIn("Industry & Domain Relevance (20%)", _FINE_SYSTEM_PROMPT)
        self.assertIn("Growth Trajectory (20%)", _FINE_SYSTEM_PROMPT)


class TestBug48ListComprehension(unittest.TestCase):
    """BUG-48: [{}]*N creates shared references. Must use list comprehension."""

    def test_batch_tailor_uses_comprehension(self):
        import inspect
        source = inspect.getsource(batch_tailor_resume)
        self.assertNotIn("[{}] *", source.replace(" ", "").replace("\t", ""),
                         "batch_tailor_resume must not use [{}]*N")

class TestBug50PrintSummaryDynamicColumns(unittest.TestCase):
    """BUG-50: _print_summary must use TAILORED_HEADERS for column lookup."""

    def test_uses_tailored_headers_not_hardcoded(self):
        """Source code should reference TAILORED_HEADERS.index(), not hardcoded col numbers."""
        import inspect
        source = inspect.getsource(_print_summary)
        self.assertIn("TAILORED_HEADERS.index", source,
                      "_print_summary must use TAILORED_HEADERS.index() for dynamic column lookup")
        # Should NOT have bare ws.cell(r, 4) style hardcoded columns for data fields
        # (col_rid, col_company etc. variables are fine)
        for hardcoded in ["ws.cell(r, 4)", "ws.cell(r, 5)", "ws.cell(r, 6)", "ws.cell(r, 7)"]:
            self.assertNotIn(hardcoded, source,
                             f"Hardcoded {hardcoded} found — must use dynamic column variable")

    def test_reads_correct_data(self):
        """_print_summary should read data from correct columns in a real Excel file."""
        from openpyxl import Workbook
        from shared.excel_store import TAILORED_HEADERS as TH
        wb = Workbook()
        ws = wb.active
        ws.title = "Tailored_Match_Results"
        ws.append(TH)
        ws.append(["resume1", "http://jd1", "TPM Lead", "Acme Corp", 72, 88, 16,
                    "/path/resume.md", "improved XYZ", "2026-01-01", "abc123"])
        ws.append(["resume1", "http://jd2", "PM Senior", "Beta Inc", 65, 80, 15,
                    "/path/resume2.md", "improved ABC", "2026-01-02", "def456"])
        ws.append(["other_id", "http://jd3", "Engineer", "Gamma", 50, 60, 10,
                    "/path/resume3.md", "improved DEF", "2026-01-03", "ghi789"])
        # Also need Tailored_Match_Results pairs to exist so _print_summary doesn't early-return
        ws2 = wb.create_sheet("Match_Results")
        from shared.excel_store import MATCH_HEADERS
        ws2.append(MATCH_HEADERS)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            xlsx_path = f.name
            wb.save(xlsx_path)
        wb.close()
        try:
            import io
            from unittest.mock import patch as _patch
            # Mock get_tailored_match_pairs to return non-empty list
            with _patch("agents.resume_optimizer.get_tailored_match_pairs",
                        return_value=[("http://jd1",), ("http://jd2",)]):
                with _patch("sys.stdout", new_callable=io.StringIO) as mock_out:
                    _print_summary(xlsx_path, "resume1")
            output = mock_out.getvalue()
            self.assertIn("Acme Corp", output)
            self.assertIn("Beta Inc", output)
            # "other_id" row should NOT appear
            self.assertNotIn("Gamma", output)
        finally:
            os.unlink(xlsx_path)


class TestRegressionFlagAssembly(unittest.TestCase):
    """Optimizer's Phase 3 assembly must include `regression = delta < 0`
    in the record dict that goes to batch_upsert_tailored_records.

    Mirrors the source pattern:
        regression = delta < 0
        results.append({..., "regression": regression})
    """

    def test_negative_delta_marks_regression_true(self):
        """When tailored < base, the assembled record must have regression=True."""
        import inspect
        src = inspect.getsource(optimizer_mod._main_inner)
        # Source-level guard: the regression key must be passed in the record dict.
        self.assertIn('"regression": regression', src,
                      "_main_inner must include regression flag in tailored record")
        self.assertIn("regression = delta < 0", src,
                      "_main_inner must compute regression from score delta")


class TestBug54KeyPoolNoneGuard(unittest.TestCase):
    """BUG-54: Functions using _KEY_POOL must raise RuntimeError when pool is None."""

    def test_tailor_resume_raises_when_pool_none(self):
        original = optimizer_mod._KEY_POOL
        try:
            optimizer_mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                tailor_resume("resume", "jd")
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            optimizer_mod._KEY_POOL = original

    def test_re_score_raises_when_pool_none(self):
        original = optimizer_mod._KEY_POOL
        try:
            optimizer_mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                re_score("tailored resume", "jd")
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            optimizer_mod._KEY_POOL = original

    def test_batch_tailor_resume_raises_when_pool_none(self):
        original = optimizer_mod._KEY_POOL
        try:
            optimizer_mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                batch_tailor_resume("resume", ["jd1", "jd2"])
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            optimizer_mod._KEY_POOL = original

    def test_re_score_raises_when_pool_none(self):
        original = optimizer_mod._KEY_POOL
        try:
            optimizer_mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                re_score("resume", "jd")
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            optimizer_mod._KEY_POOL = original


if __name__ == "__main__":
    unittest.main(verbosity=2)

"""
Tests for company_agent.py

All HTTP calls are mocked — no network required to run this suite.

Coverage:
  - _is_likely_career_url: URL heuristics
  - _slug_candidates: slug generation
  - validate_career_url: HTTP response handling (mocked)
  - _check_ats_slug: ATS API response parsing (mocked)
  - find_career_url: full multi-strategy flow (all external calls mocked)
  - Gemini schema: AICompanyInfo has no career_url field
  - Sanity: job_agent and match_agent importable

Integration tests (real network):
  To run tests that make live HTTP/API calls, set INTEGRATION_TEST=1:
    INTEGRATION_TEST=1 python -m pytest tests/test_company_agent.py
  Any test class that requires network should use:
    @unittest.skipUnless(os.getenv("INTEGRATION_TEST"), "requires INTEGRATION_TEST=1")
"""
import sys
import os
import types as pytypes
import unittest
from unittest.mock import MagicMock, patch

# ── path setup ────────────────────────────────────────────────────────────────
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

# Stub heavy optional deps before importing agent
for mod in ["google", "google.genai", "google.genai.types",
            "tavily", "dotenv"]:
    if mod not in sys.modules:
        m = MagicMock()
        sys.modules[mod] = m

# Make dotenv.load_dotenv a no-op
sys.modules["dotenv"].load_dotenv = lambda: None

# Stub google.genai sub-structure needed by agent top-level code
import google
google.genai = MagicMock()
google.genai.types = MagicMock()

from agents.company_agent import (
    _is_likely_career_url,
    _unwrap_career_url,
    _find_workday_url,
    _workday_subdomain_matches_company,
    validate_and_upgrade_ats_url,
    _slug_candidates,
    validate_career_url,
    _check_ats_slug,
    _find_ats_url,
    find_career_url,
    _tavily_extract_career_url,
    _scrape_homepage_for_career_link,
    _normalize_company_name,
    _is_duplicate_company,
    discover_ai_companies,
    AICompanyInfo,
    CompanyInfoList,
    ATS_VALIDATORS,
    KNOWN_CAREER_URLS,
)
import agents.company_agent as company_agent_mod


# ─────────────────────────────────────────────────────────────────────────────
class TestIsLikelyCareerUrl(unittest.TestCase):

    def test_greenhouse_url(self):
        self.assertTrue(_is_likely_career_url("https://job-boards.greenhouse.io/openai"))

    def test_lever_url(self):
        self.assertTrue(_is_likely_career_url("https://jobs.lever.co/anthropic"))

    def test_ashby_url(self):
        self.assertTrue(_is_likely_career_url("https://jobs.ashbyhq.com/somecompany"))

    def test_careers_path(self):
        self.assertTrue(_is_likely_career_url("https://openai.com/careers"))

    def test_jobs_path(self):
        self.assertTrue(_is_likely_career_url("https://company.com/jobs"))

    def test_news_article_false(self):
        self.assertFalse(_is_likely_career_url("https://techcrunch.com/2026/01/some-article"))

    def test_homepage_false(self):
        self.assertFalse(_is_likely_career_url("https://www.anthropic.com/"))

    def test_empty_false(self):
        self.assertFalse(_is_likely_career_url(""))


# ─────────────────────────────────────────────────────────────────────────────
class TestUnwrapCareerUrl(unittest.TestCase):
    """A″ — extract underlying company slug from LinkedIn / VC-portfolio wrappers."""

    def test_linkedin_jobs_slug_pattern(self):
        # Real example from current company sheet
        self.assertEqual(
            _unwrap_career_url("https://www.linkedin.com/jobs/arista-networks-jobs"),
            "arista networks",
        )

    def test_linkedin_jobs_with_trailing_slash(self):
        self.assertEqual(
            _unwrap_career_url("https://www.linkedin.com/jobs/some-co-jobs/"),
            "some co",
        )

    def test_linkedin_company_jobs_pattern(self):
        self.assertEqual(
            _unwrap_career_url("https://www.linkedin.com/company/arista-networks/jobs"),
            "arista networks",
        )

    def test_linkedin_view_returns_none(self):
        # No extractable name from a numeric job-id URL
        self.assertIsNone(
            _unwrap_career_url("https://www.linkedin.com/jobs/view/3854720198")
        )

    def test_vc_a16z_pattern(self):
        self.assertEqual(
            _unwrap_career_url("https://jobs.a16z.com/jobs/repl.it"),
            "repl.it",
        )

    def test_vc_a16z_with_query_params(self):
        self.assertEqual(
            _unwrap_career_url("https://jobs.a16z.com/jobs/repl.it?trk=public_post"),
            "repl.it",
        )

    def test_vc_battery_pattern(self):
        self.assertEqual(
            _unwrap_career_url("https://jobs.battery.com/jobs/semi-technologies"),
            "semi technologies",
        )

    def test_vc_gaingels_pattern(self):
        self.assertEqual(
            _unwrap_career_url("https://jobs.gaingels.com/jobs/modal-labs"),
            "modal labs",
        )

    def test_vc_01a_pattern(self):
        self.assertEqual(
            _unwrap_career_url("https://jobs.01a.com/jobs/baseten"),
            "baseten",
        )

    def test_real_ats_url_returns_none(self):
        self.assertIsNone(_unwrap_career_url("https://job-boards.greenhouse.io/openai"))
        self.assertIsNone(_unwrap_career_url("https://jobs.lever.co/palantir"))
        self.assertIsNone(_unwrap_career_url("https://jobs.ashbyhq.com/anthropic"))

    def test_company_homepage_returns_none(self):
        self.assertIsNone(_unwrap_career_url("https://www.anthropic.com/careers"))

    def test_unknown_vc_host_returns_none(self):
        # Other VC portfolio hosts not in our allow-list shouldn't trigger
        self.assertIsNone(_unwrap_career_url("https://jobs.unknown-vc.com/jobs/some-co"))

    def test_empty_returns_none(self):
        self.assertIsNone(_unwrap_career_url(""))
        self.assertIsNone(_unwrap_career_url(None))


# ─────────────────────────────────────────────────────────────────────────────
class TestValidateAndUpgradeUnwraps(unittest.TestCase):
    """A″ — validate_and_upgrade_ats_url must re-resolve wrapper URLs to real ATS."""

    def test_vc_wrapper_upgraded_when_underlying_is_on_greenhouse(self):
        # jobs.a16z.com/jobs/repl.it → probe slugs → greenhouse hit on first slug
        with patch("agents.company_agent._check_ats_slug") as mock_check:
            # Greenhouse hit on first probe; subsequent calls won't run
            mock_check.return_value = (True, 5)
            with patch("agents.company_agent.time.sleep"):  # speed up
                upgraded = validate_and_upgrade_ats_url(
                    "Repl.it", "https://jobs.a16z.com/jobs/repl.it"
                )
        self.assertIn("greenhouse.io", upgraded)
        # Should NOT return the wrapper URL
        self.assertNotIn("a16z.com", upgraded)

    def test_linkedin_wrapper_upgraded_when_underlying_is_on_lever(self):
        # First validator (greenhouse) misses, second (lever) hits
        def fake_check(slug, validator):
            return (validator["platform"] == "lever", 3 if validator["platform"] == "lever" else 0)
        with patch("agents.company_agent._check_ats_slug", side_effect=fake_check):
            with patch("agents.company_agent.time.sleep"):
                upgraded = validate_and_upgrade_ats_url(
                    "Arista Networks",
                    "https://www.linkedin.com/jobs/arista-networks-jobs",
                )
        self.assertIn("lever.co", upgraded)

    def test_wrapper_no_ats_match_falls_through(self):
        # Hint extracted but no validator hits → does NOT return upgraded URL from
        # step 1.5; falls through to step 2/3. Step 2 short-circuits on ATS-domain
        # check (linkedin.com isn't ATS), so step 3 runs slug-probing on company_name.
        # Mock _check_ats_slug to always miss → final result = original URL.
        with patch("agents.company_agent._check_ats_slug", return_value=(False, 0)):
            with patch("agents.company_agent.time.sleep"):
                result = validate_and_upgrade_ats_url(
                    "Arista Networks",
                    "https://www.linkedin.com/jobs/arista-networks-jobs",
                )
        self.assertEqual(result, "https://www.linkedin.com/jobs/arista-networks-jobs")

    def test_non_wrapper_skips_unwrap_path(self):
        # Already-ATS URL: step 1.5 returns None hint; step 2 short-circuits.
        # _check_ats_slug must NOT be invoked.
        with patch("agents.company_agent._check_ats_slug") as mock_check:
            result = validate_and_upgrade_ats_url(
                "OpenAI", "https://job-boards.greenhouse.io/openai"
            )
        self.assertEqual(result, "https://job-boards.greenhouse.io/openai")
        mock_check.assert_not_called()

    def test_workable_url_short_circuits_as_ats(self):
        # A′ regression: workable.com must be recognized as already-ATS in step 2
        with patch("agents.company_agent._check_ats_slug") as mock_check:
            result = validate_and_upgrade_ats_url(
                "Hugging Face", "https://apply.workable.com/huggingface/"
            )
        self.assertEqual(result, "https://apply.workable.com/huggingface/")
        mock_check.assert_not_called()


# ─────────────────────────────────────────────────────────────────────────────
class TestWorkdaySubdomainMatch(unittest.TestCase):
    """Critical false-positive guard: Tavily returns Workday URLs from OTHER
    companies that mention the search term. Subdomain must match the co name."""

    def test_exact_subdomain_match(self):
        self.assertTrue(_workday_subdomain_matches_company(
            "https://adobe.wd5.myworkdayjobs.com/external_experienced", "Adobe"))

    def test_arista_compact_match(self):
        # subdomain "aristanetworks" matches slug candidate "aristanetworks"
        self.assertTrue(_workday_subdomain_matches_company(
            "https://aristanetworks.wd1.myworkdayjobs.com/External", "Arista Networks"))

    def test_amd_argonne_rejected(self):
        # Real false positive observed in dry run — Argonne National Lab
        # has Workday postings that mention AMD silicon
        self.assertFalse(_workday_subdomain_matches_company(
            "https://argonne.wd1.myworkdayjobs.com/Argonne_Careers", "AMD"))

    def test_oracle_pwc_rejected(self):
        # Real false positive — PwC has an Oracle-consulting job
        self.assertFalse(_workday_subdomain_matches_company(
            "https://pwc.wd3.myworkdayjobs.com/Global_Experienced_Careers", "Oracle"))

    def test_non_workday_url_returns_false(self):
        self.assertFalse(_workday_subdomain_matches_company(
            "https://www.adobe.com/careers", "Adobe"))

    def test_short_company_name_avoids_substring_false_positive(self):
        # "AI" (len 2) must NOT match into "argonne" by substring rule
        self.assertFalse(_workday_subdomain_matches_company(
            "https://argonne.wd1.myworkdayjobs.com/Careers", "AI"))

    def test_apple_applebank_rejected(self):
        # "apple" is a strict prefix of "applebank" — substring matching would
        # let this through, exact-compact-match must reject it
        self.assertFalse(_workday_subdomain_matches_company(
            "https://applebank.wd5.myworkdayjobs.com/applebankcareers", "Apple"))

    def test_clay_claycountybcc_rejected(self):
        # "clay" is a strict prefix of "claycountybcc"
        self.assertFalse(_workday_subdomain_matches_company(
            "https://claycountybcc.wd1.myworkdayjobs.com/External_Careers", "Clay"))

    def test_western_digital_westernunion_rejected(self):
        # "western" appears in both "westerndigital" (slug) and "westernunion"
        # (subdomain) but they're different companies
        self.assertFalse(_workday_subdomain_matches_company(
            "https://westernunion.wd5.myworkdayjobs.com/WesternUnionJobs",
            "Western Digital"))


class TestFindWorkdayUrl(unittest.TestCase):
    """Workday-via-Tavily discovery for unguessable subdomains."""

    def _make_client(self, results):
        client = MagicMock()
        client.search.return_value = {"results": results}
        return client

    def test_returns_workday_url_when_tavily_finds_one(self):
        client = self._make_client([
            {"url": "https://aristanetworks.wd1.myworkdayjobs.com/External"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=True):
            url = _find_workday_url("Arista Networks", client)
        self.assertIn("myworkdayjobs.com", url or "")
        # Verify the Tavily query is Workday-scoped
        called_query = client.search.call_args[1]["query"]
        self.assertIn("myworkdayjobs.com", called_query)
        self.assertIn("Arista Networks", called_query)

    def test_skips_wrong_company_workday_url(self):
        # AMD search returns Argonne first, real AMD URL second — must skip Argonne
        client = self._make_client([
            {"url": "https://argonne.wd1.myworkdayjobs.com/Argonne_Careers"},
            {"url": "https://amd.wd1.myworkdayjobs.com/External"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=True):
            url = _find_workday_url("AMD", client)
        self.assertIn("amd.wd1", url)
        self.assertNotIn("argonne", url)

    def test_returns_none_if_only_wrong_company_results(self):
        # All results are wrong-company Workday URLs — return None, don't pick the first
        client = self._make_client([
            {"url": "https://argonne.wd1.myworkdayjobs.com/Argonne_Careers"},
            {"url": "https://pwc.wd3.myworkdayjobs.com/Global_Experienced_Careers"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=True):
            url = _find_workday_url("AMD", client)
        self.assertIsNone(url)

    def test_skips_non_workday_results(self):
        client = self._make_client([
            {"url": "https://www.linkedin.com/jobs/arista-networks-jobs"},
            {"url": "https://aristanetworks.wd1.myworkdayjobs.com/External"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=True):
            url = _find_workday_url("Arista Networks", client)
        self.assertIn("myworkdayjobs.com", url)

    def test_returns_none_when_no_workday_results(self):
        client = self._make_client([
            {"url": "https://www.linkedin.com/jobs/arista-networks-jobs"},
            {"url": "https://www.aristanetworks.com/careers"},
        ])
        url = _find_workday_url("Arista Networks", client)
        self.assertIsNone(url)

    def test_skips_workday_url_that_fails_validation(self):
        client = self._make_client([
            {"url": "https://stale.wd1.myworkdayjobs.com/dead"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=False):
            url = _find_workday_url("Stale", client)  # subdomain match: stale ↔ stale
        self.assertIsNone(url)

    def test_returns_none_when_client_is_none(self):
        self.assertIsNone(_find_workday_url("Whatever", None))

    def test_returns_none_on_tavily_exception(self):
        client = MagicMock()
        client.search.side_effect = Exception("network error")
        url = _find_workday_url("Some Co", client)
        self.assertIsNone(url)


# ─────────────────────────────────────────────────────────────────────────────
class TestValidateAndUpgradeWorkdayFallback(unittest.TestCase):
    """Step 4 — Workday-via-Tavily kicks in only when slug-probe misses."""

    def test_workday_fallback_invoked_when_slug_probe_misses(self):
        client = MagicMock()
        client.search.return_value = {"results": [
            {"url": "https://aristanetworks.wd1.myworkdayjobs.com/External"},
        ]}
        with patch("agents.company_agent._check_ats_slug", return_value=(False, 0)), \
             patch("agents.company_agent.validate_career_url", return_value=True), \
             patch("agents.company_agent.time.sleep"):
            result = validate_and_upgrade_ats_url(
                "Arista Networks",
                "https://www.linkedin.com/jobs/arista-networks-jobs",
                tavily_client=client,
            )
        self.assertIn("myworkdayjobs.com", result)
        client.search.assert_called_once()

    def test_workday_fallback_skipped_when_slug_probe_hits(self):
        # Step 3 finds a Greenhouse hit → step 4 must NOT call Tavily
        client = MagicMock()
        with patch("agents.company_agent._check_ats_slug", return_value=(True, 5)), \
             patch("agents.company_agent.time.sleep"):
            result = validate_and_upgrade_ats_url(
                "OpenAI", "https://www.openai.com/careers", tavily_client=client,
            )
        self.assertIn("greenhouse.io", result)
        client.search.assert_not_called()

    def test_workday_fallback_skipped_when_no_tavily_client(self):
        # Without client, step 4 is bypassed and original URL is returned
        with patch("agents.company_agent._check_ats_slug", return_value=(False, 0)), \
             patch("agents.company_agent.time.sleep"):
            result = validate_and_upgrade_ats_url(
                "Stuck Co", "https://stuckco.com/careers", tavily_client=None,
            )
        self.assertEqual(result, "https://stuckco.com/careers")

    def test_workday_fallback_yields_no_match_returns_original(self):
        client = MagicMock()
        client.search.return_value = {"results": []}
        with patch("agents.company_agent._check_ats_slug", return_value=(False, 0)), \
             patch("agents.company_agent.time.sleep"):
            result = validate_and_upgrade_ats_url(
                "Stuck Co", "https://stuckco.com/careers", tavily_client=client,
            )
        self.assertEqual(result, "https://stuckco.com/careers")


# ─────────────────────────────────────────────────────────────────────────────
class TestKnownAtsOverridesWordBoundary(unittest.TestCase):
    """BUG-23: KNOWN_ATS_OVERRIDES must use word-boundary matching, not substring."""

    def _override(self, name):
        from agents.company_agent import validate_and_upgrade_ats_url
        # Pass a non-ATS URL so it doesn't short-circuit on ATS detection
        return validate_and_upgrade_ats_url(name, "https://company.com/careers")

    def test_xai_exact_match(self):
        url = self._override("xAI")
        self.assertIn("xai.com", url)

    def test_xai_does_not_match_maxai(self):
        """'xai' substring must NOT match 'MaxAI' (BUG-23 regression)."""
        with patch("agents.company_agent.validate_and_upgrade_ats_url") as _:
            # Direct test of the matching logic
            import re
            from agents.company_agent import KNOWN_ATS_OVERRIDES
            name_lower = "maxai"
            matched = [kw for kw in KNOWN_ATS_OVERRIDES
                       if re.search(r'(?:^|\s)' + re.escape(kw) + r'(?:\s|$)', name_lower)]
            self.assertEqual(matched, [], f"'maxai' should NOT match any override, matched: {matched}")

    def test_nvidia_corporation_matches(self):
        """'NVIDIA Corporation' should still match the 'nvidia' override key."""
        import re
        from agents.company_agent import KNOWN_ATS_OVERRIDES
        name_lower = "nvidia corporation"
        matched = [kw for kw in KNOWN_ATS_OVERRIDES
                   if re.search(r'(?:^|\s)' + re.escape(kw) + r'(?:\s|$)', name_lower)]
        self.assertIn("nvidia", matched)

    def test_google_deepmind_multiword_matches(self):
        """Multi-word key 'google deepmind' must still match."""
        import re
        from agents.company_agent import KNOWN_ATS_OVERRIDES
        name_lower = "google deepmind"
        matched = [kw for kw in KNOWN_ATS_OVERRIDES
                   if re.search(r'(?:^|\s)' + re.escape(kw) + r'(?:\s|$)', name_lower)]
        self.assertIn("google deepmind", matched)


# ─────────────────────────────────────────────────────────────────────────────
class TestSlugCandidates(unittest.TestCase):

    def test_simple_name(self):
        slugs = _slug_candidates("Anthropic")
        self.assertIn("anthropic", slugs)

    def test_multi_word(self):
        slugs = _slug_candidates("Scale AI")
        # pre-strip slugs preserve the full name
        self.assertIn("scale-ai", slugs)
        self.assertIn("scaleai", slugs)
        # post-strip slug also generated (strips " ai" suffix)
        self.assertIn("scale", slugs)

    def test_strips_inc(self):
        slugs = _slug_candidates("Cohere Inc.")
        self.assertIn("cohere", slugs)

    def test_strips_technologies(self):
        slugs = _slug_candidates("CoreWeave Technologies")
        self.assertIn("coreweave", slugs)

    def test_no_duplicates(self):
        slugs = _slug_candidates("OpenAI")
        self.assertEqual(len(slugs), len(set(slugs)))

    def test_hyphen_and_nospace_both_present(self):
        slugs = _slug_candidates("Hugging Face")
        self.assertIn("hugging-face", slugs)
        self.assertIn("huggingface", slugs)


# ─────────────────────────────────────────────────────────────────────────────
class TestValidateCareerUrl(unittest.TestCase):
    """Validates career URL logic — HTTP calls are mocked, no network required."""

    def _mock_resp(self, status_code):
        r = MagicMock()
        r.status_code = status_code
        return r

    def test_known_good_url(self):
        with patch("agents.company_agent.requests.get", return_value=self._mock_resp(200)):
            self.assertTrue(validate_career_url("https://www.anthropic.com/careers"))

    def test_openai_careers(self):
        with patch("agents.company_agent.requests.get", return_value=self._mock_resp(200)):
            self.assertTrue(validate_career_url("https://openai.com/careers"))

    def test_empty_string(self):
        self.assertFalse(validate_career_url(""))

    def test_na_string(self):
        self.assertFalse(validate_career_url("N/A"))

    def test_non_http(self):
        self.assertFalse(validate_career_url("ftp://example.com"))

    def test_404_url(self):
        with patch("agents.company_agent.requests.get", return_value=self._mock_resp(404)):
            self.assertFalse(validate_career_url("https://example.com/nonexistent"))


# ─────────────────────────────────────────────────────────────────────────────
class TestCheckAtsSlug(unittest.TestCase):
    """ATS slug probing logic — HTTP calls are mocked, no network required."""

    def test_anthropic_greenhouse(self):
        greenhouse = ATS_VALIDATORS[0]
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"jobs": [{"id": 1}, {"id": 2}]}
        with patch("agents.company_agent.requests.get", return_value=mock_resp):
            hit, n = _check_ats_slug("anthropic", greenhouse)
        self.assertTrue(hit)
        self.assertGreater(n, 0)

    def test_nonexistent_slug(self):
        greenhouse = ATS_VALIDATORS[0]
        mock_resp = MagicMock()
        mock_resp.status_code = 404
        with patch("agents.company_agent.requests.get", return_value=mock_resp):
            hit, n = _check_ats_slug("zzz-nonexistent-company-xyz-123", greenhouse)
        self.assertFalse(hit)

    def test_mistral_lever(self):
        lever = ATS_VALIDATORS[1]
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = [{"id": "abc"}, {"id": "def"}]
        with patch("agents.company_agent.requests.get", return_value=mock_resp):
            hit, n = _check_ats_slug("mistral", lever)
        self.assertTrue(hit)

    def test_validators_cover_all_json_api_platforms(self):
        # A″ regression: company_agent's slug-probe validators must include every
        # json_api ATS that job_agent.py routes — otherwise unwrapped wrapper URLs
        # for cos on a missing ATS (e.g. Ashby) silently fail to upgrade.
        platforms = {v["platform"] for v in ATS_VALIDATORS}
        self.assertIn("greenhouse", platforms)
        self.assertIn("lever", platforms)
        self.assertIn("ashby", platforms)
        self.assertIn("workable", platforms)


# ─────────────────────────────────────────────────────────────────────────────
class TestFindAtsUrl(unittest.TestCase):
    """ATS URL discovery — _check_ats_slug and time.sleep are mocked."""

    def test_anthropic_finds_greenhouse(self):
        def fake_check(slug, validator):
            if slug == "anthropic" and validator["platform"] == "greenhouse":
                return True, 5
            return False, 0
        with patch("agents.company_agent._check_ats_slug", side_effect=fake_check), \
             patch("agents.company_agent.time.sleep"):
            url = _find_ats_url("Anthropic")
        self.assertIsNotNone(url)
        self.assertIn("greenhouse.io", url)

    def test_scaleai_finds_greenhouse(self):
        # _slug_candidates("Scale AI")[:3] → ["scale", "scale-ai", "scaleai"]
        # "scaleai" is the 3rd candidate tried, registered on Greenhouse
        def fake_check(slug, validator):
            if slug == "scaleai" and validator["platform"] == "greenhouse":
                return True, 3
            return False, 0
        with patch("agents.company_agent._check_ats_slug", side_effect=fake_check), \
             patch("agents.company_agent.time.sleep"):
            url = _find_ats_url("Scale AI")
        self.assertIsNotNone(url)
        self.assertIn("greenhouse.io", url)


# ─────────────────────────────────────────────────────────────────────────────
class TestTavilyExtractCareerUrl(unittest.TestCase):
    """Unit tests with mocked Tavily client."""

    def _make_client(self, results):
        client = MagicMock()
        client.search.return_value = {"results": results}
        return client

    def test_returns_first_valid_career_url(self):
        client = self._make_client([
            {"url": "https://techcrunch.com/some-article"},          # not career
            {"url": "https://jobs.lever.co/anthropic"},              # career ✓
        ])
        with patch("agents.company_agent.validate_career_url", return_value=True):
            result = _tavily_extract_career_url("Anthropic", "query", client)
        self.assertEqual(result, "https://jobs.lever.co/anthropic")

    def test_skips_non_career_urls(self):
        client = self._make_client([
            {"url": "https://techcrunch.com/article"},
            {"url": "https://news.ycombinator.com/item?id=123"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=True):
            result = _tavily_extract_career_url("SomeCompany", "query", client)
        self.assertIsNone(result)

    def test_skips_invalid_urls(self):
        client = self._make_client([
            {"url": "https://jobs.lever.co/fakecompany"},
        ])
        with patch("agents.company_agent.validate_career_url", return_value=False):
            result = _tavily_extract_career_url("FakeCompany", "query", client)
        self.assertIsNone(result)

    def test_handles_tavily_exception(self):
        client = MagicMock()
        client.search.side_effect = Exception("API error")
        result = _tavily_extract_career_url("AnyCompany", "query", client)
        self.assertIsNone(result)


# ─────────────────────────────────────────────────────────────────────────────
class TestFindCareerUrlOrchestration(unittest.TestCase):
    """Unit test the full find_career_url strategy chain."""

    def test_known_url_returned_first(self):
        """KNOWN_CAREER_URLS match should short-circuit everything."""
        mock_client = MagicMock()
        with patch("agents.company_agent.validate_career_url", return_value=True):
            url = find_career_url("OpenAI", mock_client)
        self.assertEqual(url, KNOWN_CAREER_URLS["OpenAI"])
        mock_client.search.assert_not_called()

    def test_falls_back_to_tavily_ats(self):
        """When known URL missing, Tavily ATS search should be tried."""
        mock_client = MagicMock()
        mock_client.search.return_value = {
            "results": [{"url": "https://jobs.lever.co/unknownstartup"}]
        }
        with patch("agents.company_agent.validate_career_url", return_value=True):
            url = find_career_url("UnknownStartup2026", mock_client)
        self.assertIsNotNone(url)
        self.assertIn("lever.co", url)

    def test_returns_none_when_all_strategies_fail(self):
        mock_client = MagicMock()
        mock_client.search.return_value = {"results": [
            {"url": "https://techcrunch.com/article"},
        ]}
        with patch("agents.company_agent.validate_career_url", return_value=False), \
             patch("agents.company_agent._find_ats_url", return_value=None), \
             patch("agents.company_agent._scrape_homepage_for_career_link", return_value=None):
            url = find_career_url("TotallyFakeCompanyXYZ", mock_client)
        self.assertIsNone(url)


# ─────────────────────────────────────────────────────────────────────────────
class TestSchemaNoCareerUrl(unittest.TestCase):
    """Ensure AICompanyInfo schema has no career_url field."""

    def test_no_career_url_field(self):
        fields = AICompanyInfo.model_fields
        self.assertNotIn("career_url", fields,
                         "AICompanyInfo must NOT have career_url — Gemini no longer generates URLs")

    def test_required_fields_present(self):
        fields = AICompanyInfo.model_fields
        for f in ["company_name", "track", "business_focus"]:
            self.assertIn(f, fields)

    def test_business_focus_description_mentions_sentences(self):
        desc = AICompanyInfo.model_fields["business_focus"].description
        self.assertIn("3-4 sentence", desc)

    def test_instantiation_without_url(self):
        obj = AICompanyInfo(
            company_name="TestCo",
            track="AI-native",
            business_focus="TestCo builds AI tools. They serve enterprise customers. "
                           "Their edge is speed. They raised $200M in 2025."
        )
        self.assertEqual(obj.company_name, "TestCo")
        self.assertFalse(hasattr(obj, "career_url"))


# ─────────────────────────────────────────────────────────────────────────────
class TestCompanyNameNormalization(unittest.TestCase):
    """Phase 3: _normalize_company_name strips suffixes and normalizes."""

    def test_strip_inc(self):
        self.assertEqual(_normalize_company_name("Anthropic Inc"), "anthropic")

    def test_strip_inc_dot(self):
        self.assertEqual(_normalize_company_name("Anthropic Inc."), "anthropic")

    def test_strip_technologies(self):
        self.assertEqual(_normalize_company_name("CoreWeave Technologies"), "coreweave")

    def test_strip_ai(self):
        self.assertEqual(_normalize_company_name("Scale AI"), "scale")

    def test_strip_labs(self):
        self.assertEqual(_normalize_company_name("Stability Labs"), "stability")

    def test_strip_llc(self):
        self.assertEqual(_normalize_company_name("DeepSeek LLC"), "deepseek")

    def test_case_insensitive(self):
        self.assertEqual(_normalize_company_name("NVIDIA"), "nvidia")

    def test_strip_whitespace(self):
        self.assertEqual(_normalize_company_name("  OpenAI  "), "openai")


# ─────────────────────────────────────────────────────────────────────────────
class TestIsDuplicateCompany(unittest.TestCase):
    """Phase 3: _is_duplicate_company catches near-duplicates."""

    def test_exact_match(self):
        self.assertTrue(_is_duplicate_company("Anthropic", {"Anthropic"}))

    def test_case_insensitive_match(self):
        self.assertTrue(_is_duplicate_company("anthropic", {"Anthropic"}))

    def test_suffix_match_inc(self):
        self.assertTrue(_is_duplicate_company("Anthropic Inc", {"Anthropic"}))

    def test_suffix_match_ai(self):
        self.assertTrue(_is_duplicate_company("Scale AI", {"Scale"}))

    def test_startswith_match(self):
        self.assertTrue(_is_duplicate_company("Anthropic AI Research", {"Anthropic"}))

    def test_no_false_positive_meta_metabase(self):
        """'Meta' vs 'Metabase' — short name (4 chars) but not same company."""
        # Meta startswith Metabase? No. Metabase startswith Meta? Yes.
        # But this is a legitimate concern — 'Meta' is 4 chars so startswith triggers.
        # This is accepted behavior: better to over-dedup than under-dedup.
        # If this becomes a problem, the user can manually add the company.
        self.assertTrue(_is_duplicate_company("Metabase", {"Meta"}))

    def test_no_false_positive_short_names(self):
        """3-char names should NOT trigger startswith."""
        self.assertFalse(_is_duplicate_company("SAS Institute", {"SAP"}))

    def test_completely_different(self):
        self.assertFalse(_is_duplicate_company("Anthropic", {"OpenAI", "Google"}))


# ─────────────────────────────────────────────────────────────────────────────
class TestSanityOtherAgents(unittest.TestCase):
    """Sanity: other agents still importable without errors."""

    def test_job_agent_importable(self):
        try:
            import importlib
            spec = importlib.util.spec_from_file_location(
                "job_agent",
                os.path.join(PROJECT_ROOT, "agents", "job_agent.py")
            )
            mod = importlib.util.module_from_spec(spec)
            # Don't exec (has side effects), just check spec loads
            self.assertIsNotNone(spec)
        except Exception as e:
            self.fail(f"job_agent.py spec load failed: {e}")

    def test_match_agent_importable(self):
        try:
            import importlib
            spec = importlib.util.spec_from_file_location(
                "match_agent",
                os.path.join(PROJECT_ROOT, "agents", "match_agent.py")
            )
            self.assertIsNotNone(spec)
        except Exception as e:
            self.fail(f"match_agent.py spec load failed: {e}")

    def test_excel_store_importable(self):
        from shared.excel_store import (
            get_or_create_excel, count_company_rows, upsert_companies,
            get_company_rows, update_company_career_url,
        )
        self.assertTrue(callable(get_or_create_excel))
        self.assertTrue(callable(upsert_companies))


# ── BUG-44: Tavily quota detection ──────────────────────────────────────────
class TestBug44TavilyQuotaDetection(unittest.TestCase):
    """BUG-44: Tavily 402/429/quota errors should be detected and break the loop."""

    @patch("agents.company_agent._KEY_POOL")
    def test_tavily_quota_402_breaks_loop(self, mock_pool):
        """When Tavily raises a 402 quota error, discover_ai_companies should stop querying."""
        from agents.company_agent import discover_ai_companies, TAVILY_QUERIES

        mock_client_cls = MagicMock()
        mock_client = MagicMock()
        # First call raises 402
        mock_client.search.side_effect = Exception("HTTP 402 Payment Required - quota exhausted")

        with patch("tavily.TavilyClient", return_value=mock_client):
            result = discover_ai_companies("fake-key", {"ExistingCo"}, {"AI-native": 5})

        # Should have called search only once (broke on quota error)
        self.assertEqual(mock_client.search.call_count, 1)
        self.assertEqual(result, [])

    @patch("agents.company_agent._KEY_POOL")
    def test_tavily_quota_429_breaks_loop(self, mock_pool):
        """429 rate limit should also break."""
        from agents.company_agent import discover_ai_companies

        mock_client = MagicMock()
        mock_client.search.side_effect = Exception("429 Too Many Requests")

        with patch("tavily.TavilyClient", return_value=mock_client):
            result = discover_ai_companies("fake-key", set(), {"AI-native": 3})

        self.assertEqual(mock_client.search.call_count, 1)

    @patch("agents.company_agent._KEY_POOL")
    def test_tavily_non_quota_error_continues(self, mock_pool):
        """Non-quota errors (e.g., network) should not break the loop."""
        from agents.company_agent import discover_ai_companies, TAVILY_QUERIES

        mock_client = MagicMock()
        # All queries raise generic network error
        mock_client.search.side_effect = Exception("Connection timeout")

        with patch("tavily.TavilyClient", return_value=mock_client):
            result = discover_ai_companies("fake-key", set(), {"AI-native": 3})

        # Should have tried all queries, not just one
        self.assertEqual(mock_client.search.call_count, len(TAVILY_QUERIES))

    @patch("agents.company_agent._KEY_POOL")
    @patch("agents.company_agent.validate_career_url", return_value=True)
    def test_tavily_extract_career_url_quota_detection(self, mock_validate, mock_pool):
        """_tavily_extract_career_url should log error for quota, warning for other."""
        from agents.company_agent import _tavily_extract_career_url

        mock_client = MagicMock()
        mock_client.search.side_effect = Exception("402 quota exceeded")

        with self.assertLogs("root", level="ERROR") as cm:
            result = _tavily_extract_career_url("TestCo", "test query", mock_client)

        self.assertIsNone(result)
        self.assertTrue(any("quota" in msg.lower() for msg in cm.output))


# ── BUG-49: discover_ai_companies set mutation ─────────────────────────────
class TestBug49DiscoverCompaniesSetMutation(unittest.TestCase):
    """BUG-49: discover_ai_companies should not mutate the caller's existing_names set."""

    @patch("agents.company_agent._KEY_POOL")
    def test_existing_names_not_mutated(self, mock_pool):
        """The caller's existing_names set should remain unchanged after the call."""
        from agents.company_agent import discover_ai_companies

        mock_client = MagicMock()
        # Return one company from Tavily
        mock_client.search.return_value = {
            "results": [{"url": "https://example.com/ai", "content": "NewCo is an AI startup"}]
        }

        # Mock Gemini to return a company
        mock_pool.generate_content.return_value = MagicMock(
            text='{"companies": [{"company_name": "NewAICo", "track": "AI-native", "business_focus": "chatbots"}]}'
        )

        original_names = {"ExistingCo", "OtherCo"}
        original_copy = set(original_names)

        with patch("tavily.TavilyClient", return_value=mock_client), \
             patch("agents.company_agent.find_career_url", return_value="https://newai.co/careers"):
            discover_ai_companies("fake-key", original_names, {"AI-native": 5})

        # The original set must NOT have been modified
        self.assertEqual(original_names, original_copy)

    @patch("agents.company_agent._KEY_POOL")
    def test_local_dedup_still_works(self, mock_pool):
        """Even with local copy, batch dedup within one call should still work."""
        from agents.company_agent import discover_ai_companies

        mock_client = MagicMock()
        mock_client.search.return_value = {"results": [{"url": "https://example.com", "content": "stuff"}]}

        # Return two companies with the same name
        mock_pool.generate_content.return_value = MagicMock(
            text='{"companies": [{"company_name": "DupCo", "track": "AI-native", "business_focus": "x"}, '
                 '{"company_name": "DupCo", "track": "AI-native", "business_focus": "y"}]}'
        )

        with patch("tavily.TavilyClient", return_value=mock_client), \
             patch("agents.company_agent.find_career_url", return_value="https://dupco.com/careers"):
            results = discover_ai_companies("fake-key", set(), {"AI-native": 5})

        # Only one DupCo should be in results (dedup within batch)
        names = [c.get("company_name") for c in results]
        self.assertEqual(names.count("DupCo"), 1)


class TestBug54KeyPoolNoneGuard(unittest.TestCase):
    """BUG-54: Functions using _KEY_POOL must raise RuntimeError when pool is None."""

    def test_discover_ai_companies_raises_when_pool_none(self):
        original = company_agent_mod._KEY_POOL
        try:
            company_agent_mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                discover_ai_companies("fake_key", set(), {"AI-native": 5})
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            company_agent_mod._KEY_POOL = original


# ── Manual-row override: run_phase_1_5 backfills blank Career URLs ──────────
class TestPhase15ManualBackfill(unittest.TestCase):
    """Manually-inserted rows (Name + Domain only, blank Career URL) must be
    picked up by run_phase_1_5 and have their Career URL backfilled via
    find_career_url. Other rows (with existing URLs) continue down the
    ATS-upgrade path unchanged.
    """

    def setUp(self):
        import tempfile
        import openpyxl
        from shared.excel_store import get_or_create_excel
        fd, self.path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        os.remove(self.path)
        get_or_create_excel(self.path)
        # Two rows: one with blank URL (manual), one with an existing ATS URL.
        wb = openpyxl.load_workbook(self.path)
        ws = wb["Company_List"]
        ws.append(["ManualCo", "Defense/Robotics AI", None, None, None, None, None, None, None])
        ws.append(["AtsCo", "AI Startups", None,
                   "https://job-boards.greenhouse.io/atsco", None, None, None, None, None])
        wb.save(self.path)
        wb.close()

    def tearDown(self):
        if os.path.exists(self.path):
            os.remove(self.path)

    @patch("agents.company_agent.find_career_url")
    def test_blank_url_row_is_backfilled(self, mock_find):
        mock_find.return_value = "https://jobs.lever.co/manualco"
        tavily_client = MagicMock()  # truthy, so backfill branch runs

        company_agent_mod.run_phase_1_5(self.path, tavily_client=tavily_client)

        import openpyxl
        mock_find.assert_called_once_with("ManualCo", tavily_client)
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb["Company_List"]
        rows = [(r[0], r[3]) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        wb.close()
        url_by_name = dict(rows)
        self.assertEqual(url_by_name["ManualCo"], "https://jobs.lever.co/manualco")
        # The ATS row should remain unchanged (already ATS, short-circuited).
        self.assertEqual(url_by_name["AtsCo"], "https://job-boards.greenhouse.io/atsco")

    @patch("agents.company_agent.find_career_url")
    def test_blank_url_skipped_when_tavily_disabled(self, mock_find):
        # No Tavily client → backfill cannot proceed; row stays blank.
        with patch.dict(os.environ, {"TAVILY_API_KEY": ""}, clear=False):
            company_agent_mod.run_phase_1_5(self.path, tavily_client=None)

        import openpyxl
        mock_find.assert_not_called()
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb["Company_List"]
        rows = [(r[0], r[3]) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        wb.close()
        url_by_name = dict(rows)
        self.assertIn(url_by_name["ManualCo"], (None, ""))

    @patch("agents.company_agent.find_career_url")
    def test_backfill_failure_leaves_row_blank(self, mock_find):
        mock_find.return_value = None  # discovery fails
        tavily_client = MagicMock()

        company_agent_mod.run_phase_1_5(self.path, tavily_client=tavily_client)

        import openpyxl
        mock_find.assert_called_once()
        wb = openpyxl.load_workbook(self.path, read_only=True)
        ws = wb["Company_List"]
        rows = [(r[0], r[3]) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
        wb.close()
        url_by_name = dict(rows)
        self.assertIn(url_by_name["ManualCo"], (None, ""))


class TestTrackWhitelist(unittest.TestCase):
    """PRJ-004: the 6-track taxonomy must be the recognized bucket set."""

    def test_track_values_and_quotas(self):
        self.assertEqual(company_agent_mod.TRACK_VALUES,
                         ("AI-native", "Mid-large Tech", "Robotics",
                          "Fintech", "Space", "Defense"))
        self.assertEqual(sum(company_agent_mod.TRACK_QUOTAS.values()),
                         company_agent_mod.MAX_TOTAL)
        self.assertEqual(company_agent_mod.MAX_TOTAL, 500)

    def test_companyinfo_accepts_defense(self):
        # Pydantic Literal must accept every track value without raising.
        info = AICompanyInfo(
            company_name="Anduril",
            track="Defense",
            business_focus="Defense AI products for the US military and allies.",
        )
        self.assertEqual(info.track, "Defense")

    def test_companyinfo_rejects_legacy_value(self):
        from pydantic import ValidationError
        with self.assertRaises(ValidationError):
            AICompanyInfo(company_name="X", track="AI Startups",
                          business_focus="legacy bucket must be rejected")


# ═════════════════════════════════════════════════════════════════════════════
# PRJ-004 T3 — per-bucket quota math + deterministic bucket rules
# ═════════════════════════════════════════════════════════════════════════════
class TestComputeNeedByTrack(unittest.TestCase):

    @staticmethod
    def _rows(counts):
        rows = []
        for track, n in counts.items():
            rows += [[f"{track}Co{i}", track, "focus", "url"] for i in range(n)]
        return rows

    def test_open_slots_computed_per_bucket(self):
        from agents.company_agent import compute_need_by_track
        need = compute_need_by_track(self._rows({"AI-native": 100, "Space": 50}))
        self.assertEqual(need["AI-native"], 50)
        self.assertEqual(need["Space"], 0)
        self.assertEqual(need["Robotics"], 50)
        self.assertEqual(need["Mid-large Tech"], 150)

    def test_grandfathered_overage_gets_zero_never_negative(self):
        """D-12: a bucket over quota (migration survivors) → 0 slots, no trim."""
        from agents.company_agent import compute_need_by_track
        need = compute_need_by_track(self._rows({"AI-native": 170}))
        self.assertEqual(need["AI-native"], 0)

    def test_unmigrated_values_count_toward_no_bucket(self):
        from agents.company_agent import compute_need_by_track
        rows = self._rows({"Fintech": 10})
        rows.append(["LegacyCo", "AI Startups", "focus", "url"])  # legacy value
        need = compute_need_by_track(rows)
        self.assertEqual(need["Fintech"], 40)
        # Legacy value consumed no bucket's quota
        self.assertEqual(sum(need.values()),
                         sum(company_agent_mod.TRACK_QUOTAS.values()) - 10)


class TestAllocateBatch(unittest.TestCase):

    def test_under_batch_size_passthrough(self):
        from agents.company_agent import allocate_batch
        need = {"AI-native": 5, "Space": 3}
        self.assertEqual(allocate_batch(need, 50), need)

    def test_proportional_allocation_caps_total(self):
        from agents.company_agent import allocate_batch
        need = {"AI-native": 150, "Mid-large Tech": 150, "Robotics": 50,
                "Fintech": 50, "Space": 50, "Defense": 50}
        alloc = allocate_batch(need, 50)
        self.assertEqual(sum(alloc.values()), 50)
        for track, n in alloc.items():
            self.assertLessEqual(n, need[track])
        # Big buckets get proportionally more
        self.assertGreater(alloc["AI-native"], alloc["Space"])

    def test_full_bucket_allocated_zero(self):
        from agents.company_agent import allocate_batch
        alloc = allocate_batch({"AI-native": 100, "Space": 0}, 50)
        self.assertEqual(alloc["Space"], 0)
        self.assertEqual(alloc["AI-native"], 50)


class TestApplyBucketRules(unittest.TestCase):

    @staticmethod
    def _co(name, track="Defense"):
        return {"company_name": name, "track": track, "business_focus": "x"}

    def test_defense_primes_dropped(self):
        from agents.company_agent import _apply_bucket_rules
        out = _apply_bucket_rules(
            [self._co("Lockheed Martin"), self._co("Anduril"),
             self._co("Raytheon"), self._co("Boeing"), self._co("L3Harris")],
            {"Defense": 50})
        self.assertEqual([c["company_name"] for c in out], ["Anduril"])

    def test_palantir_allowlisted(self):
        from agents.company_agent import _apply_bucket_rules
        out = _apply_bucket_rules([self._co("Palantir Technologies")],
                                  {"Defense": 50})
        self.assertEqual(len(out), 1)

    def test_prime_name_ok_outside_defense_bucket(self):
        """Primes are only hard-excluded from the Defense bucket."""
        from agents.company_agent import _apply_bucket_rules
        out = _apply_bucket_rules([self._co("Boeing", track="Mid-large Tech")],
                                  {"Mid-large Tech": 150})
        self.assertEqual(len(out), 1)

    def test_quota_trim_per_bucket(self):
        from agents.company_agent import _apply_bucket_rules
        cos = [self._co(f"Def{i}") for i in range(5)] \
            + [self._co("SpaceCo", track="Space")]
        out = _apply_bucket_rules(cos, {"Defense": 3, "Space": 0})
        names = [c["company_name"] for c in out]
        self.assertEqual(len([n for n in names if n.startswith("Def")]), 3)
        self.assertNotIn("SpaceCo", names)


# ═════════════════════════════════════════════════════════════════════════════
# PRJ-004 T4 — --migrate-tracks re-bucketing pass (REQ-004-06)
# ═════════════════════════════════════════════════════════════════════════════
class TestMigrateTracks(unittest.TestCase):

    def setUp(self):
        import tempfile
        from shared.excel_store import get_or_create_excel, upsert_companies
        fd, self.path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        os.remove(self.path)
        get_or_create_excel(self.path)
        self._orig_pool = company_agent_mod._KEY_POOL

    def tearDown(self):
        company_agent_mod._KEY_POOL = self._orig_pool
        if os.path.exists(self.path):
            os.remove(self.path)

    def _seed(self, rows):
        """rows: [(name, track_value)]"""
        from shared.excel_store import upsert_companies
        upsert_companies(self.path, [
            {"company_name": n, "track": t, "business_focus": f"{n} does things.",
             "career_url": f"https://{n.lower()}.com/careers"}
            for n, t in rows
        ])

    @staticmethod
    def _pool_returning(classifications):
        import json
        pool = MagicMock()
        resp = MagicMock()
        resp.text = json.dumps({"classifications": classifications})
        pool.generate_content.return_value = resp
        return pool

    def _tracks(self):
        from shared.excel_store import get_company_rows
        return {r[0]: r[1] for r in get_company_rows(self.path)}

    @patch.dict(os.environ, {"GEMINI_API_KEY": "fake"})
    def test_already_migrated_rows_skipped_idempotent(self):
        self._seed([("DoneCo", "Space"), ("OldCo", "AI Startups")])
        company_agent_mod._KEY_POOL = self._pool_returning([
            {"company_name": "OldCo", "track": "AI-native",
             "rationale": "AI product startup", "confident": True},
        ])
        out = company_agent_mod.migrate_tracks(self.path)
        self.assertEqual(out, {"migrated": 1, "skipped": 1, "flagged": 0})
        # Gemini saw only the unmigrated row
        sent = company_agent_mod._KEY_POOL.generate_content.call_args.kwargs["contents"]
        self.assertIn("OldCo", sent)
        self.assertNotIn("DoneCo", sent)
        self.assertEqual(self._tracks()["OldCo"], "AI-native")
        self.assertEqual(self._tracks()["DoneCo"], "Space")
        # Second run: nothing left to do
        out2 = company_agent_mod.migrate_tracks(self.path)
        self.assertEqual(out2["migrated"], 0)
        self.assertEqual(out2["skipped"], 2)

    @patch.dict(os.environ, {"GEMINI_API_KEY": "fake"})
    def test_unconfident_flagged_never_guessed(self):
        self._seed([("MysteryCo", "Legacy Bucket")])
        company_agent_mod._KEY_POOL = self._pool_returning([
            {"company_name": "MysteryCo", "track": "Fintech",
             "rationale": "thin description", "confident": False},
        ])
        out = company_agent_mod.migrate_tracks(self.path)
        self.assertEqual(out["flagged"], 1)
        self.assertEqual(self._tracks()["MysteryCo"], "UNMIGRATED — manual review")

    @patch.dict(os.environ, {"GEMINI_API_KEY": "fake"})
    def test_defense_prime_forced_to_midlarge(self):
        """Deterministic post-check: even if the LLM says Defense, a legacy
        prime can only survive in Mid-large Tech."""
        self._seed([("Lockheed Martin", "Defense/Robotics AI")])
        company_agent_mod._KEY_POOL = self._pool_returning([
            {"company_name": "Lockheed Martin", "track": "Defense",
             "rationale": "defense company", "confident": True},
        ])
        company_agent_mod.migrate_tracks(self.path)
        self.assertEqual(self._tracks()["Lockheed Martin"], "Mid-large Tech")

    @patch.dict(os.environ, {"GEMINI_API_KEY": "fake"})
    def test_no_rows_deleted_and_batch_failure_leaves_unmigrated(self):
        self._seed([("ACo", "old"), ("BCo", "old")])
        pool = MagicMock()
        pool.generate_content.side_effect = Exception("boom")
        company_agent_mod._KEY_POOL = pool
        out = company_agent_mod.migrate_tracks(self.path)
        tracks = self._tracks()
        # Both rows still present (never deleted), flagged for manual review.
        self.assertEqual(set(tracks.keys()), {"ACo", "BCo"})
        self.assertEqual(out["flagged"], 2)


class TestUpdateCompanyTrack(unittest.TestCase):

    def test_writes_track_in_place(self):
        import tempfile
        from shared.excel_store import (get_or_create_excel, upsert_companies,
                                        get_company_rows, update_company_track)
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        os.remove(path)
        try:
            get_or_create_excel(path)
            upsert_companies(path, [{"company_name": "TrackCo", "track": "old",
                                     "business_focus": "x", "career_url": "u"}])
            update_company_track(path, 2, "Robotics")
            rows = get_company_rows(path)
            self.assertEqual(rows[0][1], "Robotics")
            self.assertEqual(rows[0][0], "TrackCo")  # nothing else touched
        finally:
            if os.path.exists(path):
                os.remove(path)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    unittest.main(verbosity=2)

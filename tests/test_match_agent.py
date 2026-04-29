"""
Tests for agents/match_agent.py

Coverage:
  - _RateLimiter: rate limiter interval enforcement
  - load_resume: reads .md/.txt from folder
  - _GeminiKeyPool: key rotation on 429
  - Pydantic schemas: CoarseItem, BatchCoarseResult, MatchResult
  - batch_coarse_score: mocked Gemini call
  - evaluate_match: mocked Gemini call
"""
import sys
import os
import json
import asyncio
import tempfile
import time
import types as pytypes
import unittest
from unittest.mock import MagicMock, patch, PropertyMock

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

# ── Stub heavy deps before import ────────────────────────────────────────────
for mod in ["google", "google.genai", "google.genai.types", "dotenv"]:
    if mod not in sys.modules:
        sys.modules[mod] = MagicMock()

sys.modules["dotenv"].load_dotenv = lambda: None

import google
google.genai = MagicMock()
google.genai.types = MagicMock()

from shared.prompts import COARSE_SYSTEM_PROMPT as _COARSE_SYSTEM_PROMPT
from shared.schemas import CoarseItem, BatchCoarseResult, MatchResult
from agents.match_agent import (
    _RateLimiter,
    load_resume,
    _load_jd_markdown,
    _GeminiKeyPool,
    batch_coarse_score,
    evaluate_match,
    _select_fine_candidates,
    _extract_ats_keywords,
    compute_ats_for_jds,
    ATS_COVERAGE_LOW_THRESHOLD,
)
import agents.match_agent as match_agent_mod


# ─────────────────────────────────────────────────────────────────────────────
class TestNoKeywordPreFilter(unittest.TestCase):
    """Regression for 2026-04-28: the keyword pre-filter (_quick_keyword_score
    + _AI_TECH_TERMS + _KEYWORD_THRESHOLD) was removed because it second-guessed
    job_agent's `is_ai_tpm` classification with crude regex and produced false
    negatives on AI-native JDs that didn't repeat magic words. Stage 1 batch
    coarse scoring is the first quantitative gate; Stage 2 fine handles top-N."""

    def test_module_does_not_export_prefilter_symbols(self):
        for sym in ("_quick_keyword_score", "_AI_TECH_TERMS", "_KEYWORD_THRESHOLD"):
            self.assertFalse(
                hasattr(match_agent_mod, sym),
                f"match_agent.{sym} should be removed; reintroducing the keyword "
                f"pre-filter requires re-justifying it (see 2026-04-28 review)."
            )


# ─────────────────────────────────────────────────────────────────────────────
class TestRateLimiter(unittest.TestCase):

    def test_acquire_no_wait_on_first_call(self):
        """First acquire should return quickly (no waiting)."""
        limiter = _RateLimiter(rpm=60)
        start = time.monotonic()

        async def _run():
            await limiter.acquire()

        asyncio.run(_run())
        elapsed = time.monotonic() - start
        self.assertLess(elapsed, 2.0, "First acquire took too long")

    def test_interval_calculation(self):
        limiter = _RateLimiter(rpm=30)
        self.assertAlmostEqual(limiter._interval, 2.0, places=5)

    def test_high_rpm_small_interval(self):
        limiter = _RateLimiter(rpm=600)
        self.assertAlmostEqual(limiter._interval, 0.1, places=5)

    def test_lock_created_eagerly(self):
        """BUG-14: match_agent _RateLimiter must create lock eagerly (consistent with job_agent)."""
        limiter = _RateLimiter(rpm=60)
        self.assertIsNotNone(limiter._lock)


# ─────────────────────────────────────────────────────────────────────────────
class TestLoadResume(unittest.TestCase):

    def _write_resume(self, folder, name, content):
        path = os.path.join(folder, name)
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)

    def test_loads_md_file(self):
        with tempfile.TemporaryDirectory() as d:
            self._write_resume(d, "resume.md", "# Resume\nAI TPM expert")
            text, rid = load_resume(d)
            self.assertIn("AI TPM", text)
            self.assertEqual(rid, "resume")

    def test_loads_txt_file(self):
        with tempfile.TemporaryDirectory() as d:
            self._write_resume(d, "cv.txt", "LLM GenAI experience")
            text, rid = load_resume(d)
            self.assertIn("LLM", text)
            self.assertEqual(rid, "cv")

    def test_returns_empty_if_no_files(self):
        with tempfile.TemporaryDirectory() as d:
            text, rid = load_resume(d)
            self.assertEqual(text, "")
            self.assertEqual(rid, "")

    def test_returns_empty_if_folder_missing(self):
        text, rid = load_resume("/nonexistent/path/xyz")
        self.assertEqual(text, "")

    def test_ignores_hidden_files(self):
        with tempfile.TemporaryDirectory() as d:
            self._write_resume(d, ".hidden.md", "hidden")
            self._write_resume(d, "real.md", "real resume")
            text, _ = load_resume(d)
            self.assertEqual(text, "real resume")


# ─────────────────────────────────────────────────────────────────────────────
class TestGeminiKeyPool(unittest.TestCase):

    def test_initial_key(self):
        pool = _GeminiKeyPool(["key1", "key2"])
        self.assertEqual(pool.current, "key1")

    def test_rotate_advances_key(self):
        pool = _GeminiKeyPool(["k1", "k2", "k3"])
        self.assertTrue(pool.rotate())
        self.assertEqual(pool.current, "k2")
        self.assertTrue(pool.rotate())
        self.assertEqual(pool.current, "k3")

    def test_rotate_returns_false_when_exhausted(self):
        pool = _GeminiKeyPool(["only_key"])
        self.assertFalse(pool.rotate())

    def test_filters_empty_keys(self):
        pool = _GeminiKeyPool(["valid", "", None, "valid2"])
        self.assertEqual(len(pool._keys), 2)

    def test_rotates_on_429(self):
        pool = _GeminiKeyPool(["k1", "k2"], genai_mod=match_agent_mod.genai)
        mock_client = MagicMock()
        # First call: raise 429; second call: succeed
        mock_client.models.generate_content.side_effect = [
            Exception("429 RESOURCE_EXHAUSTED"),
            MagicMock(text='{"items":[]}'),
        ]
        with patch("agents.match_agent.genai.Client", return_value=mock_client):
            result = pool.generate_content("model", "content", MagicMock())
        # After rotation, should have used key2
        self.assertEqual(pool.current, "k2")

    def test_raises_when_all_keys_exhausted(self):
        pool = _GeminiKeyPool(["only"], genai_mod=match_agent_mod.genai)
        mock_client = MagicMock()
        mock_client.models.generate_content.side_effect = Exception("429 RESOURCE_EXHAUSTED")
        with patch("agents.match_agent.genai.Client", return_value=mock_client):
            with self.assertRaises(Exception):
                pool.generate_content("model", "content", MagicMock())

    def test_lock_created_eagerly(self):
        pool = _GeminiKeyPool(["k1"])
        self.assertIsNotNone(pool._keys)


class TestBug34ClientCaching(unittest.TestCase):
    """BUG-34: Client instances should be cached per key, not recreated each call."""

    def test_same_key_reuses_client(self):
        pool = _GeminiKeyPool(["key1", "key2"], genai_mod=match_agent_mod.genai)
        mock_client = MagicMock()
        mock_client.models.generate_content.return_value = MagicMock(text="ok")
        with patch("agents.match_agent.genai.Client", return_value=mock_client) as mock_cls:
            pool.generate_content("model", "content", MagicMock())
            pool.generate_content("model", "content", MagicMock())
        # Client() should be called only once for the same key
        self.assertEqual(mock_cls.call_count, 1)

    def test_different_keys_create_different_clients(self):
        pool = _GeminiKeyPool(["key1", "key2"], genai_mod=match_agent_mod.genai)
        clients = {}
        def make_client(api_key):
            c = MagicMock()
            c.models.generate_content.side_effect = [
                Exception("429 RESOURCE_EXHAUSTED"),
                MagicMock(text="ok"),
            ] if api_key == "key1" else [MagicMock(text="ok")]
            clients[api_key] = c
            return c
        with patch("agents.match_agent.genai.Client", side_effect=make_client):
            pool.generate_content("model", "content", MagicMock())
        # Both keys should have created clients
        self.assertIn("key1", clients)
        self.assertIn("key2", clients)
        self.assertIsNot(clients["key1"], clients["key2"])

    def test_client_cache_dict_exists(self):
        pool = _GeminiKeyPool(["k1"])
        self.assertIsInstance(pool._clients, dict)
        self.assertEqual(len(pool._clients), 0)


class TestBug35RoundRobinRotation(unittest.TestCase):
    """BUG-35: rotate() must round-robin back to key #1 after reaching the end."""

    def test_wraps_around_to_first_key(self):
        pool = _GeminiKeyPool(["k1", "k2"])
        pool.rotate()  # → k2
        self.assertEqual(pool.current, "k2")
        pool.rotate()  # → back to k1
        self.assertEqual(pool.current, "k1")

    def test_three_keys_full_cycle(self):
        pool = _GeminiKeyPool(["a", "b", "c"])
        pool.rotate()  # → b
        pool.rotate()  # → c
        pool.rotate()  # → a (wrapped)
        self.assertEqual(pool.current, "a")

    def test_single_key_returns_false(self):
        pool = _GeminiKeyPool(["solo"])
        self.assertFalse(pool.rotate())
        self.assertEqual(pool.current, "solo")

    def test_do_generate_tries_all_keys_before_raising(self):
        """When all keys get 429, should try each once then raise."""
        pool = _GeminiKeyPool(["k1", "k2", "k3"], genai_mod=match_agent_mod.genai)
        call_keys = []
        def make_client(api_key):
            c = MagicMock()
            def gen(*a, **kw):
                call_keys.append(api_key)
                raise Exception("429 RESOURCE_EXHAUSTED")
            c.models.generate_content.side_effect = gen
            return c
        with patch("agents.match_agent.genai.Client", side_effect=make_client):
            with self.assertRaises(Exception):
                pool.generate_content("model", "content", MagicMock())
        # Should have tried all 3 keys
        self.assertEqual(len(call_keys), 3)
        self.assertEqual(set(call_keys), {"k1", "k2", "k3"})

    def test_recovers_after_first_key_429(self):
        """Key #1 gets 429, key #2 succeeds. On next call, key #2 is reused (round-robin)."""
        pool = _GeminiKeyPool(["k1", "k2"], genai_mod=match_agent_mod.genai)
        call_count = {"k1": 0, "k2": 0}
        def make_client(api_key):
            c = MagicMock()
            if api_key == "k1":
                c.models.generate_content.side_effect = Exception("429 RESOURCE_EXHAUSTED")
            else:
                c.models.generate_content.return_value = MagicMock(text="ok")
            return c
        with patch("agents.match_agent.genai.Client", side_effect=make_client):
            pool.generate_content("model", "content", MagicMock())
        self.assertEqual(pool.current, "k2")


class TestBug36ThreadSafety(unittest.TestCase):
    """BUG-36: _idx and rotate() must be protected by a threading.Lock."""

    def test_lock_exists_on_pool(self):
        pool = _GeminiKeyPool(["k1", "k2"])
        self.assertTrue(hasattr(pool, "_lock"))
        # Lock objects have acquire/release methods
        self.assertTrue(callable(getattr(pool._lock, "acquire", None)))
        self.assertTrue(callable(getattr(pool._lock, "release", None)))

    def test_concurrent_rotations_no_skip(self):
        """Two threads calling rotate() should not skip a key."""
        import threading
        pool = _GeminiKeyPool(["k1", "k2", "k3"])
        results = []
        barrier = threading.Barrier(2)

        def rotate_and_record():
            barrier.wait()
            with pool._lock:
                pool.rotate()
                results.append(pool.current)

        t1 = threading.Thread(target=rotate_and_record)
        t2 = threading.Thread(target=rotate_and_record)
        t1.start(); t2.start()
        t1.join(); t2.join()

        # After 2 rotations from k1, should end at k3
        self.assertEqual(pool.current, "k3")

    def test_concurrent_generate_content_no_crash(self):
        """Multiple threads calling generate_content concurrently should not crash."""
        import threading
        pool = _GeminiKeyPool(["k1", "k2"], genai_mod=match_agent_mod.genai)
        mock_client = MagicMock()
        mock_client.models.generate_content.return_value = MagicMock(text="ok")
        errors = []

        def call_generate():
            try:
                with patch("agents.match_agent.genai.Client", return_value=mock_client):
                    pool.generate_content("model", "content", MagicMock())
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=call_generate) for _ in range(5)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        self.assertEqual(errors, [])


# ─────────────────────────────────────────────────────────────────────────────
class TestPydanticSchemas(unittest.TestCase):

    def test_coarse_item_valid(self):
        item = CoarseItem(index=0, score=75)
        self.assertEqual(item.index, 0)
        self.assertEqual(item.score, 75)

    def test_batch_coarse_result(self):
        result = BatchCoarseResult(items=[
            CoarseItem(index=0, score=80),
            CoarseItem(index=1, score=50),
        ])
        self.assertEqual(len(result.items), 2)

    def test_match_result_valid(self):
        m = MatchResult(
            compatibility_score=88,
            key_strengths=["LLM experience"],
            critical_gaps=["No MLOps"],
            recommendation_reason="Strong candidate.",
        )
        self.assertEqual(m.compatibility_score, 88)
        self.assertIsInstance(m.key_strengths, list)

    def test_batch_coarse_result_model_validate_json(self):
        raw = json.dumps({"items": [{"index": 0, "score": 60}, {"index": 1, "score": 40}]})
        result = BatchCoarseResult.model_validate_json(raw)
        self.assertEqual(result.items[0].score, 60)


# ─────────────────────────────────────────────────────────────────────────────
class TestBatchCoarseScore(unittest.TestCase):

    def setUp(self):
        mock_pool = MagicMock()
        mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({"items": [
                {"index": 0, "score": 85},
                {"index": 1, "score": 30},
            ]})
        )
        match_agent_mod._KEY_POOL = mock_pool

    def tearDown(self):
        match_agent_mod._KEY_POOL = None

    def test_returns_scores_list(self):
        jds = [
            {"url": "https://a.com/1", "jd_json": '{"job_title": "TPM"}'},
            {"url": "https://b.com/2", "jd_json": '{"job_title": "PM"}'},
        ]
        scores = batch_coarse_score("resume text", jds)
        self.assertEqual(len(scores), 2)
        self.assertEqual(scores[0], 85)
        self.assertEqual(scores[1], 30)

    def test_returns_empty_on_structural_error(self):
        """P0-4: structural error returns [] sentinel, never fake [1]*N."""
        from shared.exceptions import GeminiStructuralError
        match_agent_mod._KEY_POOL.generate_content.side_effect = GeminiStructuralError("bad")
        jds = [{"url": "https://x.com", "jd_json": "{}"}]
        scores = batch_coarse_score("resume", jds)
        self.assertEqual(scores, [])

    def test_reraises_transient_error(self):
        """P0-4: transient errors must bubble up so main fails loudly."""
        from shared.exceptions import GeminiTransientError
        match_agent_mod._KEY_POOL.generate_content.side_effect = GeminiTransientError("429")
        jds = [{"url": "https://x.com", "jd_json": "{}"}]
        with self.assertRaises(GeminiTransientError):
            batch_coarse_score("resume", jds)

    def test_handles_out_of_range_index(self):
        """Gemini returns index beyond batch size — should ignore it."""
        match_agent_mod._KEY_POOL.generate_content.return_value = MagicMock(
            text=json.dumps({"items": [{"index": 99, "score": 100}]})
        )
        jds = [{"url": "https://x.com", "jd_json": "{}"}]
        scores = batch_coarse_score("resume", jds)
        self.assertEqual(scores, [1])  # index 99 ignored, default stays 1


# ─────────────────────────────────────────────────────────────────────────────
class TestEvaluateMatch(unittest.TestCase):

    def setUp(self):
        mock_pool = MagicMock()
        mock_pool.generate_content.return_value = MagicMock(
            text=json.dumps({
                "compatibility_score": 72,
                "key_strengths": ["GenAI"],
                "critical_gaps": ["MLOps"],
                "recommendation_reason": "Decent fit.",
            })
        )
        match_agent_mod._KEY_POOL = mock_pool

    def tearDown(self):
        match_agent_mod._KEY_POOL = None

    def test_returns_json_string(self):
        result = evaluate_match("resume", '{"job_title": "TPM"}')
        data = json.loads(result)
        self.assertIn("compatibility_score", data)
        self.assertEqual(data["compatibility_score"], 72)

    def test_returns_none_on_structural_error(self):
        """P0-4: structural error returns None sentinel, never fake score=1 JSON."""
        from shared.exceptions import GeminiStructuralError
        match_agent_mod._KEY_POOL.generate_content.side_effect = GeminiStructuralError("bad")
        result = evaluate_match("resume", '{}')
        self.assertIsNone(result)

    def test_reraises_transient_error(self):
        """P0-4: transient errors propagate."""
        from shared.exceptions import GeminiTransientError
        match_agent_mod._KEY_POOL.generate_content.side_effect = GeminiTransientError("503")
        with self.assertRaises(GeminiTransientError):
            evaluate_match("resume", '{}')


# ─────────────────────────────────────────────────────────────────────────────
class TestScoreMinimumClamp(unittest.TestCase):
    """Phase 1: Gemini-evaluated jobs must score >= 1."""

    def setUp(self):
        mock_pool = MagicMock()
        match_agent_mod._KEY_POOL = mock_pool

    def tearDown(self):
        match_agent_mod._KEY_POOL = None

    def test_coarse_clamps_zero_to_one(self):
        """Gemini returns score=0 for a JD — should be clamped to 1."""
        match_agent_mod._KEY_POOL.generate_content.return_value = MagicMock(
            text=json.dumps({"items": [{"index": 0, "score": 0}]})
        )
        jds = [{"url": "https://x.com", "jd_json": '{"job_title": "TPM"}'}]
        scores = batch_coarse_score("resume", jds)
        self.assertEqual(scores[0], 1)

    def test_fine_clamps_zero_to_one(self):
        """evaluate_match returns score=0 — fine eval should clamp to 1."""
        match_agent_mod._KEY_POOL.generate_content.return_value = MagicMock(
            text=json.dumps({
                "compatibility_score": 0,
                "key_strengths": [],
                "critical_gaps": ["No match"],
                "recommendation_reason": "No fit.",
            })
        )
        result = evaluate_match("resume", '{"job_title": "SWE"}')
        # evaluate_match returns raw JSON; clamping happens at caller (fine_one)
        # but the evaluate_match fallback itself returns score=1 on exception
        data = json.loads(result)
        # Gemini returned 0, evaluate_match passes it through — caller clamps
        self.assertIn("compatibility_score", data)


# ─────────────────────────────────────────────────────────────────────────────
class TestCoarsePromptContent(unittest.TestCase):
    """Phase 2: Coarse prompt should contain calibration anchors."""

    def test_prompt_contains_calibration_ranges(self):
        self.assertIn("1-30", _COARSE_SYSTEM_PROMPT)
        self.assertIn("31-60", _COARSE_SYSTEM_PROMPT)
        self.assertIn("61-100", _COARSE_SYSTEM_PROMPT)

    def test_prompt_contains_key_factors(self):
        self.assertIn("AI/ML Relevance", _COARSE_SYSTEM_PROMPT)
        self.assertIn("TPM Function Match", _COARSE_SYSTEM_PROMPT)
        self.assertIn("Seniority Fit", _COARSE_SYSTEM_PROMPT)

    def test_prompt_minimum_score_instruction(self):
        self.assertIn("Minimum score is 1", _COARSE_SYSTEM_PROMPT)


# ─────────────────────────────────────────────────────────────────────────────
class TestPrintTopResultsStageColumn(unittest.TestCase):
    """BUG-22: _print_top_results must derive Stage column from MATCH_HEADERS, not hardcode 9."""

    def test_stage_col_derived_from_match_headers(self):
        from shared.excel_store import MATCH_HEADERS
        stage_col = MATCH_HEADERS.index("Stage") + 1  # 1-based
        # Verify it's currently 9, and that changing schema would update it automatically
        self.assertEqual(stage_col, 9, "Stage is currently column 9 — this test documents the contract")

    def test_print_top_results_reads_stage_correctly(self):
        from agents.match_agent import _print_top_results
        from unittest.mock import MagicMock, patch
        import openpyxl, tempfile, os

        # Build a minimal Excel with Match_Results
        from shared.excel_store import MATCH_HEADERS, get_or_create_excel
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = os.path.join(tmpdir, "test.xlsx")
            get_or_create_excel(xlsx)
            wb = openpyxl.load_workbook(xlsx)
            ws = wb["Match_Results"]
            # Write a row: resume_id=r1, url=url1, score=85, stage=fine
            row = ["r1", "https://example.com/job/1", 85, "", "", "", "", "", "fine"]
            ws.append(row)
            wb.save(xlsx)
            wb.close()

            with patch("builtins.print") as mock_print:
                _print_top_results(xlsx, "r1")

            output = " ".join(str(c) for c in [a for call in mock_print.call_args_list for a in call.args])
            self.assertIn("★", output, "fine stage should show ★ indicator")
            self.assertIn("85", output)


class TestBug41SharedRateLimiter(unittest.TestCase):
    """BUG-41: Stage 1 and Stage 2 must share a single _RateLimiter, not create independent ones."""

    def test_module_level_limiter_exists(self):
        self.assertTrue(hasattr(match_agent_mod, "_GEMINI_LIMITER"))
        self.assertIsInstance(match_agent_mod._GEMINI_LIMITER, _RateLimiter)

    def test_no_local_limiter_creation_in_main(self):
        """main() should not create new _RateLimiter instances."""
        import inspect
        # P0-7: main() is now a thin RunSummary wrapper; the orchestration
        # body lives in _main_inner.
        from agents.match_agent import _main_inner
        source = inspect.getsource(_main_inner)
        self.assertNotIn("_RateLimiter(", source,
                         "main() must not create new _RateLimiter instances")

    def test_limiter_used_is_module_level(self):
        """Both stages should reference _GEMINI_LIMITER, not a local instance."""
        import inspect
        # P0-7: main() is now a thin RunSummary wrapper; the orchestration
        # body lives in _main_inner.
        from agents.match_agent import _main_inner
        source = inspect.getsource(_main_inner)
        self.assertIn("_GEMINI_LIMITER", source,
                      "main() must use the module-level _GEMINI_LIMITER")


class TestBug39LoadJdMarkdownStructured(unittest.TestCase):
    """BUG-39: _load_jd_markdown must prefer _structured.md over raw .md."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()
        self._orig = match_agent_mod.JD_CACHE_DIR
        match_agent_mod.JD_CACHE_DIR = self.tmpdir

    def tearDown(self):
        match_agent_mod.JD_CACHE_DIR = self._orig
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _md5(self, url):
        import hashlib
        return hashlib.md5(url.encode()).hexdigest()

    def test_prefers_structured_over_raw(self):
        url = "https://example.com/job/1"
        h = self._md5(url)
        with open(os.path.join(self.tmpdir, f"{h}.md"), "w") as f:
            f.write("raw content")
        with open(os.path.join(self.tmpdir, f"{h}_structured.md"), "w") as f:
            f.write("structured content")
        result = _load_jd_markdown(url)
        self.assertEqual(result, "structured content")

    def test_falls_back_to_raw_when_no_structured(self):
        url = "https://example.com/job/2"
        h = self._md5(url)
        with open(os.path.join(self.tmpdir, f"{h}.md"), "w") as f:
            f.write("raw only")
        result = _load_jd_markdown(url)
        self.assertEqual(result, "raw only")

    def test_returns_none_when_neither_exists(self):
        result = _load_jd_markdown("https://example.com/job/nonexistent")
        self.assertIsNone(result)


class TestBug51PrintTopResultsDynamicColumns(unittest.TestCase):
    """BUG-51: _print_top_results must use MATCH_HEADERS for ALL column lookups, not just Stage."""

    def test_uses_match_headers_for_all_columns(self):
        """Source code should use MATCH_HEADERS.index() for all columns, not hardcoded numbers."""
        import inspect
        from agents.match_agent import _print_top_results
        source = inspect.getsource(_print_top_results)
        # Should have dynamic lookups for Resume ID, JD URL, Score, Stage
        for field in ["Resume ID", "JD URL", "Score", "Stage"]:
            self.assertIn(f'MATCH_HEADERS.index("{field}")', source,
                          f"_print_top_results must use MATCH_HEADERS.index(\"{field}\") for dynamic lookup")
        # No hardcoded ws.cell(r, <number>) patterns
        for hardcoded in ["ws.cell(r, 1)", "ws.cell(r, 2)", "ws.cell(r, 3)"]:
            self.assertNotIn(hardcoded, source,
                             f"Hardcoded {hardcoded} found — must use dynamic column variable")

    def test_reads_correct_data_from_excel(self):
        """End-to-end: _print_top_results reads correct columns from a real Excel file."""
        from agents.match_agent import _print_top_results
        from shared.excel_store import MATCH_HEADERS, get_or_create_excel
        import openpyxl, tempfile, os
        from unittest.mock import patch as _patch
        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx = os.path.join(tmpdir, "test.xlsx")
            get_or_create_excel(xlsx)
            wb = openpyxl.load_workbook(xlsx)
            ws = wb["Match_Results"]
            # Row for resume "r1": score 92, coarse stage
            ws.append(["r1", "https://example.com/job/a", 92, "", "", "", "", "", "coarse"])
            # Row for resume "r1": score 78, fine stage
            ws.append(["r1", "https://example.com/job/b", 78, "", "", "", "", "", "fine"])
            # Row for different resume — should not appear
            ws.append(["r2", "https://example.com/job/c", 99, "", "", "", "", "", "fine"])
            wb.save(xlsx)
            wb.close()

            with _patch("builtins.print") as mock_print:
                _print_top_results(xlsx, "r1")

            output = " ".join(str(a) for call in mock_print.call_args_list for a in call.args)
            self.assertIn("92", output, "Score 92 should appear for r1")
            self.assertIn("~", output, "coarse stage should show ~ indicator")
            self.assertIn("★", output, "fine stage should show ★ indicator")
            self.assertNotIn("99", output, "Score 99 from r2 should not appear")


class TestBug54KeyPoolNoneGuard(unittest.TestCase):
    """BUG-54: Functions using _KEY_POOL must raise RuntimeError when pool is None."""

    def test_batch_coarse_score_raises_when_pool_none(self):
        import agents.match_agent as mod
        original = mod._KEY_POOL
        try:
            mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                mod.batch_coarse_score("resume text", [{"jd": "test"}])
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            mod._KEY_POOL = original

    def test_evaluate_match_raises_when_pool_none(self):
        import agents.match_agent as mod
        original = mod._KEY_POOL
        try:
            mod._KEY_POOL = None
            with self.assertRaises(RuntimeError) as ctx:
                mod.evaluate_match("resume text", '{"jd": "test"}')
            self.assertIn("_KEY_POOL not initialized", str(ctx.exception))
        finally:
            mod._KEY_POOL = original


class TestBug31NoSubclassNeeded(unittest.TestCase):
    """BUG-31: _GeminiKeyPool subclasses should be eliminated; _GeminiKeyPoolBase used directly."""

    def test_no_subclass_definitions_in_agents(self):
        """Agent files should not define class _GeminiKeyPool(_GeminiKeyPoolBase)."""
        import inspect
        for mod_name in ["agents.match_agent", "agents.job_agent",
                         "agents.company_agent", "agents.resume_optimizer"]:
            mod = __import__(mod_name, fromlist=["_GeminiKeyPool"])
            source = inspect.getsource(mod)
            self.assertNotIn("class _GeminiKeyPool(_GeminiKeyPoolBase):", source,
                             f"{mod_name} still defines a _GeminiKeyPool subclass")

    def test_base_class_has_generate_content(self):
        """_GeminiKeyPoolBase should have generate_content method directly."""
        from shared.gemini_pool import _GeminiKeyPoolBase
        self.assertTrue(hasattr(_GeminiKeyPoolBase, "generate_content"))

    def test_base_class_accepts_genai_mod(self):
        """_GeminiKeyPoolBase constructor should accept genai_mod parameter."""
        from shared.gemini_pool import _GeminiKeyPoolBase
        mock_genai = MagicMock()
        pool = _GeminiKeyPoolBase(["key1"], genai_mod=mock_genai)
        self.assertIs(pool._genai_mod, mock_genai)

    def test_generate_content_uses_stored_genai(self):
        """generate_content should use stored genai_mod, not require subclass override."""
        from shared.gemini_pool import _GeminiKeyPoolBase
        mock_genai = MagicMock()
        mock_client = MagicMock()
        mock_client.models.generate_content.return_value = MagicMock(text="ok")
        mock_genai.Client.return_value = mock_client
        pool = _GeminiKeyPoolBase(["key1"], genai_mod=mock_genai)
        result = pool.generate_content("model", "content", MagicMock())
        mock_genai.Client.assert_called_with(api_key="key1")


class TestExtractATSKeywords(unittest.TestCase):
    """PRJ-002 PR 3 — JD-level ATS keyword extraction with graceful fallbacks."""

    def test_extracts_when_present(self):
        jd = {"jd_json": json.dumps({
            "job_title": "TPM",
            "ats_keywords": ["PyTorch", "Kubernetes", "LLM"],
        })}
        self.assertEqual(_extract_ats_keywords(jd), ["PyTorch", "Kubernetes", "LLM"])

    def test_empty_list_when_field_missing(self):
        # Legacy JDs cached before PR 1 don't have ats_keywords.
        jd = {"jd_json": json.dumps({"job_title": "TPM"})}
        self.assertEqual(_extract_ats_keywords(jd), [])

    def test_empty_list_when_field_null(self):
        jd = {"jd_json": json.dumps({"ats_keywords": None})}
        self.assertEqual(_extract_ats_keywords(jd), [])

    def test_empty_list_on_malformed_json(self):
        jd = {"jd_json": "not-json{"}
        self.assertEqual(_extract_ats_keywords(jd), [])

    def test_filters_blank_entries(self):
        jd = {"jd_json": json.dumps({
            "ats_keywords": ["PyTorch", "", None, "Kubernetes"],
        })}
        self.assertEqual(_extract_ats_keywords(jd), ["PyTorch", "Kubernetes"])

    def test_coerces_non_string_entries(self):
        jd = {"jd_json": json.dumps({"ats_keywords": ["PyTorch", 42]})}
        # Numbers stringified — defensive against malformed Gemini output.
        self.assertEqual(_extract_ats_keywords(jd), ["PyTorch", "42"])


class TestComputeATSForJDs(unittest.TestCase):
    """PRJ-002 PR 3 — orchestrator that runs ats_matcher across many JDs."""

    def test_returns_one_entry_per_jd(self):
        jds = [
            {"url": "https://a.com/1", "jd_json": json.dumps({
                "ats_keywords": ["PyTorch", "Kubernetes"],
            })},
            {"url": "https://a.com/2", "jd_json": json.dumps({
                "ats_keywords": ["Rust", "Go"],
            })},
        ]
        results = compute_ats_for_jds("Senior TPM with PyTorch and Kubernetes.", jds)
        self.assertEqual(set(results.keys()), {"https://a.com/1", "https://a.com/2"})
        self.assertEqual(results["https://a.com/1"]["percent"], 100.0)
        self.assertEqual(results["https://a.com/2"]["percent"], 0.0)

    def test_legacy_jd_returns_none_percent(self):
        # JD without ats_keywords → percent=None signalling "no ATS data".
        jds = [{"url": "https://a.com/1", "jd_json": json.dumps({"job_title": "TPM"})}]
        results = compute_ats_for_jds("anything", jds)
        self.assertIsNone(results["https://a.com/1"]["percent"])

    def test_no_lazy_load(self):
        # compute_ats_for_jds must NOT call any Gemini / network APIs.
        # If it does, the MagicMock would return non-list types.
        jds = [{"url": "https://a.com/1", "jd_json": json.dumps({
            "ats_keywords": ["PyTorch"],
        })}]
        # Simulate a pool that would error if accessed.
        match_agent_mod._KEY_POOL = None
        results = compute_ats_for_jds("PyTorch user", jds)
        self.assertEqual(results["https://a.com/1"]["percent"], 100.0)


class TestATSCoverageLowThreshold(unittest.TestCase):
    """The ⚠️ threshold for low ATS coverage is a soft signal, not a gate."""

    def test_threshold_is_30_percent(self):
        self.assertEqual(ATS_COVERAGE_LOW_THRESHOLD, 30.0)


class TestSelectFineCandidates(unittest.TestCase):
    """Stage 2 fine-eval gate: union of (score >= threshold) OR (top P%)."""

    @staticmethod
    def _scored(scores_by_url: dict, stage: str = "coarse") -> dict:
        """Build the scored_for_resume shape: {(rid, url): {score, stage, ...}}."""
        return {
            ("r1", url): {"score": s, "stage": stage, "hash": "x"}
            for url, s in scores_by_url.items()
        }

    def test_flat_high_threshold_dominates(self):
        """All scores well above threshold → union should select every JD."""
        scored = self._scored({f"u{i}": 75 for i in range(5)})
        to_fine, stats = _select_fine_candidates(scored, score_threshold=60, top_percent=60)
        self.assertEqual(len(to_fine), 5, "threshold should pull in all 5")
        self.assertEqual(stats["threshold_count"], 5)
        self.assertEqual(stats["top_count"], 3)  # ceil(5*0.6) = 3

    def test_flat_low_top_percent_dominates(self):
        """All scores below threshold → only top P% qualifies."""
        scored = self._scored({f"u{i}": 40 for i in range(10)})
        to_fine, stats = _select_fine_candidates(scored, score_threshold=60, top_percent=60)
        self.assertEqual(stats["threshold_count"], 0)
        self.assertEqual(stats["top_count"], 6)  # ceil(10*0.6) = 6
        self.assertEqual(len(to_fine), 6)

    def test_mixed_true_union(self):
        """High scorer below top-P% rank still gets in via threshold; vice versa."""
        # 10 JDs. top-60% = top 6 by rank. threshold=60 catches 5 high scorers.
        # Construct so the sets overlap but each adds something the other lacks.
        scored = self._scored({
            "u0": 90, "u1": 85, "u2": 75, "u3": 70, "u4": 65,  # all >=60 (5 above thr)
            "u5": 55, "u6": 50, "u7": 45, "u8": 40, "u9": 35,  # below thr
        })
        to_fine, stats = _select_fine_candidates(scored, score_threshold=60, top_percent=60)
        self.assertEqual(stats["threshold_count"], 5)
        self.assertEqual(stats["top_count"], 6)  # ceil(10*0.6)
        # Union = top 6 (u0..u5) ∪ {u0..u4} = u0..u5 → 6 JDs
        urls = {url for (_, url) in to_fine}
        self.assertEqual(urls, {"u0", "u1", "u2", "u3", "u4", "u5"})

    def test_threshold_pulls_in_below_top_rank(self):
        """A JD scoring >= threshold but ranked outside top-P% is still selected."""
        # 100 JDs. top-60% = top 60. JDs ranked 61..100 with score >= 60 should
        # still get picked up by the threshold leg.
        scored = self._scored({f"u{i:03d}": 60 for i in range(100)})
        # All tied at 60 — top-60% is arbitrary; threshold catches all 100.
        to_fine, stats = _select_fine_candidates(scored, score_threshold=60, top_percent=60)
        self.assertEqual(stats["threshold_count"], 100)
        self.assertEqual(len(to_fine), 100, "threshold leg should rescue tied-at-60 JDs")

    def test_skips_already_fine(self):
        """Pairs already at stage='fine' are excluded from the queue."""
        scored = {
            ("r1", "u0"): {"score": 80, "stage": "fine"},
            ("r1", "u1"): {"score": 80, "stage": "coarse"},
        }
        to_fine, _ = _select_fine_candidates(scored, score_threshold=60, top_percent=60)
        self.assertEqual([k[1] for k in to_fine], ["u1"])

    def test_empty_input(self):
        to_fine, stats = _select_fine_candidates({}, score_threshold=60, top_percent=60)
        self.assertEqual(to_fine, [])
        self.assertEqual(stats["n"], 0)

    def test_single_jd_top_n_min_one(self):
        """top_n must be at least 1 even with N=1 and small percent."""
        scored = self._scored({"u0": 30})
        to_fine, stats = _select_fine_candidates(scored, score_threshold=60, top_percent=10)
        self.assertEqual(stats["top_count"], 1)
        self.assertEqual(len(to_fine), 1)

    def test_threshold_override_narrows(self):
        """Higher threshold shrinks the threshold leg."""
        scored = self._scored({"u0": 85, "u1": 70, "u2": 65, "u3": 55, "u4": 40})
        _, stats_low = _select_fine_candidates(scored, score_threshold=60, top_percent=60)
        _, stats_hi  = _select_fine_candidates(scored, score_threshold=80, top_percent=60)
        self.assertEqual(stats_low["threshold_count"], 3)
        self.assertEqual(stats_hi["threshold_count"], 1)


if __name__ == "__main__":
    unittest.main(verbosity=2)

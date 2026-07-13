"""
Shared Excel persistence layer — used by all three agents.
All functions load → modify → save on each call (safe for single-process use).
job_agent wraps writes in an asyncio.Lock to prevent concurrent corruption.
"""
import os
import re
import json
from datetime import datetime, date
from urllib.parse import urlsplit, parse_qsl, urlencode
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from shared.config import TRACK_ORDER

# ── Path ──────────────────────────────────────────────────────────────────────
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH   = os.path.join(PROJECT_ROOT, "pathfinder_dashboard.xlsx")

# ── Sheet headers ─────────────────────────────────────────────────────────────
# PRJ-004 (D-09/D-15): "AI Domain" → "Track" (6-bucket taxonomy), "AI TPM Jobs"
# → "Qualified Jobs" (count of domain-qualified rows across all 5 tracks).
COMPANY_HEADERS          = ["Company Name", "Track", "Business Focus", "Career URL", "Updated At", "TPM Jobs", "Qualified Jobs", "No TPM Count", "Auto Archived"]
WITHOUT_TPM_HEADERS      = ["Company Name", "Track", "Business Focus", "Career URL", "Updated At", "TPM Jobs", "Qualified Jobs"]
# PRJ-004 REQ-004-14: every JD row is domain-qualified at write time, so the
# boolean "Is AI TPM" becomes the 5-value "Job Domain". "Location Tier" becomes
# the combined freshness×region "Sort Tier" (1–6, or 9 for unknown/aged rows).
JOB_DOMAIN_VALUES = ("AI", "Robotics", "Fintech", "Space", "Defense")
JD_HEADERS      = ["JD URL", "Job Title", "Company", "Location", "Salary", "Requirements",
                   "Additional Qualifications", "Responsibilities", "Job Domain", "Updated At", "MD Hash",
                   "Data Quality",
                   # PRJ-002 Phase 4 fix: ats_keywords must be persisted in JD_Tracker
                   # so it survives the round-trip from job_agent (Gemini extraction)
                   # to match_agent / resume_optimizer (consumption). Without this column
                   # the ATS dimension was always silently None.
                   "ATS Keywords",
                   "Sort Tier",
                   # PRJ-004 REQ-004-08/10/11: freshness, seniority, work-auth audit.
                   "Posted Date", "Freshness Tier", "Min YoE", "YoE Flag",
                   "Work-Auth Status", "Date Flag"]
# PRJ-002 PR 2: 3-dimension scoring columns appended at end so existing
# column indices stay valid. PR 3/PR 4 wire up the upsert paths to populate
# them — for now they're added by migration and left blank for old rows.
MATCH_HEADERS   = ["Resume ID", "JD URL", "Score", "Strengths", "Gaps", "Reason",
                   "Updated At", "Resume Hash", "Stage",
                   "ATS Coverage %", "Recruiter Score", "HM Score", "ATS Missing"]
TAILORED_HEADERS = ["Resume ID", "JD URL", "Job Title", "Company", "Original Score",
                    "Tailored Score", "Score Delta", "Tailored Resume Path",
                    "Optimization Summary", "Updated At", "Resume Hash", "Regression",
                    # 3-dimension per-stage scores + deltas (PRJ-002 PR 2):
                    "Original ATS", "Tailored ATS", "ATS Delta",
                    "Original Recruiter", "Tailored Recruiter", "Recruiter Delta",
                    "Original HM", "Tailored HM", "HM Delta",
                    # sha256 of the .md file content the optimizer last wrote.
                    # Compared against the on-disk file before any subsequent
                    # overwrite — mismatch means the user hand-edited the file,
                    # so we skip the write to preserve their edit.
                    "Last Written Hash"]

# BUG-52: pre-computed 1-based column indices for JD_Tracker lookups
_JD_COL = {h: i + 1 for i, h in enumerate(JD_HEADERS)}

# BUG-65: user-owned triage tabs. Rows the user moves here out of JD_Tracker
# are final decisions — their URLs are permanently excluded from re-scraping
# and re-insertion. Same 20-col JD_HEADERS schema as JD_Tracker.
TRIAGE_SHEETS = ("JD_ToApply", "Skipped JD")


# ── Init ──────────────────────────────────────────────────────────────────────
def get_or_create_excel(xlsx_path: str = EXCEL_PATH) -> str:
    if os.path.exists(xlsx_path):
        try:
            wb = load_workbook(xlsx_path)
        except Exception as e:
            import shutil, logging as _log
            bak = xlsx_path + ".bak"
            try:
                shutil.move(xlsx_path, bak)
            except Exception as move_err:
                raise RuntimeError(
                    f"[Excel] Corrupted file at {xlsx_path} and cannot back up to {bak}: {move_err}"
                ) from e
            _log.warning(f"[Excel] Corrupted file moved to {bak}: {e}. Recreating.")
            return get_or_create_excel(xlsx_path)
        try:
            changed = False
            # Migrate: rename old "AI_" prefixed sheet tabs to new names
            _SHEET_RENAMES = [("AI_Company_List", "Company_List"),
                              ("AI_Company_Without_TPM", "Company_Without_TPM")]
            for old_name, new_name in _SHEET_RENAMES:
                if old_name in wb.sheetnames and new_name not in wb.sheetnames:
                    wb[old_name].title = new_name
                    changed = True
                    import logging as _log
                    _log.info(f"[Excel] Migrated sheet: '{old_name}' → '{new_name}'.")
            for name, headers in ([("Company_List",         COMPANY_HEADERS),
                                   ("Company_Without_TPM",  WITHOUT_TPM_HEADERS),
                                   ("JD_Tracker",              JD_HEADERS),
                                   ("Match_Results",           MATCH_HEADERS),
                                   ("Tailored_Match_Results",  TAILORED_HEADERS)]
                                  + [(t, JD_HEADERS) for t in TRIAGE_SHEETS]):
                if name not in wb.sheetnames:
                    wb.create_sheet(name).append(headers)
                    changed = True
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]
                changed = True
            # PRJ-004 (D-09/D-15): rename Company sheet headers in place.
            # Data values migrate separately via `company_agent --migrate-tracks`.
            for _sheet_name in ("Company_List", "Company_Without_TPM"):
                if _sheet_name in wb.sheetnames:
                    ws_x = wb[_sheet_name]
                    for c in range(1, ws_x.max_column + 1):
                        v = ws_x.cell(1, c).value
                        if v == "AI Domain":
                            ws_x.cell(1, c).value = "Track"
                            changed = True
                            import logging as _log
                            _log.info(f"[Excel] Migrated {_sheet_name}: renamed 'AI Domain' to 'Track'.")
                        elif v == "AI TPM Jobs":
                            ws_x.cell(1, c).value = "Qualified Jobs"
                            changed = True
                            import logging as _log
                            _log.info(f"[Excel] Migrated {_sheet_name}: renamed 'AI TPM Jobs' to 'Qualified Jobs'.")
            # PRJ-004 REQ-004-14: the legacy JD_Tracker schema (pre-multi-track,
            # detected by its "Is AI TPM" header) is replaced wholesale. There is
            # deliberately no data migration — the user wipes JD rows first
            # (intake C3); refuse loudly if that hasn't happened so the schema
            # change can never silently apply over legacy data.
            if "JD_Tracker" in wb.sheetnames:
                import logging as _log
                ws_jd = wb["JD_Tracker"]
                header_row = [ws_jd.cell(1, c).value for c in range(1, ws_jd.max_column + 1)]
                if "Is AI TPM" in header_row:
                    if any(ws_jd.cell(r, 1).value for r in range(2, ws_jd.max_row + 1)):
                        raise RuntimeError(
                            "JD_Tracker still contains legacy rows — wipe required before "
                            "PRJ-004 schema migration (design.md §4.2). Row removal is "
                            "user-owned; nothing is deleted automatically."
                        )
                    for c in range(1, max(len(header_row), len(JD_HEADERS)) + 1):
                        ws_jd.cell(1, c).value = JD_HEADERS[c - 1] if c <= len(JD_HEADERS) else None
                    changed = True
                    _log.info("[Excel] Migrated JD_Tracker: replaced legacy headers with PRJ-004 schema.")
            # Migrate Company_List: add "TPM Jobs" and "Qualified Jobs" columns if missing
            if "Company_List" in wb.sheetnames:
                ws_co = wb["Company_List"]
                co_headers = [ws_co.cell(1, c).value for c in range(1, ws_co.max_column + 1)]
                if "TPM Jobs" not in co_headers:
                    next_col = ws_co.max_column + 1
                    ws_co.cell(1, next_col).value = "TPM Jobs"
                    ws_co.cell(1, next_col + 1).value = "Qualified Jobs"
                    changed = True
                    import logging as _log
                    _log.info("[Excel] Migrated Company_List: added 'TPM Jobs' and 'Qualified Jobs' columns.")
                # REQ-063: add "No TPM Count" and "Auto Archived" columns if missing
                co_headers = [ws_co.cell(1, c).value for c in range(1, ws_co.max_column + 1)]
                if "No TPM Count" not in co_headers:
                    next_col = ws_co.max_column + 1
                    ws_co.cell(1, next_col).value = "No TPM Count"
                    ws_co.cell(1, next_col + 1).value = "Auto Archived"
                    changed = True
                    import logging as _log
                    _log.info("[Excel] Migrated Company_List: added 'No TPM Count' and 'Auto Archived' columns.")
            # BUG-55: Migrate Company_Without_TPM: add "TPM Jobs" and "Qualified Jobs" if missing
            if "Company_Without_TPM" in wb.sheetnames:
                ws_wt = wb["Company_Without_TPM"]
                wt_headers = [ws_wt.cell(1, c).value for c in range(1, ws_wt.max_column + 1)]
                if "TPM Jobs" not in wt_headers:
                    next_col = ws_wt.max_column + 1
                    ws_wt.cell(1, next_col).value = "TPM Jobs"
                    ws_wt.cell(1, next_col + 1).value = "Qualified Jobs"
                    changed = True
                    import logging as _log
                    _log.info("[Excel] Migrated Company_Without_TPM: added 'TPM Jobs' and 'Qualified Jobs' columns.")
            # Migrate Match_Results: add "Resume Hash" (col 8) and "Stage" (col 9) if missing
            if "Match_Results" in wb.sheetnames:
                ws_mr = wb["Match_Results"]
                mr_headers = [ws_mr.cell(1, c).value for c in range(1, ws_mr.max_column + 1)]
                if "Resume Hash" not in mr_headers:
                    next_col = ws_mr.max_column + 1
                    ws_mr.cell(1, next_col).value = "Resume Hash"
                    ws_mr.cell(1, next_col + 1).value = "Stage"
                    for r in range(2, ws_mr.max_row + 1):
                        if ws_mr.cell(r, 1).value:
                            ws_mr.cell(r, next_col).value = ""
                            ws_mr.cell(r, next_col + 1).value = "fine"
                    changed = True
                    import logging as _log
                    _log.info("[Excel] Migrated Match_Results: added 'Resume Hash' and 'Stage' columns.")
                # PRJ-002 PR 2: append 3-dimension scoring columns. Leaving
                # values blank for existing rows is intentional — the user
                # re-runs match_agent to populate them after upgrade.
                mr_headers = [ws_mr.cell(1, c).value for c in range(1, ws_mr.max_column + 1)]
                _new_match_cols = ["ATS Coverage %", "Recruiter Score", "HM Score", "ATS Missing"]
                _missing = [c for c in _new_match_cols if c not in mr_headers]
                if _missing:
                    for col_name in _missing:
                        next_col = ws_mr.max_column + 1
                        ws_mr.cell(1, next_col).value = col_name
                    changed = True
                    import logging as _log
                    _log.info(f"[Excel] Migrated Match_Results: added {_missing} columns.")
            # Migrate Tailored_Match_Results: add "Regression" column if missing.
            # Backfill: existing rows are marked TRUE iff Score Delta < 0.
            if "Tailored_Match_Results" in wb.sheetnames:
                ws_tm = wb["Tailored_Match_Results"]
                tm_headers = [ws_tm.cell(1, c).value for c in range(1, ws_tm.max_column + 1)]
                if "Regression" not in tm_headers:
                    next_col = ws_tm.max_column + 1
                    ws_tm.cell(1, next_col).value = "Regression"
                    delta_col = tm_headers.index("Score Delta") + 1 if "Score Delta" in tm_headers else None
                    for r in range(2, ws_tm.max_row + 1):
                        if ws_tm.cell(r, 1).value:
                            delta = ws_tm.cell(r, delta_col).value if delta_col else 0
                            try:
                                ws_tm.cell(r, next_col).value = bool(int(delta) < 0)
                            except (TypeError, ValueError):
                                ws_tm.cell(r, next_col).value = False
                    changed = True
                    import logging as _log
                    _log.info("[Excel] Migrated Tailored_Match_Results: added 'Regression' column.")
                # PRJ-002 PR 2: append 9 per-dimension columns. Existing rows
                # left blank — user re-runs the optimizer to populate them.
                tm_headers = [ws_tm.cell(1, c).value for c in range(1, ws_tm.max_column + 1)]
                _new_tailored_cols = [
                    "Original ATS", "Tailored ATS", "ATS Delta",
                    "Original Recruiter", "Tailored Recruiter", "Recruiter Delta",
                    "Original HM", "Tailored HM", "HM Delta",
                ]
                _missing = [c for c in _new_tailored_cols if c not in tm_headers]
                if _missing:
                    for col_name in _missing:
                        next_col = ws_tm.max_column + 1
                        ws_tm.cell(1, next_col).value = col_name
                    changed = True
                    import logging as _log
                    _log.info(f"[Excel] Migrated Tailored_Match_Results: added {_missing} columns.")
                # Add "Last Written Hash" column. Existing rows leave it blank
                # — first re-run treats them as "no expected hash" and writes
                # normally (no false-positive tamper detect).
                tm_headers = [ws_tm.cell(1, c).value for c in range(1, ws_tm.max_column + 1)]
                if "Last Written Hash" not in tm_headers:
                    next_col = ws_tm.max_column + 1
                    ws_tm.cell(1, next_col).value = "Last Written Hash"
                    changed = True
                    import logging as _log
                    _log.info("[Excel] Migrated Tailored_Match_Results: added 'Last Written Hash' column.")
            if changed:
                wb.save(xlsx_path)
        finally:
            wb.close()
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Company_List"
        ws.append(COMPANY_HEADERS)
        wb.create_sheet("Company_Without_TPM").append(WITHOUT_TPM_HEADERS)
        wb.create_sheet("JD_Tracker").append(JD_HEADERS)
        wb.create_sheet("Match_Results").append(MATCH_HEADERS)
        wb.create_sheet("Tailored_Match_Results").append(TAILORED_HEADERS)
        for name in TRIAGE_SHEETS:
            wb.create_sheet(name).append(JD_HEADERS)
        wb.save(xlsx_path)
        wb.close()
    return xlsx_path


# ── Company helpers ───────────────────────────────────────────────────────────
def count_company_rows(xlsx_path: str = EXCEL_PATH) -> int:
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Company_List"]
        return sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value)
    finally:
        wb.close()


def get_company_rows(xlsx_path: str = EXCEL_PATH) -> list:
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws   = wb["Company_List"]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value or "" for c in range(1, ws.max_column + 1)]
            if any(row):
                rows.append(row)
        return rows
    finally:
        wb.close()


def get_company_rows_with_row_num(xlsx_path: str = EXCEL_PATH) -> list[tuple[int, list]]:
    """Return [(excel_row_number, row_data), ...] skipping empty rows.
    excel_row_number is the actual 1-based row index in the sheet (2 = first data row).
    Use this when you need to write back to a specific row via update_company_career_url.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws   = wb["Company_List"]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(r, c).value or "" for c in range(1, ws.max_column + 1)]
            if any(row):
                rows.append((r, row))
        return rows
    finally:
        wb.close()


def upsert_companies(xlsx_path: str, companies_data: list):
    """Upsert companies into Company_List.

    For NEW companies: write all 9 columns with TPM counts initialized to
    [0, 0, 0, "No"].
    For EXISTING companies: only update cols 1–5 (Name, Track, Focus,
    Career URL, Updated At). Cols 6–9 (TPM Jobs, Qualified Jobs, No TPM Count,
    Auto Archived) are preserved — they are managed exclusively by
    update_company_job_counts and the auto-archival pipeline.
    """
    wb = load_workbook(xlsx_path)
    try:
        ws  = wb["Company_List"]
        idx  = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
        now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for c in companies_data:
            name = c.get("company_name", "N/A")
            cols_1_to_5 = [name, c.get("track", "N/A"), c.get("business_focus", "N/A"),
                           c.get("career_url", "N/A"), now]
            if name in idx:
                for col, val in enumerate(cols_1_to_5, 1):
                    ws.cell(idx[name], col, val)
            else:
                ws.append(cols_1_to_5 + [0, 0, 0, "No"])
                idx[name] = ws.max_row
        wb.save(xlsx_path)
    finally:
        wb.close()


def get_company_names_without_tpm(xlsx_path: str = EXCEL_PATH) -> set:
    """Return a set of company names from the Company_Without_TPM sheet."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        if "Company_Without_TPM" not in wb.sheetnames:
            return set()
        ws = wb["Company_Without_TPM"]
        names = set()
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val:
                names.add(str(val).strip())
        return names
    finally:
        wb.close()


def update_company_career_url(xlsx_path: str, excel_row: int, new_url: str):
    """excel_row is the actual 1-based Excel row number (2 = first data row after header)."""
    wb = load_workbook(xlsx_path)
    try:
        ws = wb["Company_List"]
        ws.cell(excel_row, 4, new_url)
        wb.save(xlsx_path)
    finally:
        wb.close()


def get_incomplete_company_rows(xlsx_path: str = EXCEL_PATH) -> list:
    """BUG-69: return Company_List rows whose Business Focus is blank/N/A.

    Mirrors get_incomplete_jd_rows — these rows previously never self-healed
    (unlike blank Career URLs, which run_phase_1_5 backfills every run).
    Consumed by company_agent.run_reenrich_business_focus.
    Returns [{"excel_row", "name", "career_url"}].
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Company_List"]
        out = []
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or "").strip()
            if not name:
                continue
            focus = str(ws.cell(r, 3).value or "").strip()
            if focus.lower() in ("", "n/a", "none"):
                out.append({"excel_row": r, "name": name,
                            "career_url": str(ws.cell(r, 4).value or "").strip()})
        return out
    finally:
        wb.close()


def update_company_business_focus(xlsx_path: str, excel_row: int, focus: str):
    """BUG-69: write Business Focus (col 3) + refresh Updated At (col 5).
    excel_row is the actual 1-based Excel row number (2 = first data row)."""
    wb = load_workbook(xlsx_path)
    try:
        ws = wb["Company_List"]
        ws.cell(excel_row, 3, focus)
        ws.cell(excel_row, 5, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        wb.save(xlsx_path)
    finally:
        wb.close()


def update_company_track(xlsx_path: str, excel_row: int, track: str):
    """PRJ-004 REQ-004-06: write the Track column (col 2) for one row in place.
    excel_row is the actual 1-based Excel row number (2 = first data row)."""
    wb = load_workbook(xlsx_path)
    try:
        ws = wb["Company_List"]
        ws.cell(excel_row, 2, track)
        wb.save(xlsx_path)
    finally:
        wb.close()


def sort_company_list_by_track(xlsx_path: str = EXCEL_PATH) -> int:
    """Sort Company_List rows by Track (canonical TRACK_ORDER position),
    then company name (case-insensitive). Blank/custom Track values sink to
    the bottom — never dropped, visible for manual review.

    Count-preserving in-place rewrite (sort_jd_tracker_by_tier pattern): all
    columns travel with their row, rows are never deleted/inserted. Must run
    only after every excel_row-keyed write of the run has completed — sorting
    invalidates previously captured row numbers. Returns the number of data
    rows sorted. Idempotent — safe to re-run.
    """
    track_rank = {t: i for i, t in enumerate(TRACK_ORDER)}
    wb = load_workbook(xlsx_path)
    try:
        ws     = wb["Company_List"]
        n_cols = ws.max_column
        rows = []
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, n_cols + 1)]
            if not any(v not in (None, "") for v in row_vals):
                continue
            name  = str(row_vals[0] or "")
            track = str(row_vals[1] or "").strip()
            rows.append((track_rank.get(track, len(TRACK_ORDER)),
                         name.strip().lower(), row_vals))

        rows.sort(key=lambda x: (x[0], x[1]))

        max_row_before = ws.max_row
        for r in range(2, max_row_before + 1):
            for c in range(1, n_cols + 1):
                ws.cell(r, c).value = None
        for i, (_rank, _name, row_vals) in enumerate(rows):
            for c in range(1, n_cols + 1):
                ws.cell(2 + i, c).value = row_vals[c - 1]

        wb.save(xlsx_path)
        return len(rows)
    finally:
        wb.close()


# ── Archive helpers (REQ-063) ────────────────────────────────────────────────
def get_archived_companies(xlsx_path: str = EXCEL_PATH) -> set:
    """Return set of company names where Auto Archived == 'yes'."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Company_List"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        arch_col = headers.get("Auto Archived")
        if not arch_col:
            return set()
        names = set()
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            archived = str(ws.cell(r, arch_col).value or "").strip().lower()
            if name and archived == "yes":
                names.add(str(name).strip())
        return names
    finally:
        wb.close()


def update_archive_status(xlsx_path: str, company_name: str,
                          no_tpm_count: int, archived: str) -> None:
    """Update No TPM Count and Auto Archived columns for a company."""
    wb = load_workbook(xlsx_path)
    try:
        ws = wb["Company_List"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        cnt_col  = headers.get("No TPM Count")
        arch_col = headers.get("Auto Archived")
        if not cnt_col or not arch_col:
            return
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or "").strip()
            if name == company_name:
                ws.cell(r, cnt_col).value  = no_tpm_count
                ws.cell(r, arch_col).value = archived
                break
        wb.save(xlsx_path)
    finally:
        wb.close()


def unarchive_company(xlsx_path: str, company_name: str) -> None:
    """Manually restore an archived company: reset counter and archived flag."""
    update_archive_status(xlsx_path, company_name, 0, "no")


def get_company_archive_info(xlsx_path: str = EXCEL_PATH) -> dict:
    """Return {company_name: {"no_tpm_count": int, "archived": str}} for all companies."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Company_List"]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        cnt_col  = headers.get("No TPM Count")
        arch_col = headers.get("Auto Archived")
        result = {}
        if not cnt_col or not arch_col:
            return result
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or "").strip()
            if not name:
                continue
            raw_cnt = ws.cell(r, cnt_col).value
            cnt = int(raw_cnt) if isinstance(raw_cnt, (int, float)) and raw_cnt else 0
            arch = str(ws.cell(r, arch_col).value or "").strip().lower()
            result[name] = {"no_tpm_count": cnt, "archived": arch}
        return result
    finally:
        wb.close()


# ── JD helpers ────────────────────────────────────────────────────────────────
# BUG-71: known tracking-only query params. Stripping these (plus sorting the
# survivors) makes two scrapes of the same posting compare equal even when the
# source page decorated the link differently between runs (observed: LinkedIn
# pagenum/position/refid/trackingid). Job-identifying params like gh_jid are
# NOT in this list — on embedded Greenhouse boards the query IS the job key.
_TRACKING_PARAMS = frozenset([
    "gclid", "fbclid", "gh_src", "src", "source", "lever-source",
    "ref", "refid", "referrer", "trackingid", "tracking_id",
    "pagenum", "position", "page", "origin",
])
_LINKEDIN_JOB_RE = re.compile(r"/jobs/view/(?:[^/]*?-)?(\d+)/?$")
_TESLA_JOB_RE    = re.compile(r"/careers/search/job/(?:apply/)?(?:[a-z0-9-]*?-)?(\d+)/?$")


def canonical_jd_url(url: str) -> str:
    """BUG-71: reduce a JD URL to a canonical identity string so the same
    posting discovered under cosmetically different URLs dedupes correctly.

    Rules (conservative — false negatives are acceptable, false positives
    would silently drop legitimate new jobs):
      - case-normalize host, strip www., strip trailing slash;
      - job-boards.greenhouse.io == boards.greenhouse.io (same board, two hosts);
      - linkedin.com/jobs/view/<slug>-<id> → the numeric id (query dropped:
        LinkedIn queries are pure tracking);
      - tesla.com /job/apply/<id> and /job/<slug>-<id> → the numeric id;
      - otherwise drop known tracking params, keep + sort the rest
        (gh_jid-style job keys survive).
    Comparison key only — never write it back to a sheet."""
    s = urlsplit(str(url or "").strip())
    host = s.netloc.lower().lstrip(".")
    if host.startswith("www."):
        host = host[4:]
    if host == "job-boards.greenhouse.io":
        host = "boards.greenhouse.io"
    path = s.path.rstrip("/")

    if host.endswith("linkedin.com"):
        m = _LINKEDIN_JOB_RE.search(path)
        if m:
            return f"linkedin.com/jobs/view/{m.group(1)}"
    if host.endswith("tesla.com"):
        m = _TESLA_JOB_RE.search(path.lower())
        if m:
            return f"tesla.com/careers/search/job/{m.group(1)}"

    params = [(k, v) for k, v in parse_qsl(s.query, keep_blank_values=True)
              if k.lower() not in _TRACKING_PARAMS
              and not k.lower().startswith(("utm_", "campaign"))]
    query = urlencode(sorted(params))
    return f"{host}{path}" + (f"?{query}" if query else "")


def get_triaged_jd_urls(xlsx_path: str = EXCEL_PATH) -> set:
    """Return the set of CANONICAL JD URLs (canonical_jd_url) the user has
    triaged into TRIAGE_SHEETS (BUG-65). These are final decisions: the
    pipeline must never re-scrape them or re-add them to JD_Tracker. Callers
    must canonicalize before membership tests (BUG-71). Missing tabs are
    tolerated."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        urls = set()
        c_url = _JD_COL["JD URL"]
        for name in TRIAGE_SHEETS:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            for r in range(2, ws.max_row + 1):
                url = str(ws.cell(r, c_url).value or "").strip()
                if url:
                    urls.add(canonical_jd_url(url))
        return urls
    finally:
        wb.close()


def get_jd_urls(xlsx_path: str = EXCEL_PATH) -> list:
    """Return URLs that have been successfully extracted (company != N/A/empty).
    Incomplete records (failed Gemini extraction) are excluded so they get retried."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["JD_Tracker"]
        c_url = _JD_COL["JD URL"]
        c_company = _JD_COL["Company"]
        result = []
        for r in range(2, ws.max_row + 1):
            url     = ws.cell(r, c_url).value
            company = str(ws.cell(r, c_company).value or "").strip()
            if url and company not in ("", "N/A", "JSON ERROR"):
                result.append(url)
        return result
    finally:
        wb.close()


def get_jd_url_meta(xlsx_path: str = EXCEL_PATH) -> dict:
    """Returns {url: {"hash": str, "age_days": float, "title": str}} for valid, complete JD rows.
    Skips rows where company is N/A, empty, or JSON ERROR, OR where location/tech/resp are
    missing/None/N/A (incomplete records). Incomplete records must be re-processed."""
    wb  = load_workbook(xlsx_path, read_only=True)
    try:
        ws  = wb["JD_Tracker"]
        c_url  = _JD_COL["JD URL"]
        c_title = _JD_COL["Job Title"]
        c_company = _JD_COL["Company"]
        c_location = _JD_COL["Location"]
        c_req  = _JD_COL["Requirements"]
        c_resp = _JD_COL["Responsibilities"]
        c_updated = _JD_COL["Updated At"]
        c_hash = _JD_COL["MD Hash"]
        now = datetime.now()
        result = {}
        for r in range(2, ws.max_row + 1):
            url      = ws.cell(r, c_url).value
            title    = str(ws.cell(r, c_title).value or "").strip()
            company  = str(ws.cell(r, c_company).value or "").strip()
            location = str(ws.cell(r, c_location).value or "").strip()
            req      = str(ws.cell(r, c_req).value or "").strip()
            resp     = str(ws.cell(r, c_resp).value or "").strip()
            updated  = ws.cell(r, c_updated).value
            md_hash  = str(ws.cell(r, c_hash).value or "").strip()
            if not url or company in ("", "N/A", "JSON ERROR"):
                continue
            # Also skip records with missing key fields — treat as incomplete so main loop retries them
            if (location.lower() in _JD_MISSING
                    or req.lower() in _JD_MISSING
                    or resp.lower() in _JD_MISSING):
                continue
            if isinstance(updated, datetime):
                age_days = (now - updated).total_seconds() / 86400
            elif isinstance(updated, str) and updated:
                try:
                    age_days = (now - datetime.strptime(updated, "%Y-%m-%d %H:%M:%S")).total_seconds() / 86400
                except Exception:
                    age_days = 999.0
            else:
                age_days = 999.0
            result[url] = {"hash": md_hash, "age_days": age_days, "title": title}
        return result
    finally:
        wb.close()


def batch_update_jd_timestamps(xlsx_path: str, urls: list) -> int:
    """Update only Updated At column for given URLs. Returns count updated."""
    if not urls:
        return 0
    url_set = set(urls)
    wb  = load_workbook(xlsx_path)
    try:
        ws  = wb["JD_Tracker"]
        c_url = _JD_COL["JD URL"]
        c_updated = _JD_COL["Updated At"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        count = 0
        for r in range(2, ws.max_row + 1):
            url = ws.cell(r, c_url).value
            if url in url_set:
                ws.cell(r, c_updated).value = now
                count += 1
        wb.save(xlsx_path)
        return count
    finally:
        wb.close()


def get_jd_rows_for_match(xlsx_path: str = EXCEL_PATH) -> list:
    """
    Returns list of dicts for every valid JD row (REQ-004-21: all written rows
    are already domain-qualified at write time, so there is no boolean filter —
    only integrity checks: URL present, company real, extraction not failed).
    Each dict carries the row's Job Domain so the match layer can route to the
    per-track prompt pair. Reconstructs a match-ready JSON string from columns.
    """
    import logging as _log
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws   = wb["JD_Tracker"]
        c_url   = _JD_COL["JD URL"]
        c_title = _JD_COL["Job Title"]
        c_company = _JD_COL["Company"]
        c_loc   = _JD_COL["Location"]
        c_sal   = _JD_COL["Salary"]
        c_req   = _JD_COL["Requirements"]
        c_addq  = _JD_COL["Additional Qualifications"]
        c_resp  = _JD_COL["Responsibilities"]
        c_domain = _JD_COL["Job Domain"]
        c_dq    = _JD_COL["Data Quality"]
        # PRJ-002 Phase 4 fix: dynamic lookup so workbooks pre-migration (no column)
        # don't crash here — they yield ats_keywords=[] which the matcher handles.
        ws_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        c_ats = ws_headers.index("ATS Keywords") + 1 if "ATS Keywords" in ws_headers else None
        rows = []
        for r in range(2, ws.max_row + 1):
            url        = ws.cell(r, c_url).value
            job_title  = ws.cell(r, c_title).value or ""
            company    = str(ws.cell(r, c_company).value or "").strip()
            location   = ws.cell(r, c_loc).value or ""
            salary     = ws.cell(r, c_sal).value or ""
            req_raw    = ws.cell(r, c_req).value or ""
            addq_raw   = ws.cell(r, c_addq).value or ""
            resp_raw   = ws.cell(r, c_resp).value or ""
            job_domain = str(ws.cell(r, c_domain).value or "").strip()
            dq         = str(ws.cell(r, c_dq).value or "").strip().lower()
            if not url or company in ("", "N/A", "JSON ERROR") or dq == "failed":
                continue
            if job_domain not in JOB_DOMAIN_VALUES:
                # Should not occur post-G2; keep the row rather than drop it,
                # score it with the AI pair, and surface the anomaly.
                _log.warning(f"[Match] JD row {url} has invalid Job Domain "
                             f"{job_domain!r} — falling back to 'AI'.")
                job_domain = "AI"
            req_list  = [s.lstrip("• ").strip() for s in req_raw.split("\n") if s.strip()]
            addq_list = [s.lstrip("• ").strip() for s in addq_raw.split("\n") if s.strip()]
            resp_list = [s.lstrip("• ").strip() for s in resp_raw.split("\n") if s.strip()]
            ats_raw   = ws.cell(r, c_ats).value if c_ats else ""
            ats_list  = [s.lstrip("• ").strip() for s in (ats_raw or "").split("\n") if s.strip()]
            jd_json   = json.dumps({
                "job_title": job_title, "company": company, "location": location,
                "salary_range": salary, "requirements": req_list,
                "additional_qualifications": addq_list,
                "key_responsibilities": resp_list, "job_domain": job_domain,
                "ats_keywords": ats_list,
            })
            rows.append({"url": url, "jd_json": jd_json, "job_domain": job_domain})
        return rows
    finally:
        wb.close()


def _jd_row_data(jd_url: str, d: dict, now: str, markdown_hash: str) -> list:
    """Build the 20-column JD_Tracker row (single source for both upsert paths).
    Column order must match JD_HEADERS exactly."""
    req   = "\n".join(f"• {x}" for x in (d.get("requirements") or [])) or "None"
    addq  = "\n".join(f"• {x}" for x in (d.get("additional_qualifications") or [])) or "None"
    resp  = "\n".join(f"• {x}" for x in (d.get("key_responsibilities") or [])) or "None"
    ats   = "\n".join(f"• {x}" for x in (d.get("ats_keywords") or [])) or "None"
    posted = str(d.get("posted_date") or "")
    return [jd_url, d.get("job_title", "N/A"), d.get("company", "N/A"),
            d.get("location", "N/A"), d.get("salary_range", "N/A"),
            req, addq, resp,
            d.get("job_domain", ""), now, markdown_hash,
            d.get("data_quality", ""), ats,
            None,                                  # Sort Tier — written by sort_jd_tracker_by_tier
            posted,
            compute_freshness_tier(posted),        # recomputed again at every sort
            d.get("min_yoe"),
            d.get("yoe_flag", ""),
            d.get("work_auth", ""),
            d.get("date_flag", "")]


def _JD_ERROR_ROW(jd_url: str, now: str, markdown_hash: str) -> list:
    return [jd_url, "JSON ERROR", "JSON ERROR", "JSON ERROR", "JSON ERROR",
            "JSON ERROR", "JSON ERROR", "JSON ERROR", "JSON ERROR", now, markdown_hash,
            "failed", "None", None, "", None, None, "", "", ""]


def upsert_jd_record(xlsx_path: str, jd_url: str, jd_json: str, markdown_hash: str):
    wb = load_workbook(xlsx_path)
    try:
        ws   = wb["JD_Tracker"]
        idx  = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
        now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            d     = json.loads(jd_json)
            row_data = _jd_row_data(jd_url, d, now, markdown_hash)
        except json.JSONDecodeError:
            row_data = _JD_ERROR_ROW(jd_url, now, markdown_hash)
        if jd_url in idx:
            for col, val in enumerate(row_data, 1):
                ws.cell(idx[jd_url], col, val)
        else:
            ws.append(row_data)
        wb.save(xlsx_path)
    finally:
        wb.close()


def batch_upsert_jd_records(xlsx_path: str, records: list) -> int:
    """
    Write multiple JD records in a single load→modify→save cycle.
    records: list of (jd_url, jd_json, markdown_hash) tuples.
    Returns the number of records written.
    """
    if not records:
        return 0
    wb  = load_workbook(xlsx_path)
    try:
        ws  = wb["JD_Tracker"]
        # BUG-71: index by canonical URL so the same posting rediscovered under
        # a cosmetically different URL updates its existing row (never a dup).
        idx = {canonical_jd_url(ws.cell(r, 1).value): r
               for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for jd_url, jd_json, markdown_hash in records:
            try:
                d     = json.loads(jd_json)
                row_data = _jd_row_data(jd_url, d, now, markdown_hash)
            except json.JSONDecodeError:
                row_data = _JD_ERROR_ROW(jd_url, now, markdown_hash)
            canon = canonical_jd_url(jd_url)
            if canon in idx:
                for col, val in enumerate(row_data, 1):
                    ws.cell(idx[canon], col, val)
            else:
                ws.append(row_data)
                idx[canon] = ws.max_row
        wb.save(xlsx_path)
        return len(records)
    finally:
        wb.close()


# ── JD completeness helpers ───────────────────────────────────────────────────
_JD_MISSING = {"", "n/a", "none", "json error", "not specified", "not available"}

def get_incomplete_jd_rows(xlsx_path: str = EXCEL_PATH) -> list:
    """
    Return list of dicts for JD records that are missing key fields
    (location, tech_stack, or responsibilities are empty/N/A/None/JSON ERROR).
    These records should be retried regardless of whether their URL is known.
    """
    wb  = load_workbook(xlsx_path, read_only=True)
    try:
        ws  = wb["JD_Tracker"]
        c_url  = _JD_COL["JD URL"]
        c_title = _JD_COL["Job Title"]
        c_company = _JD_COL["Company"]
        c_loc  = _JD_COL["Location"]
        c_req  = _JD_COL["Requirements"]
        c_resp = _JD_COL["Responsibilities"]
        c_posted = _JD_COL["Posted Date"]
        c_dflag  = _JD_COL["Date Flag"]
        out = []
        for r in range(2, ws.max_row + 1):
            url      = ws.cell(r, c_url).value
            if not url:
                continue
            title    = str(ws.cell(r, c_title).value or "").strip()
            company  = str(ws.cell(r, c_company).value or "").strip()
            location = str(ws.cell(r, c_loc).value or "").strip()
            req      = str(ws.cell(r, c_req).value or "").strip()
            resp     = str(ws.cell(r, c_resp).value or "").strip()
            if (location.lower() in _JD_MISSING
                    or req.lower() in _JD_MISSING
                    or resp.lower() in _JD_MISSING):
                # BUG-67: carry the existing date fields so the retry path can
                # preserve them instead of wiping the row's Posted Date.
                posted_v = ws.cell(r, c_posted).value
                if isinstance(posted_v, datetime):
                    posted_v = posted_v.strftime("%Y-%m-%d")
                out.append({"url": url, "title": title, "company": company,
                            "posted_date": str(posted_v or "").strip(),
                            "date_flag": str(ws.cell(r, c_dflag).value or "").strip()})
        return out
    finally:
        wb.close()


def count_tpm_jobs_by_company(xlsx_path: str = EXCEL_PATH) -> dict:
    """
    Return {company_name: {"tpm": int, "qualified": int}} counting all valid JD
    rows. "qualified" counts rows whose Job Domain is one of the 5 track values
    (under REQ-004-09 that is every successfully classified row).
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws     = wb["JD_Tracker"]
        c_url = _JD_COL["JD URL"]
        c_company = _JD_COL["Company"]
        c_domain = _JD_COL["Job Domain"]
        counts = {}
        for r in range(2, ws.max_row + 1):
            url     = ws.cell(r, c_url).value
            company = str(ws.cell(r, c_company).value or "").strip()
            domain  = str(ws.cell(r, c_domain).value or "").strip()
            if not url or company.lower() in ("", "n/a", "json error"):
                continue
            if company not in counts:
                counts[company] = {"tpm": 0, "qualified": 0}
            counts[company]["tpm"] += 1
            if domain in JOB_DOMAIN_VALUES:
                counts[company]["qualified"] += 1
        return counts
    finally:
        wb.close()


def count_valid_tpm_jobs_by_company(xlsx_path: str = EXCEL_PATH) -> dict:
    """Return {company_name: int} counting JD rows where data_quality != 'failed'.

    Used by REQ-063 archive logic: only non-failed records count toward
    determining whether a company has TPM jobs.
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["JD_Tracker"]
        c_url = _JD_COL["JD URL"]
        c_company = _JD_COL["Company"]
        c_dq = _JD_COL["Data Quality"]
        counts = {}
        for r in range(2, ws.max_row + 1):
            url     = ws.cell(r, c_url).value
            company = str(ws.cell(r, c_company).value or "").strip()
            if not url or company.lower() in ("", "n/a", "json error"):
                continue
            if c_dq:
                dq = str(ws.cell(r, c_dq).value or "").strip().lower()
                if dq == "failed":
                    continue
            counts[company] = counts.get(company, 0) + 1
        return counts
    finally:
        wb.close()


def update_company_job_counts(xlsx_path: str, counts: dict) -> None:
    """
    Write TPM Jobs / Qualified Jobs columns in Company_List.
    counts: {company_name: {"tpm": int, "qualified": int}}
    """
    wb = load_workbook(xlsx_path)
    try:
        ws_co   = wb["Company_List"]
        headers = {ws_co.cell(1, c).value: c for c in range(1, ws_co.max_column + 1)}
        tpm_col   = headers.get("TPM Jobs")
        qual_col  = headers.get("Qualified Jobs")
        if not tpm_col or not qual_col:
            return
        for r in range(2, ws_co.max_row + 1):
            name = str(ws_co.cell(r, 1).value or "").strip()
            if name in counts:
                ws_co.cell(r, tpm_col).value  = counts[name]["tpm"]
                ws_co.cell(r, qual_col).value = counts[name]["qualified"]
        wb.save(xlsx_path)
    finally:
        wb.close()


# ── Location tier (sort + highlight) ─────────────────────────────────────────
# Greater Seattle / Puget Sound cities. A location segment qualifies as
# "Greater Seattle" only if it pairs one of these names with the WA state
# token (", WA") — guards against false positives like "Kent, OH".
_GREATER_SEATTLE_CITIES = (
    "seattle", "bellevue", "redmond", "kirkland", "bothell",
    "sammamish", "issaquah", "renton", "tukwila", "kent",
    "lynnwood", "everett", "tacoma",
)

# "Washington" without a city pairing is ambiguous (state vs D.C.). We accept
# only forms that explicitly qualify it as the state — never bare "Washington".
_WASHINGTON_STATE_FORMS = frozenset({
    "washington state",
    "washington, us",
    "washington, usa",
    "washington, united states",
    "washington, u.s.",
    "washington, u.s.a.",
    "washington, united states of america",
})

# Remote tier accepts: bare "Remote" (US-default) + any explicit US qualifier.
# Explicit non-US country (e.g. "Remote, Canada", "Remote, UK") → Other.
_US_REMOTE_QUALIFIERS = frozenset({
    "us", "usa", "united states", "u.s.", "u.s.a.", "united states of america",
})

_TIER_PRIORITY = {"Greater Seattle": 0, "Remote": 1, "Other": 2}

# PRJ-004 REQ-004-12: CA/TX detection for the tightened geo filter. City hints
# are unambiguous major metros (design §2.2.4); state-token pairing (", ca")
# and full state names also qualify. SoCal cities included deliberately —
# space/defense hiring concentrates there (El Segundo, Hawthorne, Long Beach…).
_CA_CITY_HINTS = (
    "san francisco", "palo alto", "mountain view", "sunnyvale", "san jose",
    "santa clara", "menlo park", "cupertino", "oakland", "berkeley", "fremont",
    "redwood city", "south san francisco", "los angeles", "irvine", "san diego",
    "santa monica", "el segundo", "hawthorne", "long beach", "torrance",
    "pasadena", "culver city", "sacramento",
)
_TX_CITY_HINTS = (
    "austin", "dallas", "houston", "san antonio", "fort worth", "plano",
    "irving", "richardson", "el paso",
)

# Sort-group precedence inside compute_sort_tier / best-region selection.
_REGION_PRIORITY = {"Seattle": 0, "Remote": 1, "CA": 2, "TX": 3, "Other": 8, "Unknown": 9}

_SORT_TIER_FILL = {
    1: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # Excel "Good" green
    2: PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    3: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Excel "Neutral" yellow
    4: PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}
_NO_FILL = PatternFill(fill_type=None)


def classify_location(location: str) -> str:
    """Return 'Greater Seattle' / 'Remote' / 'Other' for a JD Location string.

    Greater Seattle: any semicolon-separated segment that either
      - names a Puget Sound city paired with ", WA" (e.g. "Bellevue, WA"), or
      - matches a whitelisted Washington-state form (e.g. "Washington, US",
        "Washington, USA", "Washington State"). Bare "Washington" is NOT
        accepted because it collides with Washington, D.C.

    Remote: any segment that is bare "Remote" (US-default) OR "Remote, <q>"
    where <q> ∈ {US, USA, United States, U.S., U.S.A., United States of
    America}. Explicit non-US country qualifiers (e.g. "Remote, Canada")
    do NOT qualify.

    Greater Seattle wins when both signals are present.
    """
    if not location:
        return "Other"
    text = str(location).strip()
    if not text or text.lower() in _JD_MISSING:
        return "Other"
    has_seattle = False
    has_remote  = False
    for raw_seg in text.split(";"):
        seg = raw_seg.strip().lower()
        if not seg:
            continue
        # Greater Seattle: city + ", WA" pairing
        if ", wa" in seg and any(city in seg for city in _GREATER_SEATTLE_CITIES):
            has_seattle = True
        # Greater Seattle: explicitly state-qualified Washington (whitelist)
        elif seg in _WASHINGTON_STATE_FORMS:
            has_seattle = True
        # Remote: bare "remote" (US-default) or US-qualified form
        if seg == "remote":
            has_remote = True
        elif seg.startswith("remote,"):
            qualifier = seg[len("remote,"):].strip()
            if qualifier in _US_REMOTE_QUALIFIERS:
                has_remote = True
            # explicit non-US country (e.g. "remote, canada") → not remote
    if has_seattle:
        return "Greater Seattle"
    if has_remote:
        return "Remote"
    return "Other"


def _classify_region_segment(seg: str) -> str:
    """Classify one lowercase location segment → Seattle/Remote/CA/TX/Other."""
    if ", wa" in seg and any(city in seg for city in _GREATER_SEATTLE_CITIES):
        return "Seattle"
    if seg in _WASHINGTON_STATE_FORMS:
        return "Seattle"
    if seg == "remote":
        return "Remote"
    if seg.startswith("remote,"):
        qualifier = seg[len("remote,"):].strip()
        if qualifier in _US_REMOTE_QUALIFIERS:
            return "Remote"
        return "Other"  # explicit non-US remote (e.g. "Remote, Canada")
    if ", ca" in seg or "california" in seg or any(c in seg for c in _CA_CITY_HINTS):
        return "CA"
    if ", tx" in seg or "texas" in seg or any(c in seg for c in _TX_CITY_HINTS):
        return "TX"
    return "Other"


def classify_region(location: str) -> str:
    """PRJ-004 REQ-004-12: classify a JD Location string into
    'Seattle' / 'Remote' / 'CA' / 'TX' / 'Other' / 'Unknown'.

    Multi-location strings ('; '-separated) qualify if ANY segment qualifies;
    the best region wins (precedence Seattle > Remote > CA > TX) so the sort
    tier reflects the most desirable posting location. Blank / placeholder
    locations return 'Unknown' — callers keep those rows (conservative: only
    confirmed out-of-region rows are dropped; see BUG-66).
    """
    if not location:
        return "Unknown"
    text = str(location).strip()
    if not text or text.lower() in _JD_MISSING:
        return "Unknown"
    best = "Other"
    for raw_seg in text.split(";"):
        seg = raw_seg.strip().lower()
        if not seg:
            continue
        region = _classify_region_segment(seg)
        if _REGION_PRIORITY[region] < _REGION_PRIORITY[best]:
            best = region
    return best


def compute_freshness_tier(posted_date: str, today: date | None = None) -> int | None:
    """PRJ-004 REQ-004-10: freshness tier from a 'YYYY-MM-DD' posted date.

    age 0–2d → 1; 3–7d → 2; 8–14d → 3; ≥15d, blank, or unparsable → None.
    The ≤14-day tier ceiling matches the write-time keep gate exactly (pinned
    boundary — see design.md §2.2.3): every kept dated posting is tierable.
    Future-dated postings (clock skew) are treated as age 0.
    """
    if not posted_date:
        return None
    try:
        d = datetime.strptime(str(posted_date).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None
    age = ((today or datetime.now().date()) - d).days
    if age < 0:
        age = 0
    if age <= 2:
        return 1
    if age <= 7:
        return 2
    if age <= 14:
        return 3
    return None


def compute_sort_tier(freshness_tier: int | None, region: str) -> int:
    """PRJ-004 REQ-004-15: combined 1–6 sort tier (freshness primary,
    Seattle+Remote > CA/TX secondary):

        (T1, Sea/Rem)=1  (T1, CA/TX)=2
        (T2, Sea/Rem)=3  (T2, CA/TX)=4
        (T3, Sea/Rem)=5  (T3, CA/TX)=6

    Unknown-date rows, aged grandfathered rows, and Other/Unknown regions → 9
    (sink to the bottom, visible for manual review — never deleted).
    """
    if freshness_tier not in (1, 2, 3):
        return 9
    if region in ("Seattle", "Remote"):
        group = 0
    elif region in ("CA", "TX"):
        group = 1
    else:
        return 9
    return (freshness_tier - 1) * 2 + 1 + group


def sort_jd_tracker_by_tier(xlsx_path: str = EXCEL_PATH) -> int:
    """Sort JD_Tracker rows by the combined 1–6 Sort Tier and apply highlights.

    Per row: recompute Freshness Tier from Posted Date (tiers drift daily — the
    scrape-time value alone goes stale), classify the region, compute Sort Tier,
    write both columns back, then sort by (Sort Tier asc, Posted Date desc,
    Updated At desc). Tier-9 rows (unknown date / aged / Other region) sink to
    the bottom. Fills: tiers 1–2 green, 3–4 yellow, 5+ none.

    Count-preserving: rows are rewritten in place, never deleted (REQ-004-16).
    Returns the number of data rows sorted. Idempotent — safe to re-run.
    """
    wb = load_workbook(xlsx_path)
    try:
        ws = wb["JD_Tracker"]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "Sort Tier" not in headers:
            ws.cell(1, ws.max_column + 1).value = "Sort Tier"
            headers.append("Sort Tier")
        n_cols  = len(headers)
        c_loc   = headers.index("Location") + 1
        c_upd   = headers.index("Updated At") + 1
        c_tier  = headers.index("Sort Tier") + 1
        c_post  = headers.index("Posted Date") + 1 if "Posted Date" in headers else None
        c_fresh = headers.index("Freshness Tier") + 1 if "Freshness Tier" in headers else None

        today = datetime.now().date()
        rows = []
        for r in range(2, ws.max_row + 1):
            url = ws.cell(r, 1).value
            if not url:
                continue
            row_vals = [ws.cell(r, c).value for c in range(1, n_cols + 1)]
            location = str(row_vals[c_loc - 1] or "")
            updated  = str(row_vals[c_upd - 1] or "")
            posted   = str(row_vals[c_post - 1] or "") if c_post else ""
            fresh    = compute_freshness_tier(posted, today)
            if c_fresh:
                row_vals[c_fresh - 1] = fresh
            tier     = compute_sort_tier(fresh, classify_region(location))
            row_vals[c_tier - 1] = tier
            rows.append((tier, posted, updated, row_vals))

        # Stable three-pass sort: Updated At desc, Posted Date desc, tier asc.
        rows.sort(key=lambda x: x[2], reverse=True)
        rows.sort(key=lambda x: x[1], reverse=True)
        rows.sort(key=lambda x: x[0])

        # Clear all existing data rows (values + fills) before rewriting.
        max_row_before = ws.max_row
        for r in range(2, max_row_before + 1):
            for c in range(1, n_cols + 1):
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = _NO_FILL

        # Rewrite sorted rows + apply tier fill.
        for i, (tier, _post, _upd, row_vals) in enumerate(rows):
            r = 2 + i
            for c in range(1, n_cols + 1):
                ws.cell(r, c).value = row_vals[c - 1]
            fill = _SORT_TIER_FILL.get(tier)
            if fill is not None:
                for c in range(1, n_cols + 1):
                    ws.cell(r, c).fill = fill

        wb.save(xlsx_path)
        return len(rows)
    finally:
        wb.close()


# ── Match helpers ─────────────────────────────────────────────────────────────
def get_match_pairs(xlsx_path: str = EXCEL_PATH) -> dict:
    """Returns {(resume_id, jd_url): {"score": int, "hash": str, "stage": str}}."""
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws    = wb["Match_Results"]
        pairs = {}
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, 1).value
            url = ws.cell(r, 2).value
            if rid and url:
                score  = ws.cell(r, 3).value or 0
                rhash  = (ws.cell(r, 8).value or "") if ws.max_column >= 8 else ""
                stage  = (ws.cell(r, 9).value or "fine") if ws.max_column >= 9 else "fine"
                pairs[(str(rid), str(url))] = {
                    "score": int(score) if isinstance(score, (int, float)) else 0,
                    "hash":  str(rhash),
                    "stage": str(stage),
                }
        return pairs
    finally:
        wb.close()


def upsert_match_record(xlsx_path: str, resume_id: str, jd_url: str, match_json: str,
                        resume_hash: str = "", stage: str = "fine"):
    wb  = load_workbook(xlsx_path)
    try:
        ws  = wb["Match_Results"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            d         = json.loads(match_json)
            strengths = "\n".join(f"• {x}" for x in d.get("key_strengths", [])) or "None"
            gaps      = "\n".join(f"• {x}" for x in d.get("critical_gaps", [])) or "None"
            row_data  = [resume_id, jd_url, d.get("compatibility_score", 0),
                         strengths, gaps, d.get("recommendation_reason","N/A"), now,
                         resume_hash, stage]
        except json.JSONDecodeError:
            row_data = [resume_id, jd_url, 0, "JSON ERROR", "JSON ERROR", "JSON ERROR", now,
                        resume_hash, stage]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == resume_id and ws.cell(r, 2).value == jd_url:
                for col, val in enumerate(row_data, 1):
                    ws.cell(r, col, val)
                wb.save(xlsx_path)
                return
        ws.append(row_data)
        wb.save(xlsx_path)
    finally:
        wb.close()


def batch_upsert_match_records(xlsx_path: str, records: list) -> int:
    """
    Write multiple match records in a single load→modify→save cycle.

    Accepts two record shapes (interchangeable within a single call):

      Legacy 5-tuple (back-compat):
          (resume_id, jd_url, match_json, resume_hash, stage)
          Writes the 9 legacy columns; new 3-dim columns left untouched.

      Dict (PRJ-002 PR 3+):
          {
            "resume_id": str, "jd_url": str, "match_json": str,
            "resume_hash": str, "stage": str,
            # all of the following are OPTIONAL — keys absent from the dict
            # leave the corresponding Excel column unchanged so Stage 1 / Stage 2
            # writes can update only their own dimension:
            "ats_coverage_percent": float | None,    # written if KEY PRESENT
            "ats_missing":          list[str],       # joined ', ', top 5
            "recruiter_score":      int | None,
            "hm_score":             int | None,
          }

    "Key absent" semantics matter for Stage 2: it writes a fine record
    that updates HM Score, Strengths, Gaps, Reason, and Stage="fine" — but
    must NOT clobber the ATS Coverage % / Recruiter Score / ATS Missing
    that Stage 1 already wrote.
    """
    if not records:
        return 0
    wb  = load_workbook(xlsx_path)
    try:
        ws  = wb["Match_Results"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        idx = {}
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, 1).value
            url = ws.cell(r, 2).value
            if rid and url:
                idx[(str(rid), str(url))] = r
        # Resolve dynamic column positions for new dim columns. They may not
        # exist yet on a freshly migrated file from before the PRJ-002 migration
        # block ran (defensive); we skip writes for missing columns silently.
        ws_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_ats_pct = ws_headers.index("ATS Coverage %") + 1 if "ATS Coverage %" in ws_headers else None
        col_recruit = ws_headers.index("Recruiter Score") + 1 if "Recruiter Score" in ws_headers else None
        col_hm      = ws_headers.index("HM Score") + 1 if "HM Score" in ws_headers else None
        col_atsmiss = ws_headers.index("ATS Missing") + 1 if "ATS Missing" in ws_headers else None

        for rec in records:
            if isinstance(rec, dict):
                resume_id   = rec["resume_id"]
                jd_url      = rec["jd_url"]
                match_json  = rec["match_json"]
                resume_hash = rec.get("resume_hash", "")
                stage       = rec.get("stage", "fine")
            else:
                resume_id, jd_url, match_json, resume_hash, stage = rec
            try:
                d         = json.loads(match_json)
                strengths = "\n".join(f"• {x}" for x in d.get("key_strengths", [])) or "None"
                gaps      = "\n".join(f"• {x}" for x in d.get("critical_gaps", [])) or "None"
                row_data  = [resume_id, jd_url, d.get("compatibility_score", 0),
                             strengths, gaps, d.get("recommendation_reason","N/A"), now,
                             resume_hash, stage]
            except json.JSONDecodeError:
                row_data = [resume_id, jd_url, 0, "JSON ERROR", "JSON ERROR", "JSON ERROR", now,
                            resume_hash, stage]
            key = (str(resume_id), str(jd_url))
            if key in idx:
                target_row = idx[key]
                for col, val in enumerate(row_data, 1):
                    ws.cell(target_row, col, val)
            else:
                ws.append(row_data)
                idx[key] = ws.max_row
                target_row = idx[key]
            # Optional 3-dim writes — only when key is PRESENT in dict (preserve
            # otherwise). Tuple records skip this block entirely.
            if isinstance(rec, dict):
                if "ats_coverage_percent" in rec and col_ats_pct:
                    ws.cell(target_row, col_ats_pct, rec["ats_coverage_percent"])
                if "recruiter_score" in rec and col_recruit:
                    ws.cell(target_row, col_recruit, rec["recruiter_score"])
                if "hm_score" in rec and col_hm:
                    ws.cell(target_row, col_hm, rec["hm_score"])
                if "ats_missing" in rec and col_atsmiss:
                    missing = rec["ats_missing"] or []
                    # Empty list → None (openpyxl reads back "" as None anyway).
                    ws.cell(target_row, col_atsmiss,
                            ", ".join(missing[:5]) if missing else None)
        wb.save(xlsx_path)
        return len(records)
    finally:
        wb.close()


# ── Tailored match helpers ───────────────────────────────────────────────────
def get_scored_matches(xlsx_path: str = EXCEL_PATH,
                       stage: str | None = "fine") -> list:
    """Return match records from Match_Results.

    By default only returns Stage 2 ('fine') Gemini-scored rows. Stage 1
    ('coarse') heuristic scores are excluded because they are not
    comparable to the tailored re-score (which always uses the fine prompt);
    mixing them produces systemically inflated improvement deltas in
    Tailored_Match_Results.

    Pass stage=None to include all stages (used by tests / migration).

    Returns list of dicts:
      {
        resume_id, jd_url, score, stage, resume_hash,
        # PRJ-002 PR 4: per-dim originals for the optimizer's delta calc.
        # Any of these may be None if the column doesn't exist or the row's
        # value is blank (legacy data — caller falls back to `score`).
        ats_coverage_percent, recruiter_score, hm_score,
      }
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb["Match_Results"]
        # Dynamic column lookup so old workbooks (pre-PR 2 migration) still work.
        ws_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_ats_pct = ws_headers.index("ATS Coverage %") + 1 if "ATS Coverage %" in ws_headers else None
        col_recruit = ws_headers.index("Recruiter Score") + 1 if "Recruiter Score" in ws_headers else None
        col_hm      = ws_headers.index("HM Score") + 1 if "HM Score" in ws_headers else None
        results = []
        for r in range(2, ws.max_row + 1):
            rid   = ws.cell(r, 1).value
            url   = ws.cell(r, 2).value
            score = ws.cell(r, 3).value or 0
            rhash = (ws.cell(r, 8).value or "") if ws.max_column >= 8 else ""
            row_stage = (ws.cell(r, 9).value or "fine") if ws.max_column >= 9 else "fine"
            if not (rid and url and isinstance(score, (int, float)) and int(score) >= 0):
                continue
            if stage is not None and str(row_stage) != stage:
                continue
            ats_pct  = ws.cell(r, col_ats_pct).value if col_ats_pct else None
            recruit  = ws.cell(r, col_recruit).value if col_recruit else None
            hm_score = ws.cell(r, col_hm).value if col_hm else None
            results.append({
                "resume_id": str(rid),
                "jd_url": str(url),
                "score": int(score),
                "stage": str(row_stage),
                "resume_hash": str(rhash),
                "ats_coverage_percent": (
                    float(ats_pct) if isinstance(ats_pct, (int, float)) else None
                ),
                "recruiter_score": (
                    int(recruit) if isinstance(recruit, (int, float)) else None
                ),
                "hm_score": (
                    int(hm_score) if isinstance(hm_score, (int, float)) else None
                ),
            })
        return results
    finally:
        wb.close()


def get_tailored_match_pairs(xlsx_path: str = EXCEL_PATH) -> dict:
    """Returns {(resume_id, jd_url): {"tailored_score": int, "resume_hash": str,
                                       "last_written_hash": str}}.

    `last_written_hash` is the sha256 the optimizer recorded the last time it
    wrote the .md file. Empty string for legacy rows (before this column was
    added) — the optimizer treats empty as "no expected hash, safe to write".
    """
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        if "Tailored_Match_Results" not in wb.sheetnames:
            return {}
        ws    = wb["Tailored_Match_Results"]
        ws_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_lwh = ws_headers.index("Last Written Hash") + 1 if "Last Written Hash" in ws_headers else None
        pairs = {}
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, 1).value
            url = ws.cell(r, 2).value
            if rid and url:
                t_score = ws.cell(r, 6).value or 0
                rhash   = ws.cell(r, 11).value or ""
                lwh     = (ws.cell(r, col_lwh).value or "") if col_lwh else ""
                pairs[(str(rid), str(url))] = {
                    "tailored_score": int(t_score) if isinstance(t_score, (int, float)) else 0,
                    "resume_hash": str(rhash),
                    "last_written_hash": str(lwh),
                }
        return pairs
    finally:
        wb.close()


def batch_upsert_tailored_records(xlsx_path: str, records: list) -> int:
    """
    Write multiple tailored match records in a single load→modify→save cycle.

    records: list of dicts. Required keys:
        resume_id, jd_url

    Legacy (single-dimension) keys:
        job_title, company, original_score, tailored_score, score_delta,
        tailored_resume_path, optimization_summary, resume_hash, regression

    PRJ-002 PR 4 — per-dim keys (optional):
        original_ats, tailored_ats, ats_delta,
        original_recruiter, tailored_recruiter, recruiter_delta,
        original_hm, tailored_hm, hm_delta

    Regression semantics (PRJ-002 REQ-108):
      * If caller passes `regression` explicitly → use it.
      * Else if `hm_delta` is present → regression = (hm_delta < 0).
        ATS / Recruiter delta < 0 do NOT trigger regression — the ATS
        dimension can drop when the tailor reshuffles emphasis.
      * Else (legacy single-score path) → regression = (score_delta < 0).

    Per-dim columns are written when keys are PRESENT in the dict (None is
    a valid "no data" value). Missing keys leave the corresponding cell
    untouched on update; on insert, the cell stays blank.
    """
    if not records:
        return 0
    wb = load_workbook(xlsx_path)
    try:
        ws  = wb["Tailored_Match_Results"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        idx = {}
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, 1).value
            url = ws.cell(r, 2).value
            if rid and url:
                idx[(str(rid), str(url))] = r
        ws_headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        # Per-dim column index lookup (None when column missing).
        _dim_cols = {
            name: (ws_headers.index(name) + 1 if name in ws_headers else None)
            for name in (
                "Original ATS", "Tailored ATS", "ATS Delta",
                "Original Recruiter", "Tailored Recruiter", "Recruiter Delta",
                "Original HM", "Tailored HM", "HM Delta",
                "Last Written Hash",
            )
        }
        for rec in records:
            # Regression precedence: explicit > hm_delta > score_delta (legacy).
            regression = rec.get("regression")
            if regression is None:
                if "hm_delta" in rec and rec["hm_delta"] is not None:
                    regression = bool(int(rec["hm_delta"]) < 0)
                else:
                    regression = bool(int(rec.get("score_delta", 0)) < 0)
            row_data = [
                rec["resume_id"], rec["jd_url"], rec.get("job_title", ""),
                rec.get("company", ""), rec.get("original_score", 0),
                rec.get("tailored_score", 0), rec.get("score_delta", 0),
                rec.get("tailored_resume_path", ""),
                rec.get("optimization_summary", ""), now,
                rec.get("resume_hash", ""), bool(regression),
            ]
            key = (str(rec["resume_id"]), str(rec["jd_url"]))
            if key in idx:
                target_row = idx[key]
                for col, val in enumerate(row_data, 1):
                    ws.cell(target_row, col, val)
            else:
                ws.append(row_data)
                idx[key] = ws.max_row
                target_row = idx[key]
            # PR 4: per-dim writes — only for keys PRESENT in dict (None ok).
            _dim_keys = (
                ("original_ats",        "Original ATS"),
                ("tailored_ats",        "Tailored ATS"),
                ("ats_delta",           "ATS Delta"),
                ("original_recruiter",  "Original Recruiter"),
                ("tailored_recruiter",  "Tailored Recruiter"),
                ("recruiter_delta",     "Recruiter Delta"),
                ("original_hm",         "Original HM"),
                ("tailored_hm",         "Tailored HM"),
                ("hm_delta",            "HM Delta"),
            )
            for rec_key, col_name in _dim_keys:
                if rec_key in rec and _dim_cols[col_name]:
                    ws.cell(target_row, _dim_cols[col_name], rec[rec_key])
            # sha256 of the .md content the optimizer just wrote. Same
            # "key present → write" semantics as per-dim cols above.
            if "last_written_hash" in rec and _dim_cols["Last Written Hash"]:
                ws.cell(target_row, _dim_cols["Last Written Hash"],
                        rec["last_written_hash"])
        wb.save(xlsx_path)
        return len(records)
    finally:
        wb.close()
